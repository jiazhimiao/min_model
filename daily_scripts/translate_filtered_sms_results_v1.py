"""
优化版翻译脚本 V1

优化点：
1. 批量翻译 - 每次请求翻译多条短信，减少API调用次数
2. 并行处理 - 多线程并发翻译，大幅提升速度
3. 进度显示 - 实时显示翻译进度和预估剩余时间
"""

from __future__ import annotations

import argparse
import json
import os
import re
import shutil
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from pathlib import Path
from threading import Lock

from openai import OpenAI
from openpyxl import load_workbook

API_KEY = os.getenv("DASHSCOPE_API_KEY", "sk-2f6faaacbae74df08ec9f616c5aad97c")

MODEL_LIMITS = {
    "qwen-mt-flash": {
        "rpm": 60,
        "tpm": 35000,
        "default_request_delay": 1.05,
        "target_batch_tokens": 520,
        "default_batch_size": 12,
    },
    "qwen-mt-turbo": {
        "rpm": 60,
        "tpm": 25000,
        "default_request_delay": 1.05,
        "target_batch_tokens": 380,
        "default_batch_size": 8,
    },
}


@dataclass
class TranslationStats:
    total: int = 0
    cached: int = 0
    translated: int = 0
    failed: int = 0
    start_time: float = 0.0

    def elapsed(self) -> float:
        return time.time() - self.start_time

    def speed(self) -> float:
        if self.elapsed() > 0:
            return (self.cached + self.translated) / self.elapsed()
        return 0.0

    def eta(self, remaining: int) -> float:
        if self.speed() > 0:
            return remaining / self.speed()
        return 0.0


def estimate_text_tokens(text: str) -> int:
    text = str(text or "").strip()
    if not text:
        return 0
    return max(1, (len(text) + 3) // 4)


def strip_code_fences(text: str) -> str:
    text = text.strip()
    if text.startswith("```") and text.endswith("```"):
        lines = text.splitlines()
        if len(lines) >= 3:
            return "\n".join(lines[1:-1]).strip()
    return text


class DashScopeTranslateClientV1:
    def __init__(
        self,
        cache_path: str | Path = "translation_cache_zh.json",
        model: str = "qwen-mt-flash",
        max_retries: int = 5,
        batch_size: int = 12,
        max_workers: int = 1,
        retry_backoff_base: float = 3.0,
        save_every: int = 100,
        request_delay: float = 1.05,
    ) -> None:
        model_limits = MODEL_LIMITS.get(model, {})
        self.cache_path = Path(cache_path)
        self.model = model
        self.max_retries = max_retries
        self.batch_size = batch_size or model_limits.get("default_batch_size", 10)
        self.max_workers = max_workers or 1
        self.retry_backoff_base = retry_backoff_base
        self.save_every = save_every
        self.request_delay = request_delay or model_limits.get("default_request_delay", 1.0)
        self.target_batch_tokens = model_limits.get("target_batch_tokens", 450)

        api_key = API_KEY
        if not api_key:
            raise EnvironmentError("请设置环境变量 DASHSCOPE_API_KEY")

        self.client = OpenAI(
            api_key=api_key,
            base_url="https://dashscope.aliyuncs.com/compatible-mode/v1",
        )
        self.cache: dict[str, str] = {}
        self.cache_lock = Lock()
        self._load_cache()

    def _load_cache(self) -> None:
        if not self.cache_path.exists():
            return
        try:
            self.cache = json.loads(self.cache_path.read_text(encoding="utf-8"))
        except Exception:
            broken_path = self.cache_path.with_suffix(self.cache_path.suffix + ".broken")
            shutil.copyfile(self.cache_path, broken_path)
            print(f"缓存文件损坏，已备份到: {broken_path}")
            self.cache = {}

    def save_cache(self) -> None:
        with self.cache_lock:
            tmp_path = self.cache_path.with_suffix(self.cache_path.suffix + ".tmp")
            tmp_path.write_text(
                json.dumps(self.cache, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
            tmp_path.replace(self.cache_path)

    def translate_single(self, text: str) -> str:
        text = str(text or "").strip()
        if not text:
            return ""

        with self.cache_lock:
            if text in self.cache:
                return self.cache[text]

        prompt = (
            "请把下面的西班牙语短信翻译成自然、准确、简洁的中文。"
            "不要解释，不要总结，只输出中文译文。\n\n"
            f"{text}"
        )

        for attempt in range(self.max_retries):
            try:
                time.sleep(self.request_delay)
                completion = self.client.chat.completions.create(
                    model=self.model,
                    messages=[{"role": "user", "content": prompt}],
                )
                translated = completion.choices[0].message.content.strip()
                with self.cache_lock:
                    self.cache[text] = translated
                return translated
            except Exception as exc:
                if attempt < self.max_retries - 1:
                    wait = self.retry_backoff_base * (attempt + 1)
                    time.sleep(wait)
                else:
                    print(f"翻译失败: {text[:50]}... 错误: {exc}")
                    return ""

        return ""

    def translate_batch(self, texts: list[str]) -> dict[str, str]:
        if not texts:
            return {}

        results = {}
        uncached = []

        with self.cache_lock:
            for text in texts:
                if text in self.cache:
                    results[text] = self.cache[text]
                else:
                    uncached.append(text)

        if not uncached:
            return results

        combined_text = "\n---SEPARATOR---\n".join(uncached)
        prompt = (
            "请把下面的西班牙语短信翻译成中文。"
            "每条短信用 ---SEPARATOR--- 分隔，请按相同顺序输出翻译结果，"
            "每条翻译也用 ---SEPARATOR--- 分隔。\n\n"
            f"{combined_text}"
        )

        for attempt in range(self.max_retries):
            try:
                time.sleep(self.request_delay)
                completion = self.client.chat.completions.create(
                    model=self.model,
                    messages=[{"role": "user", "content": prompt}],
                )
                response = completion.choices[0].message.content.strip()
                translations = response.split("---SEPARATOR---")

                translations = [t.strip() for t in translations if t.strip()]

                if len(translations) >= len(uncached):
                    for i, text in enumerate(uncached):
                        translated = translations[i] if i < len(translations) else ""
                        results[text] = translated
                        with self.cache_lock:
                            self.cache[text] = translated
                    return results
                else:
                    raise ValueError(f"翻译数量不匹配: 期望 {len(uncached)}, 实际 {len(translations)}")

            except Exception as exc:
                if attempt < self.max_retries - 1:
                    wait = self.retry_backoff_base * (attempt + 1)
                    time.sleep(wait)
                else:
                    print(f"批量翻译失败，回退到单条翻译: {exc}")
                    for text in uncached:
                        results[text] = self.translate_single(text)
                    return results

        return results

    def translate_many_parallel(
        self,
        texts: list[str],
        max_items: int | None = None,
        burst_size: int = 10,
        burst_interval: float = 10.0,
    ) -> dict[str, str]:
        unique_texts = list({str(t or "").strip() for t in texts if str(t or "").strip()})

        if max_items is not None:
            unique_texts = unique_texts[:max_items]

        stats = TranslationStats()
        stats.total = len(unique_texts)
        stats.start_time = time.time()

        results: dict[str, str] = {}
        to_translate: list[str] = []

        with self.cache_lock:
            for text in unique_texts:
                if text in self.cache:
                    results[text] = self.cache[text]
                    stats.cached += 1
                else:
                    to_translate.append(text)

        print(f"总计 {stats.total} 条，已缓存 {stats.cached} 条，需翻译 {len(to_translate)} 条")
        print(f"突发模式: 每 {burst_interval}秒 发送 {burst_size} 个请求，每请求翻译 {self.batch_size} 条")
        print(f"理论速度: 每{burst_interval}秒翻译 {burst_size * self.batch_size} 条")

        if not to_translate:
            return results

        batches = [
            to_translate[i:i + self.batch_size]
            for i in range(0, len(to_translate), self.batch_size)
        ]

        print(f"共 {len(batches)} 个批次")

        dirty_count = 0
        last_save_time = time.time()
        burst_count = 0

        for burst_start in range(0, len(batches), burst_size):
            burst_count += 1
            burst_batches = batches[burst_start:burst_start + burst_size]
            
            if burst_start > 0:
                print(f"等待 {burst_interval} 秒后发送下一批...")
                time.sleep(burst_interval)

            with ThreadPoolExecutor(max_workers=burst_size) as executor:
                future_to_batch = {
                    executor.submit(self.translate_batch, batch): batch
                    for batch in burst_batches
                }

                for future in as_completed(future_to_batch):
                    batch = future_to_batch[future]
                    try:
                        batch_results = future.result()
                        results.update(batch_results)
                        stats.translated += len(batch_results)
                        dirty_count += len(batch_results)

                        if dirty_count >= self.save_every or time.time() - last_save_time > 30:
                            self.save_cache()
                            dirty_count = 0
                            last_save_time = time.time()

                        remaining = len(to_translate) - stats.translated - stats.failed
                        eta_str = f", 预计剩余 {stats.eta(remaining) / 60:.1f} 分钟" if stats.speed() > 0 else ""
                        print(
                            f"[突发{burst_count}] 进度: {stats.translated + stats.cached}/{stats.total} "
                            f"(速度: {stats.speed():.1f} 条/秒{eta_str})"
                        )

                    except Exception as exc:
                        stats.failed += len(batch)
                        print(f"批次翻译失败: {exc}")

        self.save_cache()

        elapsed = stats.elapsed()
        print(f"\n翻译完成!")
        print(f"  - 缓存命中: {stats.cached} 条")
        print(f"  - 新翻译: {stats.translated} 条")
        print(f"  - 失败: {stats.failed} 条")
        print(f"  - 耗时: {elapsed:.1f} 秒 ({elapsed / 60:.1f} 分钟)")
        print(f"  - 平均速度: {stats.speed():.1f} 条/秒")

        return results


def translate_result_workbook(
    input_file: str | Path,
    output_file: str | Path,
    translation_cache: str | Path = "translation_cache_zh.json",
    model: str = "qwen-mt-flash",
    batch_size: int = 70,
    burst_size: int = 10,
    burst_interval: float = 10.0,
    max_items: int | None = None,
) -> None:
    print(f"读取文件: {input_file}")
    wb = load_workbook(input_file)
    ws = wb[wb.sheetnames[0]]

    headers = [cell.value for cell in ws[1]]
    header_map = {str(v).strip(): idx + 1 for idx, v in enumerate(headers) if v is not None}

    if "content" not in header_map:
        raise ValueError("缺少 content 列")

    content_col = header_map["content"]

    if "content_zh" in header_map:
        content_zh_col = header_map["content_zh"]
    else:
        content_zh_col = ws.max_column + 1
        ws.cell(row=1, column=content_zh_col, value="content_zh")

    texts: list[str] = []
    for row_idx in range(2, ws.max_row + 1):
        text = str(ws.cell(row=row_idx, column=content_col).value or "").strip()
        if text:
            texts.append(text)

    print(f"共 {len(texts)} 条短信待处理")

    translator = DashScopeTranslateClientV1(
        cache_path=translation_cache,
        model=model,
        batch_size=batch_size,
        max_workers=burst_size,
    )

    translations = translator.translate_many_parallel(
        texts, 
        max_items=max_items,
        burst_size=burst_size,
        burst_interval=burst_interval,
    )

    print("回填翻译结果...")
    translated_keys = set(translations.keys())
    for row_idx in range(2, ws.max_row + 1):
        text = str(ws.cell(row=row_idx, column=content_col).value or "").strip()
        if text in translated_keys:
            ws.cell(row=row_idx, column=content_zh_col, value=translations.get(text, ""))
        if row_idx % 50000 == 0:
            print(f"  已回填 {row_idx - 1}/{ws.max_row - 1} 行")

    for column_cells in ws.columns:
        header = str(column_cells[0].value or "")
        if header in {"content", "normalized_text", "content_zh"}:
            ws.column_dimensions[column_cells[0].column_letter].width = 90

    wb.save(output_file)
    print(f"输出完成: {output_file}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="优化版翻译脚本 - 支持突发模式批量翻译"
    )
    parser.add_argument("input_file")
    parser.add_argument("output_file")
    parser.add_argument(
        "--translation-cache",
        default="translation_cache_zh.json",
        help="翻译缓存文件路径",
    )
    parser.add_argument(
        "--model",
        default="qwen-mt-flash",
        help="DashScope 模型名",
    )
    parser.add_argument(
        "--batch-size",
        type=int,
        default=70,
        help="每个请求翻译的短信数量 (默认: 70)",
    )
    parser.add_argument(
        "--burst-size",
        type=int,
        default=10,
        help="每次突发发送的请求数 (默认: 10)",
    )
    parser.add_argument(
        "--burst-interval",
        type=float,
        default=10.0,
        help="突发间隔秒数 (默认: 10.0)",
    )
    parser.add_argument(
        "--max-items",
        type=int,
        default=None,
        help="仅翻译前 N 条（测试用）",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    translate_result_workbook(
        input_file=args.input_file,
        output_file=args.output_file,
        translation_cache=args.translation_cache,
        model=args.model,
        batch_size=args.batch_size,
        burst_size=args.burst_size,
        burst_interval=args.burst_interval,
        max_items=args.max_items,
    )


if __name__ == "__main__":
    main()
