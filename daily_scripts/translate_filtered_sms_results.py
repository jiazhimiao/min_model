from __future__ import annotations

import argparse
import json
import os
import shutil
import time
from pathlib import Path

from openai import OpenAI
from openpyxl import load_workbook

HARDCODED_DASHSCOPE_API_KEY = "sk-2f6faaacbae74df08ec9f616c5aad97c"


class DashScopeTranslateClient:
    def __init__(
        self,
        cache_path: str | Path = "translation_cache_zh.json",
        model: str = "qwen-mt-flash",
        max_retries: int = 5,
        sleep_seconds: float = 0.2,
        retry_backoff_base: float = 5.0,
        save_every: int = 50,
    ) -> None:
        self.cache_path = Path(cache_path)
        self.model = model
        self.max_retries = max_retries
        self.sleep_seconds = sleep_seconds
        self.retry_backoff_base = retry_backoff_base
        self.save_every = save_every
        api_key = HARDCODED_DASHSCOPE_API_KEY or os.getenv("DASHSCOPE_API_KEY")
        if not api_key:
            raise EnvironmentError("缺少 DashScope API key，请设置 HARDCODED_DASHSCOPE_API_KEY 或 DASHSCOPE_API_KEY。")
        self.client = OpenAI(
            api_key=api_key,
            base_url="https://dashscope.aliyuncs.com/compatible-mode/v1",
        )
        self.cache: dict[str, str] = {}
        self._load_cache()

    def _load_cache(self) -> None:
        if not self.cache_path.exists():
            return
        try:
            self.cache = json.loads(self.cache_path.read_text(encoding="utf-8"))
        except Exception:
            broken_path = self.cache_path.with_suffix(self.cache_path.suffix + ".broken")
            shutil.copyfile(self.cache_path, broken_path)
            print(f"检测到损坏缓存，已备份到: {broken_path}")
            self.cache = {}

    def save_cache(self) -> None:
        tmp_path = self.cache_path.with_suffix(self.cache_path.suffix + ".tmp")
        tmp_path.write_text(
            json.dumps(self.cache, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        tmp_path.replace(self.cache_path)

    def translate_es_to_zh(self, text: str) -> str:
        text = str(text or "").strip()
        if not text:
            return ""
        if text in self.cache:
            return self.cache[text]

        prompt = (
            "请把下面的西班牙语短信翻译成自然、准确、简洁的中文。"
            "不要解释，不要总结，只输出中文译文。\n\n"
            f"{text}"
        )

        last_error: Exception | None = None
        for i in range(self.max_retries):
            try:
                completion = self.client.chat.completions.create(
                    model=self.model,
                    messages=[{"role": "user", "content": prompt}],
                    extra_body={
                        "translation_options": {
                            "source_lang": "Spanish",
                            "target_lang": "Chinese",
                        }
                    },
                )
                translated = completion.choices[0].message.content.strip()
                self.cache[text] = translated
                time.sleep(self.sleep_seconds)
                return translated
            except Exception as exc:
                last_error = exc
                wait_seconds = self.retry_backoff_base * (i + 1)
                print(f"翻译失败，准备重试 {i + 1}/{self.max_retries}，等待 {wait_seconds:.1f}s。错误: {type(exc).__name__}")
                time.sleep(wait_seconds)

        if last_error is not None:
            raise last_error
        return ""

    def translate_many(self, texts: list[str], max_items: int | None = None) -> dict[str, str]:
        unique_texts: list[str] = []
        seen: set[str] = set()
        for text in texts:
            key = str(text or "").strip()
            if key and key not in seen:
                unique_texts.append(key)
                seen.add(key)

        if max_items is not None:
            unique_texts = unique_texts[:max_items]

        print(f"准备翻译 {len(unique_texts)} 条去重后的短信内容...")
        newly_translated = 0
        dirty_since_save = 0

        for idx, text in enumerate(unique_texts, start=1):
            if text not in self.cache:
                self.translate_es_to_zh(text)
                newly_translated += 1
                dirty_since_save += 1
                if dirty_since_save >= self.save_every:
                    self.save_cache()
                    dirty_since_save = 0
            if idx % 50 == 0:
                print(f"  已处理 {idx}/{len(unique_texts)} 条，新增翻译 {newly_translated} 条")

        self.save_cache()
        print(f"翻译完成，本次新增翻译 {newly_translated} 条。")
        return {text: self.cache.get(text, "") for text in unique_texts}


def translate_result_workbook(
    input_file: str | Path,
    output_file: str | Path,
    translation_cache: str | Path = "translation_cache_zh.json",
    model: str = "qwen-mt-flash",
    max_items: int | None = None,
) -> None:
    print(f"读取结果文件: {input_file}")
    wb = load_workbook(input_file)
    ws = wb[wb.sheetnames[0]]

    headers = [cell.value for cell in ws[1]]
    header_map = {str(v).strip(): idx + 1 for idx, v in enumerate(headers) if v is not None}

    if "content" not in header_map:
        raise ValueError("结果文件缺少 content 列，无法翻译。")

    content_col = header_map["content"]

    if "content_zh" in header_map:
        content_zh_col = header_map["content_zh"]
    else:
        content_zh_col = ws.max_column + 1
        ws.cell(row=1, column=content_zh_col, value="content_zh")

    texts: list[str] = []
    for row_idx in range(2, ws.max_row + 1):
        text = str(ws.cell(row=row_idx, column=content_col).value or "").strip()
        if text:
            texts.append(text)

    translator = DashScopeTranslateClient(
        cache_path=translation_cache,
        model=model,
    )
    translations = translator.translate_many(texts, max_items=max_items)

    print("开始回填中文翻译...")
    translated_keys = set(translations.keys())
    for row_idx in range(2, ws.max_row + 1):
        text = str(ws.cell(row=row_idx, column=content_col).value or "").strip()
        if text in translated_keys:
            ws.cell(row=row_idx, column=content_zh_col, value=translations.get(text, ""))
        if row_idx % 50000 == 0:
            print(f"  已回填 {row_idx - 1}/{ws.max_row - 1} 行")

    for column_cells in ws.columns:
        header = str(column_cells[0].value or "")
        if header in {"content", "normalized_text", "content_zh"}:
            ws.column_dimensions[column_cells[0].column_letter].width = 90

    wb.save(output_file)
    print(f"输出完成: {output_file}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="为已有筛选结果文件补充 content_zh 翻译列。")
    parser.add_argument("input_file")
    parser.add_argument("output_file")
    parser.add_argument(
        "--translation-cache",
        default="translation_cache_zh.json",
        help="翻译缓存文件路径，避免重复翻译相同短信。",
    )
    parser.add_argument(
        "--model",
        default="qwen-mt-flash",
        help="DashScope 兼容接口模型名。",
    )
    parser.add_argument(
        "--max-items",
        type=int,
        default=None,
        help="仅翻译前 N 条去重短信，用于测试或分批跑。",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    translate_result_workbook(
        input_file=args.input_file,
        output_file=args.output_file,
        translation_cache=args.translation_cache,
        model=args.model,
        max_items=args.max_items,
    )


if __name__ == "__main__":
    main()
