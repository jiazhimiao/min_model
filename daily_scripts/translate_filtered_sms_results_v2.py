"""
优化版翻译脚本 V2

优化点：
1. 批量翻译，减少 API 调用次数
2. 三线程并行处理，尽量利用请求等待时间
3. 全局请求限流，控制整体请求速率，避免冲击 RPM
4. 基于缓存复用历史翻译结果，避免重复翻译
5. 按估算 token 动态拆批，尽量贴近 TPM 上限
"""

from __future__ import annotations

import argparse
import json
import os
import re
import shutil
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from pathlib import Path
from threading import Lock

from openai import OpenAI
from openpyxl import load_workbook

API_KEY = os.getenv("DASHSCOPE_API_KEY", "sk-2f6faaacbae74df08ec9f616c5aad97c")

# 不同模型的默认限流与批处理参数。
MODEL_LIMITS = {
    "qwen-mt-flash": {
        "rpm": 60,
        "tpm": 35000,
        "default_request_delay": 1.02,
        "target_batch_tokens": 520,
        "default_batch_size": 12,
        "default_max_workers": 3,
    },
    "qwen-mt-turbo": {
        "rpm": 60,
        "tpm": 25000,
        "default_request_delay": 1.02,
        "target_batch_tokens": 380,
        "default_batch_size": 8,
        "default_max_workers": 3,
    },
}


@dataclass
class TranslationStats:
    """记录翻译过程中的进度、速度与耗时。"""

    total: int = 0
    cached: int = 0
    translated: int = 0
    failed: int = 0
    start_time: float = 0.0

    def elapsed(self) -> float:
        return time.time() - self.start_time

    def speed(self) -> float:
        elapsed = self.elapsed()
        if elapsed <= 0:
            return 0.0
        return (self.cached + self.translated) / elapsed

    def eta(self, remaining: int) -> float:
        speed = self.speed()
        if speed <= 0:
            return 0.0
        return remaining / speed


class RequestRateLimiter:
    """全局请求限流器，确保多线程共享同一个发请求节奏。"""

    def __init__(self, min_interval_seconds: float) -> None:
        self.min_interval_seconds = max(0.0, min_interval_seconds)
        self._lock = Lock()
        self._next_allowed_time = 0.0

    def acquire(self) -> None:
        while True:
            with self._lock:
                now = time.monotonic()
                if now >= self._next_allowed_time:
                    self._next_allowed_time = now + self.min_interval_seconds
                    return
                wait_seconds = self._next_allowed_time - now
            if wait_seconds > 0:
                time.sleep(wait_seconds)


def estimate_text_tokens(text: str) -> int:
    """粗略估算文本 token 数，用于控制批次大小。"""

    text = str(text or "").strip()
    if not text:
        return 0
    return max(1, (len(text) + 3) // 4)


def strip_code_fences(text: str) -> str:
    """去掉模型偶尔返回的 Markdown 代码块外壳。"""

    text = text.strip()
    if text.startswith("```") and text.endswith("```"):
        lines = text.splitlines()
        if len(lines) >= 3:
            return "\n".join(lines[1:-1]).strip()
    return text


def extract_tagged_translations(response: str) -> dict[int, str]:
    """从带 [ZH-001] 这类标签的响应中提取翻译结果。"""

    result: dict[int, str] = {}
    for line in strip_code_fences(response).splitlines():
        matched = re.match(r"\[ZH-(\d{3})\]\s*(.*)", line.strip())
        if matched:
            result[int(matched.group(1))] = matched.group(2).strip()
    return result


class DashScopeTranslateClientV2:
    """DashScope 翻译客户端，支持缓存、批量、并行与全局限流。"""

    def __init__(
        self,
        cache_path: str | Path = "translation_cache_zh.json",
        model: str = "qwen-mt-flash",
        max_retries: int = 5,
        batch_size: int | None = None,
        max_workers: int | None = None,
        retry_backoff_base: float = 3.0,
        save_every: int = 100,
        request_delay: float | None = None,
    ) -> None:
        model_limits = MODEL_LIMITS.get(model, {})
        self.cache_path = Path(cache_path)
        self.model = model
        self.max_retries = max_retries
        self.batch_size = batch_size or model_limits.get("default_batch_size", 10)
        self.max_workers = max_workers or model_limits.get("default_max_workers", 1)
        self.retry_backoff_base = retry_backoff_base
        self.save_every = save_every
        self.request_delay = request_delay or model_limits.get("default_request_delay", 1.0)
        self.target_batch_tokens = model_limits.get("target_batch_tokens", 450)
        self.rate_limiter = RequestRateLimiter(self.request_delay)

        api_key = API_KEY
        if not api_key:
            raise EnvironmentError("请设置环境变量 DASHSCOPE_API_KEY")

        self.client = OpenAI(
            api_key=api_key,
            base_url="https://dashscope.aliyuncs.com/compatible-mode/v1",
        )
        self.cache: dict[str, str] = {}
        self.cache_lock = Lock()
        self._load_cache()

    def _load_cache(self) -> None:
        if not self.cache_path.exists():
            return
        try:
            self.cache = json.loads(self.cache_path.read_text(encoding="utf-8"))
        except Exception:
            broken_path = self.cache_path.with_suffix(self.cache_path.suffix + ".broken")
            shutil.copyfile(self.cache_path, broken_path)
            print(f"检测到缓存文件损坏，已备份到: {broken_path}")
            self.cache = {}

    def save_cache(self) -> None:
        with self.cache_lock:
            tmp_path = self.cache_path.with_suffix(self.cache_path.suffix + ".tmp")
            tmp_path.write_text(
                json.dumps(self.cache, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
            tmp_path.replace(self.cache_path)

    def translate_single(self, text: str) -> str:
        """单条翻译，作为批量失败时的兜底路径。"""

        text = str(text or "").strip()
        if not text:
            return ""

        with self.cache_lock:
            cached = self.cache.get(text)
        if cached is not None:
            return cached

        prompt = (
            "Translate the following Spanish SMS into concise natural Chinese. "
            "Return only the Chinese translation.\n\n"
            f"{text}"
        )

        for attempt in range(self.max_retries):
            try:
                self.rate_limiter.acquire()
                completion = self.client.chat.completions.create(
                    model=self.model,
                    messages=[{"role": "user", "content": prompt}],
                )
                translated = (completion.choices[0].message.content or "").strip()
                with self.cache_lock:
                    self.cache[text] = translated
                return translated
            except Exception as exc:
                if attempt < self.max_retries - 1:
                    time.sleep(self.retry_backoff_base * (attempt + 1))
                else:
                    print(f"单条翻译失败: {exc}")
                    return ""
        return ""

    def translate_batch(self, texts: list[str]) -> dict[str, str]:
        """批量翻译一组短信，优先使用缓存，失败时回退到单条翻译。"""

        if not texts:
            return {}

        results: dict[str, str] = {}
        uncached: list[str] = []

        with self.cache_lock:
            for text in texts:
                cached = self.cache.get(text)
                if cached is None:
                    uncached.append(text)
                else:
                    results[text] = cached

        if not uncached:
            return results

        payload_lines = [f"[SMS-{idx:03d}] {text}" for idx, text in enumerate(uncached, start=1)]
        prompt = (
            "Translate each Spanish SMS to concise natural Chinese.\n"
            "Keep the same order and return one line per item.\n"
            "Use this exact format:\n"
            "[ZH-001] translation\n"
            "[ZH-002] translation\n\n"
            f"{chr(10).join(payload_lines)}"
        )

        for attempt in range(self.max_retries):
            try:
                self.rate_limiter.acquire()
                completion = self.client.chat.completions.create(
                    model=self.model,
                    messages=[{"role": "user", "content": prompt}],
                )
                parsed = extract_tagged_translations(completion.choices[0].message.content or "")
                if len(parsed) < len(uncached):
                    raise ValueError(f"期望 {len(uncached)} 条翻译，实际只返回 {len(parsed)} 条")

                for idx, text in enumerate(uncached, start=1):
                    translated = parsed.get(idx, "")
                    results[text] = translated
                    with self.cache_lock:
                        self.cache[text] = translated
                return results
            except Exception as exc:
                if attempt < self.max_retries - 1:
                    time.sleep(self.retry_backoff_base * (attempt + 1))
                else:
                    print(f"批量翻译失败，回退到单条模式: {exc}")
                    for text in uncached:
                        results[text] = self.translate_single(text)
                    return results

        return results

    def build_batches(self, texts: list[str]) -> list[list[str]]:
        """按最大条数和估算 token 数量动态拆分批次。"""

        batches: list[list[str]] = []
        current_batch: list[str] = []
        current_tokens = 0

        for text in texts:
            estimated_tokens = estimate_text_tokens(text) * 2 + 12
            hit_batch_limit = len(current_batch) >= self.batch_size
            hit_token_limit = current_batch and (current_tokens + estimated_tokens > self.target_batch_tokens)

            if hit_batch_limit or hit_token_limit:
                batches.append(current_batch)
                current_batch = []
                current_tokens = 0

            current_batch.append(text)
            current_tokens += estimated_tokens

        if current_batch:
            batches.append(current_batch)

        return batches

    def translate_many(self, texts: list[str], max_items: int | None = None) -> dict[str, str]:
        """批量翻译整批短信，含去重、缓存命中、并发处理与进度显示。"""

        unique_texts: list[str] = []
        seen: set[str] = set()
        for text in texts:
            key = str(text or "").strip()
            if key and key not in seen:
                unique_texts.append(key)
                seen.add(key)

        if max_items is not None:
            unique_texts = unique_texts[:max_items]

        stats = TranslationStats(total=len(unique_texts), start_time=time.time())
        results: dict[str, str] = {}
        to_translate: list[str] = []

        with self.cache_lock:
            for text in unique_texts:
                cached = self.cache.get(text)
                if cached is None:
                    to_translate.append(text)
                else:
                    results[text] = cached
                    stats.cached += 1

        print(f"总计 {stats.total} 条，缓存命中 {stats.cached} 条，待翻译 {len(to_translate)} 条")
        print(
            f"模型={self.model}，线程数={self.max_workers}，"
            f"单批最多 {self.batch_size} 条，请求间隔 {self.request_delay:.2f} 秒，"
            f"目标批次 token 约 {self.target_batch_tokens}"
        )

        if not to_translate:
            return results

        batches = self.build_batches(to_translate)
        print(f"共拆分为 {len(batches)} 个批次")

        dirty_count = 0
        last_save_time = time.time()

        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            future_to_batch = {executor.submit(self.translate_batch, batch): batch for batch in batches}
            for future in as_completed(future_to_batch):
                batch = future_to_batch[future]
                try:
                    batch_results = future.result()
                    results.update(batch_results)
                    stats.translated += len(batch_results)
                    dirty_count += len(batch_results)

                    if dirty_count >= self.save_every or time.time() - last_save_time > 30:
                        self.save_cache()
                        dirty_count = 0
                        last_save_time = time.time()

                    remaining = len(to_translate) - stats.translated - stats.failed
                    print(
                        f"进度: {stats.translated + stats.cached}/{stats.total}，"
                        f"速度: {stats.speed():.1f} 条/秒，"
                        f"预计剩余: {stats.eta(remaining) / 60:.1f} 分钟"
                    )
                except Exception as exc:
                    stats.failed += len(batch)
                    print(f"批次处理失败: {exc}")

        self.save_cache()
        print(
            f"翻译完成，缓存命中 {stats.cached} 条，新翻译 {stats.translated} 条，"
            f"失败 {stats.failed} 条，总耗时 {stats.elapsed():.1f} 秒"
        )
        return results


def translate_result_workbook(
    input_file: str | Path,
    output_file: str | Path,
    translation_cache: str | Path = "translation_cache_zh.json",
    model: str = "qwen-mt-flash",
    batch_size: int | None = None,
    max_workers: int | None = None,
    request_delay: float | None = None,
    max_items: int | None = None,
) -> None:
    """读取 Excel，翻译 content 列，并回填到 content_zh 列。"""

    print(f"读取工作簿: {input_file}")
    wb = load_workbook(input_file)
    ws = wb[wb.sheetnames[0]]

    headers = [cell.value for cell in ws[1]]
    header_map = {str(v).strip(): idx + 1 for idx, v in enumerate(headers) if v is not None}

    if "content" not in header_map:
        raise ValueError("缺少 content 列")

    content_col = header_map["content"]

    if "content_zh" in header_map:
        content_zh_col = header_map["content_zh"]
    else:
        content_zh_col = ws.max_column + 1
        ws.cell(row=1, column=content_zh_col, value="content_zh")

    texts: list[str] = []
    for row_idx in range(2, ws.max_row + 1):
        text = str(ws.cell(row=row_idx, column=content_col).value or "").strip()
        if text:
            texts.append(text)

    print(f"检测到 {len(texts)} 条待处理短信")

    translator = DashScopeTranslateClientV2(
        cache_path=translation_cache,
        model=model,
        batch_size=batch_size,
        max_workers=max_workers,
        request_delay=request_delay,
    )
    translations = translator.translate_many(texts, max_items=max_items)

    print("开始回填翻译结果...")
    translated_keys = set(translations.keys())
    for row_idx in range(2, ws.max_row + 1):
        text = str(ws.cell(row=row_idx, column=content_col).value or "").strip()
        if text in translated_keys:
            ws.cell(row=row_idx, column=content_zh_col, value=translations.get(text, ""))
        if row_idx % 50000 == 0:
            print(f"已回填 {row_idx - 1}/{ws.max_row - 1} 行")

    for column_cells in ws.columns:
        header = str(column_cells[0].value or "")
        if header in {"content", "normalized_text", "content_zh"}:
            ws.column_dimensions[column_cells[0].column_letter].width = 90

    wb.save(output_file)
    print(f"输出完成: {output_file}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="为结果工作簿补充 content_zh 列，支持缓存、批量翻译、并行处理与全局限流。"
    )
    parser.add_argument("input_file")
    parser.add_argument("output_file")
    parser.add_argument("--translation-cache", default="translation_cache_zh.json", help="翻译缓存文件路径")
    parser.add_argument("--model", default="qwen-mt-flash", help="DashScope 模型名")
    parser.add_argument("--batch-size", type=int, default=12, help="单次请求最多翻译的短信数")
    parser.add_argument("--max-workers", type=int, default=3, help="并行线程数")
    parser.add_argument("--request-delay", type=float, default=1.02, help="全局请求最小间隔秒数")
    parser.add_argument("--max-items", type=int, default=None, help="仅翻译前 N 条去重短信，用于测试")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    translate_result_workbook(
        input_file=args.input_file,
        output_file=args.output_file,
        translation_cache=args.translation_cache,
        model=args.model,
        batch_size=args.batch_size,
        max_workers=args.max_workers,
        request_delay=args.request_delay,
        max_items=args.max_items,
    )


if __name__ == "__main__":
    main()
