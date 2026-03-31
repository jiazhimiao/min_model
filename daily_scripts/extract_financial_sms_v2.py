from __future__ import annotations

import argparse
import json
import re
import sys
import time
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import requests
from openpyxl import Workbook, load_workbook


# =============================================================================
# 整体设计说明
# =============================================================================
# 这个脚本用于处理 Excel 的全部 sheet，对每条短信做“关键词命中 + 规则修正”
# 的金融识别，并输出筛选后的结果文件。
#
# v2 的重点在于：
# 1. 不只给出是否像现金贷，还补充更细的 finance_segment
# 2. 输出时可以保留原始字段，也可以附带中文翻译
# 3. 翻译支持本地 JSON 缓存，避免重复请求
#
# 整体主流程：
# 读取全部 sheet -> 提取短信内容 -> 归一化与分词 -> 打分与分类 -> 可选翻译 -> 输出
#

# =============================================================================
# 关键词分组说明
# =============================================================================
# 当前关键词分为 5 类：
#
# 1) strong
#    高置信关键词。
#    单独出现时就很像贷款申请、放款、还款、逾期、催收、法务升级语境。
#    这是整个规则体系里最主要的加分来源。
#
# 2) medium
#    中等信号词。
#    这类词在贷款短信里很常见，但在一般金融通知、账户提醒、支付类短信里
#    也会大量出现，所以只能作为辅助上下文，权重不能太高。
#
# 3) low_weight_noise
#    噪音词 / 泛金融词。
#    常见于银行登录提醒、验证码、账单、汇款、电信、保险等场景。
#    它们不能证明短信属于现金贷，所以当前策略是轻微扣分，而不是直接忽略。
#
# 4) ignore
#    高度常见但几乎没有区分度的停用词。
#    会在分词后直接剔除，不参与任何评分。
#
# 5) brand_hints
#    历史样本里出现过的平台名或产品名提示词。
#    单独命中不一定能直接定性，但和借贷词、审批词、账户词一起出现时，
#    往往能明显增强“贷款平台短信”的可信度。
# =============================================================================

SMS_LOAN_KEYWORDS = {
    "strong": {
        # ---------------------------------------------------------------------
        # A. 借贷核心词
        # ---------------------------------------------------------------------
        # 直接表达“借款 / 贷款 / 信用贷 / 小额贷”的概念。
        # 这是最基础也最关键的一组词。
        #
        # 常见短信场景：
        # - tu prestamo esta disponible
        # - credito aprobado
        # - microcredito al instante
        "prestamo", "prestamos", "préstamo", "préstamos",
        "credito", "creditos", "crédito", "créditos",
        "microcredito", "microcreditos", "microcrédito", "microcréditos",
        "prestamista", "prestamistas",

        # ---------------------------------------------------------------------
        # B. 贷前阶段：申请 / 审批 / 授信 / 额度
        # ---------------------------------------------------------------------
        # 这组词常见于“申请成功”“审批通过”“给额度”“额度提升”等贷前营销语境。
        # 如果它们和 brand_hints、credito、cuenta 一起出现，通常很像平台拉新短信。
        "solicitud", "solicitudes",
        "aprobado", "aprobada", "aprobados", "aprobadas",
        "aprobacion", "aprobaciones", "aprobación",
        "rechazado", "rechazada", "rechazo",
        "limite", "limites", "límite", "límites",
        "monto", "montos",
        "capacidad",
        "financiamiento",

        # ---------------------------------------------------------------------
        # C. 放款 / 到账 / 转账
        # ---------------------------------------------------------------------
        # 用于识别“已放款 / 已到账 / 已转入账户”的场景。
        # 在贷款类短信里，这组词通常意味着用户已进入放款后阶段。
        "depositado", "depositada", "depositados", "depositadas",
        "deposito", "depositos", "depósito", "depósitos",
        "desembolso", "desembolsos",
        "transferencia", "transferencias",

        # ---------------------------------------------------------------------
        # D. 贷后还款 / 上传凭证 / 分期
        # ---------------------------------------------------------------------
        # 这组词主要出现在贷后管理和还款提醒里。
        # 例如上传凭证、支付分期、结清、取消中的上下文。
        "comprobante", "comprobantes",
        "recibo", "recibos",
        "cuota", "cuotas",
        "abono", "abonos",
        "liquidado", "liquidada", "liquidados", "liquidadas",
        "liquidacion", "liquidación",
        "cancelar", "cancele", "cancelacion", "cancelación",

        # ---------------------------------------------------------------------
        # E. 到期 / 逾期 / 拖欠
        # ---------------------------------------------------------------------
        # 这是催收和贷后提醒里最稳定的一层信号。
        # 如果与 prestamo / credito / cuota 共同出现，区分度会非常高。
        "vence", "vencen",
        "vencido", "vencida", "vencidos", "vencidas",
        "vencimiento", "vencimientos",
        "atraso", "atrasos",
        "atrasado", "atrasada", "atrasados", "atrasadas",
        "pendiente", "pendientes",
        "mora", "moras",
        "moroso", "morosa", "morosos", "morosas",
        "morosidad",

        # ---------------------------------------------------------------------
        # F. 利息 / 罚金 / 费用 / 折扣
        # ---------------------------------------------------------------------
        # 常见于逾期施压、催收提醒和促还款营销中。
        # 尤其是“今天支付有折扣”“已产生罚金/利息”这类文案。
        "interes", "intereses", "interés",
        "penalizacion", "penalizaciones", "penalización",
        "cargo", "cargos",
        "recargo", "recargos",
        "costo", "costos",
        "descuento", "descuentos",

        # ---------------------------------------------------------------------
        # G. 催收 / 法务升级 / 压力传导
        # ---------------------------------------------------------------------
        # 这是识别法务催收、司法施压、催收升级的关键词组。
        # 即使短信里没有明确写 prestamo，也常常能单独构成强信号。
        "cobranza", "cobranzas",
        "cobro", "cobros",
        "demanda", "demandas",
        "legal", "legales",
        "juridico", "juridicos", "jurídico", "jurídicos",
        "judicial", "judiciales",
        "embargo", "embargos",
        "proceso", "procesos",
        "expediente", "expedientes",
        "convenio", "convenios",
        "historial",
        "crediticio", "crediticia", "crediticios", "crediticias",
        "record", "récord",
        "cartera",
    },
    "medium": {
        # ---------------------------------------------------------------------
        # 中等信号词
        # ---------------------------------------------------------------------
        # 这组词在贷款短信和普通金融短信里都会出现。
        # 它们更适合作为上下文辅助，而不是单独决定结果。
        "pagar", "paga", "pague", "pago", "pagos",
        "regularice", "regularizar", "regularizando",
        "importe", "importes",
        "fecha", "fechas",
        "linea", "línea",
        "sube", "subido", "subida", "subir",
        "canales", "oficiales",
        "aplicacion", "aplicaciones", "aplicación",
        "app",
        "rapido", "rapida", "rápido", "rápida",
        "inmediato", "inmediata",
        "beneficio", "beneficios",
        "calificacion", "calificación",
        "reembolso",
        "saldo", "saldos",
        "cuenta", "cuentas",
        "bancaria", "bancarias", "bancario", "bancarios",
    },
    "low_weight_noise": {
        # ---------------------------------------------------------------------
        # 噪音词 / 泛金融词
        # ---------------------------------------------------------------------
        # 这些词虽然和金融、支付、账户、账单有关，但并不能说明就是现金贷。
        # 如果给这组词太高权重，会把验证码、银行通知、电费账单、电信业务、
        # remesa 汇款之类短信误判进来。
        "banco", "bancos",
        "token", "tokens",
        "codigo", "codigos", "código", "códigos",
        "verificacion", "verificación",
        "registro",
        "tarjeta", "tarjetas",
        "dinero",
        "efectivo",
        "remesa", "remesas",
        "ahorro", "ahorros",
        "seguro", "seguros",
        "iva",
        "factura", "facturas",
        "servicio", "servicios",
        "energia", "energía",
        "musica", "música",
        "plan", "planes",
    },
    "ignore": {
        # ---------------------------------------------------------------------
        # 完全忽略词
        # ---------------------------------------------------------------------
        # 这些词在各种营销短信、服务通知、普通提示里都太常见，
        # 几乎没有区分度，因此直接从 token 集合中剔除。
        "cliente", "estimado", "hola", "hoy", "manana", "mañana",
        "dia", "dias", "días", "aqui", "aquí",
    },
    "brand_hints": {
        # ---------------------------------------------------------------------
        # 平台名 / 产品名提示词
        # ---------------------------------------------------------------------
        # 来源于历史样本里的平台或品牌痕迹。
        # 它们不是绝对证据，但当短信同时出现额度、审批、账户、还款等词时，
        # 这些品牌提示会显著增强贷款平台语境。
        "creditofacil",
        "creditya",
        "tuprestamo",
        "intiprestamo",
        "confimoneda",
        "lemo",
        "bula",
        "vana",
        "activagroup",
        "globalcredit",
        "unicoservi",
        "interconsumo",
        "micro",
        "seda",
        "mano",
    },
}


def _strip_accents(text: str) -> str:
    # 去掉西语重音，统一 préstamo / prestamo 这类写法。
    return "".join(
        ch for ch in unicodedata.normalize("NFD", text)
        if unicodedata.category(ch) != "Mn"
    )


def normalize_sms_text(text: str) -> str:
    # 对短信正文做归一化。
    # 主要处理：
    # 1. 统一小写
    # 2. 修正常见数字替字母混写，如 dep0sito / apr0bada
    # 3. 去掉重音
    # 4. 删除 URL
    # 5. 清理标点和多余空格
    text = str(text or "").lower()
    text = (
        text.replace("0", "o")
        .replace("1", "i")
        .replace("3", "e")
        .replace("4", "a")
        .replace("5", "s")
    )
    text = _strip_accents(text)
    text = re.sub(r"https?://\S+", " ", text)
    text = re.sub(r"[^a-z0-9\s]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def normalize_keyword_set(words: Iterable[str]) -> set[str]:
    # 对词典也做同样的归一化，保证“短信文本”和“关键词表”在同一规则下比较。
    return {nw for w in words if (nw := normalize_sms_text(w))}


NORMALIZED_KEYWORDS = {
    key: normalize_keyword_set(values)
    for key, values in SMS_LOAN_KEYWORDS.items()
}


# 统一封装单条短信的识别结果。
# 后续写 Excel 时，直接从这里取分数、标签、命中词和细分类别。
@dataclass
class ClassificationResult:
    normalized_text: str
    label: str
    finance_segment: str
    score: float
    strong_hits: list[str]
    medium_hits: list[str]
    brand_hits: list[str]
    noise_hits: list[str]
    rule_hits: list[str]
    exclusion_hits: list[str]


def tokenize(text: str) -> set[str]:
    # 分词后直接去掉 ignore 里的停用词，只保留对识别有帮助的 token。
    normalized = normalize_sms_text(text)
    if not normalized:
        return set()
    return set(normalized.split()) - NORMALIZED_KEYWORDS["ignore"]


def rule_based_boost(tokens: set[str]) -> tuple[float, list[str]]:
    # 组合规则加分。
    # 单个关键词有时不够稳定，所以这里额外捕捉几组高置信共现模式。
    # 例如：
    # - 借贷核心词 + 逾期词
    # - 借贷核心词 + 还款凭证/额度词
    # - 审批/放款词 + 借贷核心词
    # - 催收/法务词
    score = 0.0
    hits: list[str] = []

    loan_core = {"prestamo", "credito", "microcredito"}
    overdue = {"vence", "vencido", "vencida", "mora", "morosidad", "atraso", "atrasado", "pendiente"}
    approval = {"solicitud", "aprobado", "aprobada", "aprobacion", "rechazo", "desembolso", "depositado", "depositada"}
    collection = {"convenio", "legal", "juridico", "judicial", "embargo", "expediente", "cobranza", "demanda"}
    proof = {"comprobante", "recibo", "abono", "cuota", "historial", "crediticio", "limite"}
    urgent = {"hoy", "ahora", "pronto"}

    if tokens & loan_core and tokens & overdue:
        score += 4.0
        hits.append("loan+overdue")

    if tokens & loan_core and tokens & proof:
        score += 3.0
        hits.append("loan+proof_or_limit")

    if tokens & approval and (tokens & loan_core or {"linea", "financiamiento"} & tokens):
        score += 3.5
        hits.append("approval_or_disbursement")

    if tokens & collection:
        score += 3.5
        hits.append("collection_or_legal")

    if {"paga", "pagar", "pague"} & tokens and urgent & tokens and tokens & loan_core:
        score += 2.0
        hits.append("urgent_payment")

    return score, hits


def exclusion_rule_adjustment(tokens: set[str]) -> tuple[float, list[str]]:
    # 排除场景修正。
    # 这部分不是为了识别贷款，而是为了压低容易误判的上下文，
    # 尤其是电信、公共事业、信用卡充值、普通账单这些场景。
    #
    # 设计思路是：
    # - 如果明显像电信/水电/话费类短信，就大幅扣分
    # - 如果更像信用卡、银行账户通知，就避免被当成现金贷
    # - 但对带有平台痕迹和贷后语境的“服务类”短信要留保护口
    score = 0.0
    hits: list[str] = []

    telecom_core = {
        "recarga", "paquete", "datos", "llamadas", "conexion", "internet",
        "wifi", "saldo", "compras", "debito", "factura", "facturas",
        "servicio", "corte", "suspendido", "plan", "planes",
    }
    telecom_brands = {
        "claro", "tigo", "movistar", "energuate",
    }
    utility_terms = {
        "energia", "agua", "luz", "electricidad", "oficinavirtual",
    }
    loan_core = {"prestamo", "microcredito", "cuota", "cuotas", "vencimiento", "vence", "vencido", "vencida"}
    collection_core = {"cobranza", "juridico", "embargo", "demanda", "expediente", "convenio", "legal"}
    bank_core = {"credito", "banco", "tarjeta", "cuenta", "saldo", "debito"}
    loan_app_context = {"credito", "cuota", "cuotas", "vence", "vencimiento", "mora", "atraso", "atrasos"}
    platform_hints = NORMALIZED_KEYWORDS["brand_hints"]
    bank_statement_context = {"pago", "pagos", "corte", "cobros", "cargos", "tc", "tarjeta", "banco"}

    telecom_hit_count = len(tokens & telecom_core) + len(tokens & telecom_brands) + len(tokens & utility_terms)
    has_loan_signal = bool(tokens & loan_core)
    has_collection_signal = bool(tokens & collection_core)
    has_finance_signal = bool(tokens & bank_core)
    has_platform_hint = bool(tokens & platform_hints)
    protected_loan_service_context = bool(tokens & loan_app_context) and (has_platform_hint or bool({"credito"} & tokens))

    if telecom_hit_count >= 2 and not has_loan_signal and not has_collection_signal:
        score -= 4.0
        hits.append("telecom_or_utility_context")

    if {"tarjeta"} & tokens and {"recarga"} & tokens and not has_loan_signal:
        score -= 3.0
        hits.append("card_or_recharge_context")

    if (
        {"factura", "facturas", "servicio", "corte"} & tokens
        and not has_loan_signal
        and not has_collection_signal
        and not protected_loan_service_context
        and not (tokens & bank_statement_context and (tokens & {"banco", "tarjeta", "tc", "credito"}))
    ):
        score -= 2.5
        hits.append("billing_service_context")

    if {"credito"} & tokens and {"tarjeta"} & tokens and {"recarga"} & tokens and not has_loan_signal:
        score -= 2.0
        hits.append("credito_card_context")

    if {"credito"} & tokens and {"banco", "tarjeta", "cuenta"} & tokens and not has_loan_signal and not has_collection_signal:
        score += 1.0
        hits.append("banking_context")

    if has_finance_signal and not has_loan_signal and not has_collection_signal:
        score += 0.5
        hits.append("generic_finance_context")

    return score, hits


def determine_finance_segment(tokens: set[str], result_label: str, exclusion_hits: list[str]) -> str:
    # 在总标签之外补一个更细的金融类别。
    # 这一步主要服务人工复核与后续分析，帮助区分：
    # - 贷款/放款
    # - 催收/法务
    # - 银行账户类通知
    # - OTP/安全验证
    # - 一般金融营销
    # - 已命中排除规则的非目标场景
    #
    # 注意这里输出的是“细分类别”，不是 score。
    # 也就是说，一条短信即使进入输出结果，仍然可以被进一步标成
    # collection_like / finance_otp_like / bank_or_finance_like 等不同子类。
    loan_core = {"prestamo", "microcredito", "desembolso", "cuota", "mora", "vencido", "vencida"}
    collection_core = {"cobranza", "juridico", "embargo", "demanda", "expediente", "convenio", "legal"}
    bank_core = {"credito", "banco", "tarjeta", "cuenta", "saldo", "debito"}
    otp_core = {"codigo", "token", "verificacion", "registro", "seguridad", "clave"}
    otp_strict = {"codigo", "token", "verificacion", "registro", "clave"}
    marketing_core = {
        "aprobado", "aprobada", "solicitud", "monto", "disponible", "financiamiento",
        "desembolso", "segundos", "tramites", "directo", "pedi", "necesitas",
    }
    platform_hints = NORMALIZED_KEYWORDS["brand_hints"]

    if any(hit in exclusion_hits for hit in {"telecom_or_utility_context", "card_or_recharge_context", "billing_service_context", "credito_card_context"}):
        return "excluded_non_finance"
    if tokens & collection_core:
        return "collection_like"
    if (tokens & otp_strict or ({"seguridad"} & tokens and {"registrarse", "registro"} & tokens)) and (tokens & bank_core or tokens & platform_hints):
        return "finance_otp_like"
    if (tokens & marketing_core and (tokens & bank_core or tokens & platform_hints)) or (tokens & platform_hints and {"credito", "cuenta"} & tokens):
        return "loan_marketing_like"
    if tokens & loan_core or result_label == "cash_loan_like":
        return "loan_like"
    if tokens & bank_core:
        return "bank_or_finance_like"
    return "other_finance_like"


def score_sms(text: str, loan_threshold: float = 6.0, min_output_score: float = 3.0) -> ClassificationResult:
    # 单条短信的总评分入口。
    # 这里会汇总：
    # - 强/中/品牌/噪音关键词命中
    # - 组合规则加分
    # - 排除场景修正
    # 最终输出统一的 ClassificationResult。
    #
    # 标签规则保持两层：
    # - score >= loan_threshold        -> cash_loan_like
    # - min_output_score <= score <   -> suspicious
    # 低于输出阈值的短信不会进入最终结果文件。
    normalized_text = normalize_sms_text(text)
    tokens = tokenize(text)

    strong_hits = sorted(tokens & NORMALIZED_KEYWORDS["strong"])
    medium_hits = sorted(tokens & NORMALIZED_KEYWORDS["medium"])
    brand_hits = sorted(tokens & NORMALIZED_KEYWORDS["brand_hints"])
    noise_hits = sorted(tokens & NORMALIZED_KEYWORDS["low_weight_noise"])

    score = 0.0
    score += len(strong_hits) * 3.0
    score += len(medium_hits) * 1.0
    score += len(brand_hits) * 2.0
    score -= len(noise_hits) * 0.35

    rule_score, rule_hits = rule_based_boost(tokens)
    score += rule_score
    exclusion_score, exclusion_hits = exclusion_rule_adjustment(tokens)
    score += exclusion_score

    if not ({"prestamo", "credito", "microcredito"} & tokens) and not (
        {"cobranza", "juridico", "embargo", "demanda", "expediente", "convenio"} & tokens
    ):
        score -= 1.5

    score = round(score, 2)

    if score >= loan_threshold:
        label = "cash_loan_like"
    elif score >= min_output_score:
        label = "suspicious"
    else:
        label = "ignored"

    finance_segment = determine_finance_segment(tokens, label, exclusion_hits)

    return ClassificationResult(
        normalized_text=normalized_text,
        label=label,
        finance_segment=finance_segment,
        score=score,
        strong_hits=strong_hits,
        medium_hits=medium_hits,
        brand_hits=brand_hits,
        noise_hits=noise_hits,
        rule_hits=rule_hits,
        exclusion_hits=exclusion_hits,
    )


# 翻译客户端采用“在线请求 + 本地缓存”的轻量设计。
# 重点是让重复运行时尽量复用已有 content -> content_zh 的映射，
# 降低重复翻译成本。
class GoogleTranslateClient:
    def __init__(
        self,
        enabled: bool,
        cache_path: str | Path | None,
        target_lang: str = "zh-CN",
        timeout: int = 20,
        sleep_seconds: float = 0.15,
    ) -> None:
        self.enabled = enabled
        self.target_lang = target_lang
        self.timeout = timeout
        self.sleep_seconds = sleep_seconds
        self.session = requests.Session()
        self.cache_path = Path(cache_path) if cache_path else None
        self.cache: dict[str, str] = {}
        if self.cache_path and self.cache_path.exists():
            try:
                self.cache = json.loads(self.cache_path.read_text(encoding="utf-8"))
            except Exception:
                self.cache = {}

    def save_cache(self) -> None:
        # 将当前翻译缓存写回本地 JSON 文件。
        if not self.cache_path:
            return
        self.cache_path.write_text(
            json.dumps(self.cache, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    def translate_text(self, text: str) -> str:
        # 单条翻译入口：优先查缓存，未命中再请求在线翻译。
        text = str(text or "").strip()
        if not text:
            return ""
        if text in self.cache:
            return self.cache[text]
        if not self.enabled:
            return ""

        url = "https://translate.googleapis.com/translate_a/single"
        params = {
            "client": "gtx",
            "sl": "auto",
            "tl": self.target_lang,
            "dt": "t",
            "q": text,
        }
        try:
            response = self.session.get(url, params=params, timeout=self.timeout)
            response.raise_for_status()
            data = response.json()
            translated = "".join(part[0] for part in data[0] if part and part[0])
        except Exception:
            translated = ""

        self.cache[text] = translated
        time.sleep(self.sleep_seconds)
        return translated

    def translate_many(self, texts: Iterable[str]) -> dict[str, str]:
        # 批量翻译包装器。
        # 这里会先去重，再逐条调用 translate_text，
        # 适合在导出结果前统一处理所有短信内容。
        if not self.enabled:
            return {}

        unique_texts: list[str] = []
        seen: set[str] = set()
        for text in texts:
            key = str(text or "").strip()
            if key and key not in seen:
                unique_texts.append(key)
                seen.add(key)

        if not unique_texts:
            return {}

        print(f"准备翻译 {len(unique_texts)} 条去重后的短信内容...")
        translated_count = 0
        for idx, text in enumerate(unique_texts, start=1):
            if text not in self.cache:
                self.translate_text(text)
                translated_count += 1
                if idx % 200 == 0:
                    print(f"  已翻译 {idx}/{len(unique_texts)} 条")
                    self.save_cache()

        self.save_cache()
        print(f"翻译完成，本次新增翻译 {translated_count} 条。")
        return {text: self.cache.get(text, "") for text in unique_texts}


RESULT_COLUMNS = [
    "sheet_name",
    "label",
    "finance_segment",
    "score",
    "normalized_text",
    "content_zh",
    "strong_hits",
    "medium_hits",
    "brand_hits",
    "noise_hits",
    "rule_hits",
    "exclusion_hits",
]

COMPACT_INPUT_COLUMNS = ["orderId", "submit", "phone", "content", "time"]


def collect_rows(input_file: str | Path) -> tuple[list[str], list[dict[str, object]]]:
    # 读取所有 sheet，汇总成统一的行列表。
    # 每行额外补充 _sheet_name，方便后续回溯来源工作表。
    print(f"读取文件: {input_file}")
    wb = load_workbook(input_file, read_only=False, data_only=False)

    all_columns: list[str] = []
    all_rows: list[dict[str, object]] = []

    for sheet_idx, ws in enumerate(wb.worksheets, start=1):
        print(f"读取 sheet {sheet_idx}/{len(wb.worksheets)}: {ws.title}")
        header_row = [cell.value for cell in ws[1]]
        headers = [str(v).strip() for v in header_row if v is not None]
        header_map = {header: idx for idx, header in enumerate(headers)}

        if "content" not in header_map:
            print(f"  跳过 sheet '{ws.title}'，缺少列: ['content']")
            continue

        for header in headers:
            if header not in all_columns:
                all_columns.append(header)

        sheet_rows = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict: dict[str, object] = {}
            for header, idx in header_map.items():
                row_dict[header] = row[idx] if idx < len(row) else None
            row_dict["_sheet_name"] = ws.title
            all_rows.append(row_dict)
            sheet_rows += 1

        print(f"  读取 {sheet_rows} 行")

    print(f"\n总共读取 {len(all_rows)} 行")
    return all_columns, all_rows


def build_filtered_rows(
    rows: list[dict[str, object]],
    loan_threshold: float,
    min_output_score: float,
) -> list[tuple[dict[str, object], ClassificationResult]]:
    # 对短信逐条评分，只保留达到输出阈值的记录。
    filtered_rows: list[tuple[dict[str, object], ClassificationResult]] = []
    total_rows = len(rows)
    print("开始评分...")
    for idx, row_dict in enumerate(rows):
        if idx % 50000 == 0:
            print(f"已处理 {idx}/{total_rows} 行, 当前保留 {len(filtered_rows)} 行...")

        content = str(row_dict.get("content") or "")
        result = score_sms(content, loan_threshold=loan_threshold, min_output_score=min_output_score)
        if result.score >= min_output_score:
            filtered_rows.append((row_dict, result))

    print(f"评分完成，保留 {len(filtered_rows)} 行。")
    return filtered_rows


def save_output(
    output_file: str | Path,
    original_columns: list[str],
    filtered_rows: list[tuple[dict[str, object], ClassificationResult]],
    translator: GoogleTranslateClient,
    output_mode: str,
) -> None:
    # 生成结果工作簿：
    # - compact: 只保留五个核心输入字段
    # - full: 保留原始全部字段
    # 两种模式都会追加识别结果列与可选翻译列。
    wb = Workbook()
    ws = wb.active
    ws.title = "all_sheets_scored"

    if output_mode == "compact":
        selected_columns = [col for col in COMPACT_INPUT_COLUMNS if col in original_columns]
    else:
        selected_columns = original_columns

    ws.append(selected_columns + RESULT_COLUMNS)

    translations = translator.translate_many(
        str(row_dict.get("content") or "")
        for row_dict, _ in filtered_rows
    )

    for row_dict, result in filtered_rows:
        original_values = [row_dict.get(col) for col in selected_columns]
        content = str(row_dict.get("content") or "")
        ws.append(
            original_values + [
                row_dict.get("_sheet_name", ""),
                result.label,
                result.finance_segment,
                result.score,
                result.normalized_text,
                translations.get(content, ""),
                ",".join(result.strong_hits),
                ",".join(result.medium_hits),
                ",".join(result.brand_hits),
                ",".join(result.noise_hits),
                ",".join(result.rule_hits),
                ",".join(result.exclusion_hits),
            ]
        )

    for column_cells in ws.columns:
        column_letter = column_cells[0].column_letter
        header = str(column_cells[0].value or "")
        if header in {"content", "normalized_text", "content_zh"}:
            width = 90
        elif header in {"strong_hits", "medium_hits", "brand_hits", "noise_hits", "rule_hits", "exclusion_hits"}:
            width = 35
        else:
            width = min(max(len(header) + 4, 12), 28)
        ws.column_dimensions[column_letter].width = width

    wb.save(output_file)
    print(f"输出完成: {output_file}")


def process_all_sheets(
    input_file: str | Path,
    output_file: str | Path,
    loan_threshold: float = 6.0,
    min_output_score: float = 3.0,
    translate_zh: bool = False,
    translation_cache: str | Path | None = None,
    output_mode: str = "compact",
) -> None:
    # 主流程：读取 -> 打分筛选 -> 可选翻译 -> 输出结果。
    original_columns, rows = collect_rows(input_file)
    filtered_rows = build_filtered_rows(rows, loan_threshold, min_output_score)

    translator = GoogleTranslateClient(
        enabled=translate_zh,
        cache_path=translation_cache,
    )
    save_output(output_file, original_columns, filtered_rows, translator, output_mode)


def parse_args(argv: list[str]) -> argparse.Namespace:
    # 命令行参数入口。
    parser = argparse.ArgumentParser(
        description="保留筛选后的原始列，并可选追加中文翻译的金融短信筛选脚本。"
    )
    parser.add_argument("input_file", nargs="?", default="input.xlsx")
    parser.add_argument("output_file", nargs="?", default="output_scored_all_sheets_v2.xlsx")
    parser.add_argument("--loan-threshold", type=float, default=6.0)
    parser.add_argument("--min-output-score", type=float, default=3.0)
    parser.add_argument("--translate-zh", action="store_true", help="调用在线翻译并输出 content_zh 列。")
    parser.add_argument(
        "--translation-cache",
        default="translation_cache_zh.json",
        help="翻译缓存文件路径，避免重复翻译相同短信。",
    )
    parser.add_argument(
        "--output-mode",
        choices=["compact", "full"],
        default="compact",
        help="compact 仅保留前五列加结果列；full 保留全部原始列加结果列。",
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> None:
    # 脚本入口。
    args = parse_args(argv or sys.argv[1:])
    process_all_sheets(
        input_file=args.input_file,
        output_file=args.output_file,
        loan_threshold=args.loan_threshold,
        min_output_score=args.min_output_score,
        translate_zh=args.translate_zh,
        translation_cache=args.translation_cache,
        output_mode=args.output_mode,
    )


if __name__ == "__main__":
    main()
