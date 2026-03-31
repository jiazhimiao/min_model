"""
处理短信 Excel 文件的第一张 sheet，只输出以下原始字段：
orderId, submit, phone, content, time
并追加新的识别结果字段。

行为说明：
- 只读取第一个 sheet
- 只保留分数 >= 3.0 的行
- 标签规则：
    score >= 6.0        -> cash_loan_like
    3.0 <= score < 6.0 -> suspicious
- 输出为新的 xlsx 文件

使用方式：
    python process_sms_xlsx_first_sheet_zh.py input.xlsx output.xlsx

示例：
    python process_sms_xlsx_first_sheet_zh.py raw_sms.xlsx scored_sms.xlsx

如果不传命令行参数，则默认使用：
    input.xlsx
    output_scored_first_sheet.xlsx
"""

from __future__ import annotations

import re
import sys
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook


# =============================================================================
# 整体设计说明
# =============================================================================
# 这个脚本是一个“基于规则 + 关键词打分”的危地马拉西语现金贷短信识别器。
#
# 为什么不用模型，先用规则？
# 1. 短信文本很短，很多时候单词信号比复杂语义更稳定。
# 2. 短信里经常出现脏数据和故意混淆写法，例如：
#       pag0 / apr0bada / dep0sitado
# 3. 西语短信经常存在重音不一致：
#       préstamo / prestamo
#       crédito  / credito
# 4. 空格、标点、URL 都不稳定，直接依赖完整短语容易漏掉。
#
# 所以当前方案是：
#   1) 先做文本归一化
#   2) 再切成单词 token
#   3) 根据不同关键词分组进行加权打分
#   4) 对高置信组合额外加分
#
# 这个脚本的目标不是“识别所有金融短信”，而是重点识别：
#   - 互联网现金贷短信
#   - 还款提醒短信
#   - 逾期 / 催收短信
#   - 法务升级 / 施压催收短信
#
# 文件处理逻辑：
# - 只读取 Excel 的第一张 sheet
# - 只保留原始的 5 个字段：
#       orderId, submit, phone, content, time
# - 再追加打分结果字段
# =============================================================================


# =============================================================================
# 关键词分组说明
# =============================================================================
# 当前关键词分为 5 类：
#
# 1) strong
#    高置信关键词。
#    单独出现时就很像贷款/催收语境。
#    每命中一个，分值较高。
#
# 2) medium
#    中等信号词。
#    在贷款短信里常见，但在普通银行/账单/支付短信里也会出现。
#    所以加分较低，只作为辅助上下文。
#
# 3) low_weight_noise
#    噪音词 / 泛金融词。
#    在验证码、银行通知、remesa、水电账单等短信里非常常见。
#    这类词不应该作为贷款短信的强依据，因此做轻微扣分。
#
# 4) ignore
#    几乎没有区分度的高频词。
#    直接从 token 集合里去掉，不参与评分。
#
# 5) brand_hints
#    历史样本里出现过的现金贷 / 催收平台名称。
#    它们不是绝对证据，但如果和其他贷款词一起出现，会显著增强可信度。
# =============================================================================
SMS_LOAN_KEYWORDS = {
    "strong": {
        # ---------------------------------------------------------------------
        # A. 借贷核心词
        # ---------------------------------------------------------------------
        # 这组词直接表达“借款/贷款/信用贷/小额贷”的概念。
        # 是整个规则体系里最关键的一层。
        #
        # 常见短信场景：
        # - su prestamo vence hoy
        # - tu credito ha sido aprobado
        # - microcredito disponible
        "prestamo", "prestamos", "préstamo", "préstamos",
        "credito", "creditos", "crédito", "créditos",
        "microcredito", "microcreditos", "microcrédito", "microcréditos",
        "prestamista", "prestamistas",

        # ---------------------------------------------------------------------
        # B. 贷前阶段：申请 / 审批 / 授信 / 额度
        # ---------------------------------------------------------------------
        # 这组词常见于互联网现金贷的“申请成功 / 审批通过 / 给额度”场景。
        #
        # 例如：
        # - solicitud de credito
        # - aprobado
        # - linea de financiamiento
        # - limite incrementado
        "solicitud", "solicitudes",
        "aprobado", "aprobada", "aprobados", "aprobadas",
        "aprobacion", "aprobaciones", "aprobación", "aprobaciones",
        "rechazado", "rechazada", "rechazo",
        "limite", "limites", "límite", "límites",
        "monto", "montos",
        "capacidad",
        "financiamiento",

        # ---------------------------------------------------------------------
        # C. 放款 / 到账 / 转账
        # ---------------------------------------------------------------------
        # 用于识别“已经放款 / 已转账 / 已存入账户”的贷款短信。
        #
        # 例如：
        # - fue depositado a su cuenta
        # - desembolso realizado
        # - transferencia completada
        "depositado", "depositada", "depositados", "depositadas",
        "deposito", "depositos", "depósito", "depósitos",
        "desembolso", "desembolsos",
        "transferencia", "transferencias",

        # ---------------------------------------------------------------------
        # D. 贷后还款 / 上传凭证 / 分期
        # ---------------------------------------------------------------------
        # 这组词常见于“提醒还款”“要求上传凭证”“提示分期/分笔还款”场景。
        #
        # 例如：
        # - sube tu comprobante
        # - realiza tu abono
        # - paga tu cuota
        "comprobante", "comprobantes",
        "recibo", "recibos",
        "cuota", "cuotas",
        "abono", "abonos",
        "liquidado", "liquidada", "liquidados", "liquidadas",
        "liquidacion", "liquidación",
        "cancelar", "cancele", "cancelacion", "cancelación",

        # ---------------------------------------------------------------------
        # E. 到期 / 逾期 / 拖欠
        # ---------------------------------------------------------------------
        # 这是催收短信最稳定、最强的一组信号词。
        #
        # 例如：
        # - vence hoy
        # - prestamo vencido
        # - mora grave
        # - saldo pendiente
        #
        # 如果和 prestamo / credito 一起出现，区分度会非常高。
        "vence", "vencen",
        "vencido", "vencida", "vencidos", "vencidas",
        "vencimiento", "vencimientos",
        "atraso", "atrasos",
        "atrasado", "atrasada", "atrasados", "atrasadas",
        "pendiente", "pendientes",
        "mora", "moras",
        "moroso", "morosa", "morosos", "morosas",
        "morosidad",

        # ---------------------------------------------------------------------
        # F. 利息 / 罚金 / 费用 / 折扣
        # ---------------------------------------------------------------------
        # 这组词常见于催收施压或促还款短信。
        #
        # 例如：
        # - intereses
        # - recargos
        # - penalizacion
        # - descuento si paga hoy
        "interes", "intereses", "interés",
        "penalizacion", "penalizaciones", "penalización", "penalizaciones",
        "cargo", "cargos",
        "recargo", "recargos",
        "costo", "costos",
        "descuento", "descuentos",

        # ---------------------------------------------------------------------
        # G. 催收 / 法务升级 / 组合施压
        # ---------------------------------------------------------------------
        # 这组词非常关键，用于识别法务催收和升级施压场景。
        #
        # 例如：
        # - traslado a legal
        # - convenio de pago
        # - cobro judicial
        # - embargo
        # - expediente
        #
        # 即使短信里没有明确写 prestamo / credito，只要这组词命中很多，
        # 也很可能是贷款催收短信。
        "cobranza", "cobranzas",
        "cobro", "cobros",
        "demanda", "demandas",
        "legal", "legales",
        "juridico", "juridicos", "jurídico", "jurídicos",
        "judicial", "judiciales",
        "embargo", "embargos",
        "proceso", "procesos",
        "expediente", "expedientes",
        "convenio", "convenios",
        "historial",
        "crediticio", "crediticia", "crediticios", "crediticias",
        "record", "récord",
        "cartera",
    },

    "medium": {
        # ---------------------------------------------------------------------
        # 中等信号词
        # ---------------------------------------------------------------------
        # 这组词在贷款短信里经常出现，但在普通银行 / 支付 / 账单短信里
        # 也会大量出现，因此不适合给太高权重。
        #
        # 比如：
        # - pago / pagar
        # - saldo
        # - cuenta
        # - app
        # - rapido / inmediato
        #
        # 它们更多是辅助上下文，不应单独决定结果。
        "pagar", "paga", "pague", "pago", "pagos",
        "regularice", "regularizar", "regularizando",
        "importe", "importes",
        "fecha", "fechas",
        "linea", "línea",
        "sube", "subido", "subida", "subir",
        "canales", "oficiales",
        "aplicacion", "aplicaciones", "aplicación", "aplicaciones",
        "app",
        "rapido", "rapida", "rápido", "rápida",
        "inmediato", "inmediata",
        "beneficio", "beneficios",
        "calificacion", "calificación",
        "reembolso",
        "saldo", "saldos",
        "cuenta", "cuentas",
        "bancaria", "bancarias", "bancario", "bancarios",
    },

    "low_weight_noise": {
        # ---------------------------------------------------------------------
        # 噪音词 / 泛金融词
        # ---------------------------------------------------------------------
        # 这些词在你的历史短信里很常见，但不能说明短信就是现金贷。
        #
        # 它们常见于：
        # - 银行登录提醒
        # - OTP / token / 注册验证码
        # - remesa 汇款
        # - 水电煤账单
        # - 保险 / 普通金融营销
        #
        # 如果这些词被当成强信号，会大幅增加误判。
        # 所以当前策略是：轻微扣分，而不是直接忽略。
        "banco", "bancos",
        "token", "tokens",
        "codigo", "codigos", "código", "códigos",
        "verificacion", "verificación",
        "registro",
        "tarjeta", "tarjetas",
        "dinero",
        "efectivo",
        "remesa", "remesas",
        "ahorro", "ahorros",
        "seguro", "seguros",
        "iva",
        "factura", "facturas",
        "servicio", "servicios",
        "energia", "energía",
        "musica", "música",
        "plan", "planes",
    },

    "ignore": {
        # ---------------------------------------------------------------------
        # 完全忽略词
        # ---------------------------------------------------------------------
        # 这些词在各种短信里都很常见，单独几乎没有分类价值。
        # 所以直接从 token 集合中去掉，不参与任何评分。
        "cliente", "estimado", "hola", "hoy", "manana", "mañana",
        "dia", "dias", "días", "aqui", "aquí",
    },

    "brand_hints": {
        # ---------------------------------------------------------------------
        # 平台 / 品牌 / 催收主体提示词
        # ---------------------------------------------------------------------
        # 这些名称来自你的历史贷款 / 催收短信样本。
        #
        # 注意：
        # - 品牌词不是绝对证据
        # - 但如果和其他贷款词一起出现，可信度会明显提升
        #
        # 例如：
        # - creditofacil
        # - creditya
        # - vana
        # - activagroup
        # - unicoservi
        "creditofacil",
        "creditya",
        "tuprestamo",
        "intiprestamo",
        "confimoneda",
        "lemo",
        "bula",
        "vana",
        "activagroup",
        "globalcredit",
        "unicoservi",
        "interconsumo",
        "micro",
        "seda",
        "mano",
    }
}


def _strip_accents(text: str) -> str:
    """
    去掉西语重音符号。

    例如：
        préstamo -> prestamo
        crédito  -> credito
    """
    return "".join(
        ch for ch in unicodedata.normalize("NFD", text)
        if unicodedata.category(ch) != "Mn"
    )


def normalize_sms_text(text: str) -> str:
    """
    对短信文本做归一化。

    归一化的目的：
    1. 统一大小写
    2. 处理数字替字母的混淆写法
    3. 去掉重音
    4. 去掉 URL
    5. 去掉标点，只保留字母/数字/空格
    6. 合并多余空格

    例如：
        "Tu préstamo fue dep0sitado!!!"
    会变成：
        "tu prestamo fue depositado"
    """
    text = str(text or "").lower()

    # 常见短信混淆写法：用数字代替字母
    text = (
        text.replace("0", "o")
            .replace("1", "i")
            .replace("3", "e")
            .replace("4", "a")
            .replace("5", "s")
    )

    text = _strip_accents(text)

    # URL 对单词匹配价值不大，直接删掉
    text = re.sub(r"https?://\S+", " ", text)

    # 只保留字母 / 数字 / 空格
    text = re.sub(r"[^a-z0-9\s]", " ", text)

    # 合并多空格
    text = re.sub(r"\s+", " ", text).strip()
    return text


def normalize_keyword_set(words: Iterable[str]) -> set[str]:
    """
    用和短信正文完全相同的归一化规则处理关键词集合。

    这样可以保证：
    - préstamo 和 prestamo 能统一匹配
    - pag0 和 pago 能统一匹配
    """
    out = set()
    for w in words:
        nw = normalize_sms_text(w)
        if nw:
            out.add(nw)
    return out


# 预先把所有关键词做归一化，避免每条短信重复处理关键词
NORMALIZED_KEYWORDS = {
    key: normalize_keyword_set(values)
    for key, values in SMS_LOAN_KEYWORDS.items()
}


@dataclass
class ClassificationResult:
    """
    单条短信的识别结果。
    """
    normalized_text: str
    label: str
    score: float
    strong_hits: list[str]
    medium_hits: list[str]
    brand_hits: list[str]
    noise_hits: list[str]
    rule_hits: list[str]


def tokenize(text: str) -> set[str]:
    """
    对短信做归一化后分词，并移除 ignore 词。
    """
    normalized = normalize_sms_text(text)
    if not normalized:
        return set()
    tokens = set(normalized.split())
    return tokens - NORMALIZED_KEYWORDS["ignore"]


def rule_based_boost(tokens: set[str]) -> tuple[float, list[str]]:
    """
    对高置信词组合进行额外加分。

    为什么要做组合规则？
    因为简单的单词计数虽然有效，但“词组合”往往更能说明真实语义。

    例如：
        prestamo + vence
    的组合，比单独出现 prestamo 或 vence 的意义更强。

    这里定义了几组常见高置信模式：
    - 贷款核心词 + 逾期词
    - 贷款核心词 + 凭证/额度/历史词
    - 审批/放款词 + 贷款词
    - 法务/催收词
    - 紧急还款词
    """
    score = 0.0
    hits: list[str] = []

    loan_core = {"prestamo", "credito", "microcredito"}
    overdue = {"vence", "vencido", "vencida", "mora", "morosidad", "atraso", "atrasado", "pendiente"}
    approval = {"solicitud", "aprobado", "aprobada", "aprobacion", "rechazo", "desembolso", "depositado", "depositada"}
    collection = {"convenio", "legal", "juridico", "judicial", "embargo", "expediente", "cobranza", "demanda"}
    proof = {"comprobante", "recibo", "abono", "cuota", "historial", "crediticio", "limite"}
    urgent = {"hoy", "ahora", "pronto"}

    # 贷款词 + 逾期词：很强的贷后/催收信号
    if tokens & loan_core and tokens & overdue:
        score += 4.0
        hits.append("loan+overdue")

    # 贷款词 + 凭证/额度/历史：常见于还款或风险控制短信
    if tokens & loan_core and tokens & proof:
        score += 3.0
        hits.append("loan+proof_or_limit")

    # 审批/放款 + 贷款词：很强的贷前/放款信号
    if tokens & approval and (tokens & loan_core or {"linea", "financiamiento"} & tokens):
        score += 3.5
        hits.append("approval_or_disbursement")

    # 催收/法务词：单独就很强
    if tokens & collection:
        score += 3.5
        hits.append("collection_or_legal")

    # 紧急还款：常见于施压催收
    if {"paga", "pagar", "pague"} & tokens and urgent & tokens and tokens & loan_core:
        score += 2.0
        hits.append("urgent_payment")

    return score, hits


def score_sms(text: str, loan_threshold: float = 6.0, min_output_score: float = 3.0) -> ClassificationResult:
    """
    给单条短信打分。

    当前打分逻辑：
    - strong 命中：每个 +3
    - medium 命中：每个 +1
    - brand_hints 命中：每个 +2
    - noise 命中：每个 -0.35
    - 组合规则命中：额外加分

    额外还有一个“底部惩罚”：
    如果短信里既没有核心贷款词，也没有法务/催收词，
    那么再轻微扣一点分，避免普通金融短信分数虚高。
    """
    normalized_text = normalize_sms_text(text)
    tokens = tokenize(text)

    strong_hits = sorted(tokens & NORMALIZED_KEYWORDS["strong"])
    medium_hits = sorted(tokens & NORMALIZED_KEYWORDS["medium"])
    brand_hits = sorted(tokens & NORMALIZED_KEYWORDS["brand_hints"])
    noise_hits = sorted(tokens & NORMALIZED_KEYWORDS["low_weight_noise"])

    score = 0.0
    score += len(strong_hits) * 3.0
    score += len(medium_hits) * 1.0
    score += len(brand_hits) * 2.0
    score -= len(noise_hits) * 0.35

    rule_score, rule_hits = rule_based_boost(tokens)
    score += rule_score

    # 如果没有贷款核心词，也没有催收/法务词，则轻微扣分
    if not ({"prestamo", "credito", "microcredito"} & tokens) and not (
        {"cobranza", "juridico", "embargo", "demanda", "expediente", "convenio"} & tokens
    ):
        score -= 1.5

    score = round(score, 2)

    if score >= loan_threshold:
        label = "cash_loan_like"
    elif score >= min_output_score:
        label = "suspicious"
    else:
        label = "ignored"

    return ClassificationResult(
        normalized_text=normalized_text,
        label=label,
        score=score,
        strong_hits=strong_hits,
        medium_hits=medium_hits,
        brand_hits=brand_hits,
        noise_hits=noise_hits,
        rule_hits=rule_hits,
    )


# =============================================================================
# 输入 / 输出列定义
# =============================================================================
# 输入：
#   只从第一张 sheet 中读取下面 5 列
#
# 输出：
#   保留这 5 列，再追加识别结果字段
# =============================================================================
REQUIRED_INPUT_COLUMNS = ["orderId", "submit", "phone", "content", "time"]

NEW_COLUMNS = [
    "label",
    "score",
    "normalized_text",
    "strong_hits",
    "medium_hits",
    "brand_hits",
    "noise_hits",
    "rule_hits",
]


def process_all_sheets(
    input_file: str | Path,
    output_file: str | Path,
    loan_threshold: float = 6.0,
    min_output_score: float = 3.0,
) -> None:
    """
    处理 Excel 的所有 sheet，合并后输出。

    处理逻辑：
    1. 读取所有 sheet
    2. 从中取出指定的 5 列
    3. 用 content 列做短信打分
    4. 只保留 score >= min_output_score 的行
    5. 输出新的 workbook
    """
    print(f"读取文件: {input_file}")
    wb = load_workbook(input_file, read_only=False, data_only=False)
    
    all_rows = []
    total_rows = 0
    
    for sheet_idx, ws in enumerate(wb.worksheets):
        print(f"读取 sheet {sheet_idx + 1}/{len(wb.worksheets)}: {ws.title}")
        
        header_row = [cell.value for cell in ws[1]]
        header_map = {str(v).strip(): idx for idx, v in enumerate(header_row) if v is not None}
        
        missing = [c for c in REQUIRED_INPUT_COLUMNS if c not in header_map]
        if missing:
            print(f"  跳过 sheet '{ws.title}'，缺少列: {missing}")
            continue
        
        sheet_rows = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            for col in REQUIRED_INPUT_COLUMNS:
                row_dict[col] = row[header_map[col]] if header_map[col] < len(row) else None
            row_dict["_sheet_name"] = ws.title
            all_rows.append(row_dict)
            sheet_rows += 1
        
        total_rows += sheet_rows
        print(f"  读取 {sheet_rows} 行")
    
    print(f"\n总共读取 {total_rows} 行，来自 {len(wb.worksheets)} 个 sheet")
    print("开始评分...")
    
    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "all_sheets_scored"
    
    out_ws.append(REQUIRED_INPUT_COLUMNS + ["sheet_name"] + NEW_COLUMNS)
    
    kept = 0
    for idx, row_dict in enumerate(all_rows):
        if idx % 50000 == 0:
            print(f"已处理 {idx}/{total_rows} 行, 保留 {kept} 行...")
        
        content = row_dict.get("content", "")
        result = score_sms(content, loan_threshold=loan_threshold, min_output_score=min_output_score)
        
        if result.score < min_output_score:
            continue
        
        kept += 1
        out_ws.append([
            row_dict.get("orderId"),
            row_dict.get("submit"),
            row_dict.get("phone"),
            row_dict.get("content"),
            row_dict.get("time"),
            row_dict.get("_sheet_name", ""),
            result.label,
            result.score,
            result.normalized_text,
            ",".join(result.strong_hits),
            ",".join(result.medium_hits),
            ",".join(result.brand_hits),
            ",".join(result.noise_hits),
            ",".join(result.rule_hits),
        ])
    
    width_map = {
        "A": 18, "B": 12, "C": 18, "D": 90, "E": 22,
        "F": 15, "G": 18, "H": 10, "I": 90, "J": 35,
        "K": 30, "L": 25, "M": 25, "N": 25,
    }
    for col_letter, width in width_map.items():
        out_ws.column_dimensions[col_letter].width = width
    
    out_wb.save(output_file)
    print(f"\n处理完成! 总行数: {total_rows}, 保留: {kept}")


def main():
    """
    命令行入口。

    如果传了两个参数：
        python script.py input.xlsx output.xlsx

    否则使用默认文件名：
        input.xlsx
        output_scored_all_sheets.xlsx
    """
    if len(sys.argv) >= 3:
        input_file = sys.argv[1]
        output_file = sys.argv[2]
    else:
        input_file = "input.xlsx"
        output_file = "output_scored_all_sheets.xlsx"

    process_all_sheets(input_file, output_file)


if __name__ == "__main__":
    main()
