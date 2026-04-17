from __future__ import annotations

import argparse
from functools import lru_cache
import re
import sys
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import polars as pl
from openpyxl import Workbook


# =============================================================================
# 整体设计说明
# =============================================================================
# 这个脚本用于处理 Excel 的全部 sheet，对每条短信做“关键词命中 + 规则修正”
# 的金融识别，并输出筛选后的结果文件。
#
# v3.1 的重点在于：
# 1. 在 v2 规则识别内核基础上，输出更贴近业务使用的分层结果
# 2. 读取阶段改为 polars + calamine，提高大 Excel 文件的读取速度
# 3. 主流程改为边读边评边保留，降低海量短信场景下的中间内存占用
# 4. 输出结果进一步压缩为适合人工复核与分桶使用的关键字段
#
# 整体主流程：
# 读取全部 sheet -> 提取短信内容 -> 归一化与分词 -> 打分与分类 -> 业务分层 -> 输出
#

# =============================================================================
# 关键词分组说明
# =============================================================================
# 当前关键词分为 5 类：
#
# 1) strong
#    高置信关键词。
#    单独出现时就很像贷款申请、放款、还款、逾期、催收、法务升级语境。
#    这是整个规则体系里最主要的加分来源。
#
# 2) medium
#    中等信号词。
#    这类词在贷款短信里很常见，但在一般金融通知、账户提醒、支付类短信里
#    也会大量出现，所以只能作为辅助上下文，权重不能太高。
#
# 3) low_weight_noise
#    噪音词 / 泛金融词。
#    常见于银行登录提醒、验证码、账单、汇款、电信、保险等场景。
#    它们不能证明短信属于现金贷，所以当前策略是轻微扣分，而不是直接忽略。
#
# 4) ignore
#    高度常见但几乎没有区分度的停用词。
#    会在分词后直接剔除，不参与任何评分。
#
# 5) brand_hints
#    历史样本里出现过的平台名或产品名提示词。
#    单独命中不一定能直接定性，但和借贷词、审批词、账户词一起出现时，
#    往往能明显增强“贷款平台短信”的可信度。
# =============================================================================

SMS_LOAN_KEYWORDS = {
    "strong": {
        # ---------------------------------------------------------------------
        # A. 借贷核心词
        # ---------------------------------------------------------------------
        # 直接表达“借款 / 贷款 / 信用贷 / 小额贷”的概念。
        # 这是最基础也最关键的一组词。
        #
        # 常见短信场景：
        # - tu prestamo esta disponible
        # - credito aprobado
        # - microcredito al instante
        "prestamo", "prestamos", "préstamo", "préstamos",
        "credito", "creditos", "crédito", "créditos",
        "microcredito", "microcreditos", "microcrédito", "microcréditos",
        "prestamista", "prestamistas",

        # ---------------------------------------------------------------------
        # B. 贷前阶段：申请 / 审批 / 授信 / 额度
        # ---------------------------------------------------------------------
        # 这组词常见于“申请成功”“审批通过”“给额度”“额度提升”等贷前营销语境。
        # 如果它们和 brand_hints、credito、cuenta 一起出现，通常很像平台拉新短信。
        "solicitud", "solicitudes",
        "aprobado", "aprobada", "aprobados", "aprobadas",
        "aprobacion", "aprobaciones", "aprobación",
        "rechazado", "rechazada", "rechazo",
        "limite", "limites", "límite", "límites",
        "monto", "montos",
        "capacidad",
        "financiamiento",

        # ---------------------------------------------------------------------
        # C. 放款 / 到账 / 转账
        # ---------------------------------------------------------------------
        # 用于识别“已放款 / 已到账 / 已转入账户”的场景。
        # 在贷款类短信里，这组词通常意味着用户已进入放款后阶段。
        "depositado", "depositada", "depositados", "depositadas",
        "deposito", "depositos", "depósito", "depósitos",
        "desembolso", "desembolsos",
        "transferencia", "transferencias",

        # ---------------------------------------------------------------------
        # D. 贷后还款 / 上传凭证 / 分期
        # ---------------------------------------------------------------------
        # 这组词主要出现在贷后管理和还款提醒里。
        # 例如上传凭证、支付分期、结清、取消中的上下文。
        "comprobante", "comprobantes",
        "recibo", "recibos",
        "cuota", "cuotas",
        "abono", "abonos",
        "liquidado", "liquidada", "liquidados", "liquidadas",
        "liquidacion", "liquidación",
        "cancelar", "cancele", "cancelacion", "cancelación",

        # ---------------------------------------------------------------------
        # E. 到期 / 逾期 / 拖欠
        # ---------------------------------------------------------------------
        # 这是催收和贷后提醒里最稳定的一层信号。
        # 如果与 prestamo / credito / cuota 共同出现，区分度会非常高。
        "vence", "vencen",
        "vencido", "vencida", "vencidos", "vencidas",
        "vencimiento", "vencimientos",
        "atraso", "atrasos",
        "atrasado", "atrasada", "atrasados", "atrasadas",
        "pendiente", "pendientes",
        "mora", "moras",
        "moroso", "morosa", "morosos", "morosas",
        "morosidad",

        # ---------------------------------------------------------------------
        # F. 利息 / 罚金 / 费用 / 折扣
        # ---------------------------------------------------------------------
        # 常见于逾期施压、催收提醒和促还款营销中。
        # 尤其是“今天支付有折扣”“已产生罚金/利息”这类文案。
        "interes", "intereses", "interés",
        "penalizacion", "penalizaciones", "penalización",
        "cargo", "cargos",
        "recargo", "recargos",
        "costo", "costos",
        "descuento", "descuentos",

        # ---------------------------------------------------------------------
        # G. 催收 / 法务升级 / 压力传导
        # ---------------------------------------------------------------------
        # 这是识别法务催收、司法施压、催收升级的关键词组。
        # 即使短信里没有明确写 prestamo，也常常能单独构成强信号。
        "cobranza", "cobranzas",
        "cobro", "cobros",
        "demanda", "demandas",
        "legal", "legales",
        "juridico", "juridicos", "jurídico", "jurídicos",
        "judicial", "judiciales",
        "embargo", "embargos",
        "proceso", "procesos",
        "expediente", "expedientes",
        "convenio", "convenios",
        "historial",
        "crediticio", "crediticia", "crediticios", "crediticias",
        "record", "récord",
        "cartera",
    },
    "medium": {
        # ---------------------------------------------------------------------
        # 中等信号词
        # ---------------------------------------------------------------------
        # 这组词在贷款短信和普通金融短信里都会出现。
        # 它们更适合作为上下文辅助，而不是单独决定结果。
        "pagar", "paga", "pague", "pago", "pagos",
        "regularice", "regularizar", "regularizando",
        "importe", "importes",
        "fecha", "fechas",
        "linea", "línea",
        "sube", "subido", "subida", "subir",
        "canales", "oficiales",
        "aplicacion", "aplicaciones", "aplicación",
        "app",
        "rapido", "rapida", "rápido", "rápida",
        "inmediato", "inmediata",
        "beneficio", "beneficios",
        "calificacion", "calificación",
        "reembolso",
        "saldo", "saldos",
        "cuenta", "cuentas",
        "bancaria", "bancarias", "bancario", "bancarios",
    },
    "low_weight_noise": {
        # ---------------------------------------------------------------------
        # 噪音词 / 泛金融词
        # ---------------------------------------------------------------------
        # 这些词虽然和金融、支付、账户、账单有关，但并不能说明就是现金贷。
        # 如果给这组词太高权重，会把验证码、银行通知、电费账单、电信业务、
        # remesa 汇款之类短信误判进来。
        "banco", "bancos",
        "token", "tokens",
        "codigo", "codigos", "código", "códigos",
        "verificacion", "verificación",
        "registro",
        "tarjeta", "tarjetas",
        "dinero",
        "efectivo",
        "remesa", "remesas",
        "ahorro", "ahorros",
        "seguro", "seguros",
        "iva",
        "factura", "facturas",
        "servicio", "servicios",
        "energia", "energía",
        "musica", "música",
        "plan", "planes",
    },
    "ignore": {
        # ---------------------------------------------------------------------
        # 完全忽略词
        # ---------------------------------------------------------------------
        # 这些词在各种营销短信、服务通知、普通提示里都太常见，
        # 几乎没有区分度，因此直接从 token 集合中剔除。
        "cliente", "estimado", "hola", "hoy", "manana", "mañana",
        "dia", "dias", "días", "aqui", "aquí",
    },
    "brand_hints": {
        # ---------------------------------------------------------------------
        # 平台名 / 产品名提示词
        # ---------------------------------------------------------------------
        # 来源于历史样本里的平台或品牌痕迹。
        # 它们不是绝对证据，但当短信同时出现额度、审批、账户、还款等词时，
        # 这些品牌提示会显著增强贷款平台语境。
        "creditofacil",
        "creditya",
        "tuprestamo",
        "intiprestamo",
        "confimoneda",
        "lemo",
        "bula",
        "vana",
        "activagroup",
        "globalcredit",
        "unicoservi",
        "interconsumo",
        "micro",
        "seda",
        "mano",
    },
}

COMPACT_INPUT_COLUMNS = ["orderId", "submit", "phone", "content", "time"]


def _strip_accents(text: str) -> str:
    # 去掉西语重音，统一 préstamo / prestamo 这类写法。
    return "".join(
        ch for ch in unicodedata.normalize("NFD", text)
        if unicodedata.category(ch) != "Mn"
    )


@lru_cache(maxsize=200000)
def normalize_sms_text(text: str) -> str:
    # 对短信正文做归一化。
    # 主要处理：
    # 1. 统一小写
    # 2. 修正常见数字替字母混写，如 dep0sito / apr0bada
    # 3. 去掉重音
    # 4. 删除 URL
    # 5. 清理标点和多余空格
    text = str(text or "").lower()
    text = (
        text.replace("0", "o")
        .replace("1", "i")
        .replace("3", "e")
        .replace("4", "a")
        .replace("5", "s")
    )
    text = _strip_accents(text)
    text = re.sub(r"https?://\S+", " ", text)
    text = re.sub(r"[^a-z0-9\s]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def normalize_keyword_set(words: Iterable[str]) -> set[str]:
    # 对词典也做同样的归一化，保证“短信文本”和“关键词表”在同一规则下比较。
    return {normalized for word in words if (normalized := normalize_sms_text(word))}


NORMALIZED_KEYWORDS = {
    key: normalize_keyword_set(values)
    for key, values in SMS_LOAN_KEYWORDS.items()
}


# 统一封装单条短信的识别结果。
# 后续写 Excel 时，直接从这里取分数、标签、命中词和细分类别。
@dataclass
class ClassificationResult:
    normalized_text: str
    label: str
    finance_segment: str
    score: float
    strong_hits: list[str]
    medium_hits: list[str]
    brand_hits: list[str]
    noise_hits: list[str]
    rule_hits: list[str]
    exclusion_hits: list[str]


FINANCE_SEGMENT_ZH = {
    "excluded_non_finance": "排除类/非目标金融短信",
    "collection_like": "催收法务类",
    "finance_otp_like": "金融验证码/注册安全类",
    "loan_marketing_like": "贷款营销/授信触达类",
    "loan_like": "贷款主链路类",
    "bank_or_finance_like": "银行/账户/信用卡/普通金融服务类",
    "other_finance_like": "其他弱金融相关类",
}


def get_score_band(score: float) -> tuple[str, str]:
    if score < 3:
        return "S0", "低于保留线"
    if score < 5:
        return "S1", "弱金融相关"
    if score < 8:
        return "S2", "中等相关"
    if score < 12:
        return "S3", "高相关"
    return "S4", "极高相关"


def get_priority_and_bucket(finance_segment: str, score: float) -> tuple[str, str, str]:
    if finance_segment == "excluded_non_finance" or score < 3:
        return "D", "排除或低优先级", "主要为运营商/公用事业/明显非目标短信，或证据太弱"

    if finance_segment in {"loan_like", "collection_like"} and score >= 8:
        return "A", "核心贷款催收", "贷款/催收证据较强，适合作为最高优先级池"

    if finance_segment in {"loan_like", "collection_like"} and score < 5:
        return "B-", "边界贷款样本", "类别偏贷款/催收，但总证据偏弱，建议单独复核"

    if finance_segment == "loan_marketing_like" and score < 5:
        return "C+", "低优先级贷款营销", "明显属于贷款营销，但证据较弱，不建议与排除类混放"

    if (
        finance_segment in {"loan_like", "collection_like"} and 5 <= score < 8
    ) or (
        finance_segment == "loan_marketing_like" and score >= 5
    ):
        return "B", "贷款相关", "与贷款业务高度相关，但不一定是强催收"

    if finance_segment in {"bank_or_finance_like", "finance_otp_like", "other_finance_like"} and score >= 3:
        return "C", "银行/金融服务", "银行、验证码、普通金融服务类，保留但优先级低于贷款主池"

    return "D", "排除或低优先级", "不满足主要金融分层口径"


RESULT_COLUMNS_V3_1 = [
    "sheet_name",
    "finance_segment_zh",
    "priority_level",
    "business_bucket",
]

RULE_LOAN_CORE = frozenset({"prestamo", "credito", "microcredito"})
RULE_OVERDUE = frozenset({"vence", "vencido", "vencida", "mora", "morosidad", "atraso", "atrasado", "pendiente"})
RULE_APPROVAL = frozenset({"solicitud", "aprobado", "aprobada", "aprobacion", "rechazo", "desembolso", "depositado", "depositada"})
RULE_COLLECTION = frozenset({"convenio", "legal", "juridico", "judicial", "embargo", "expediente", "cobranza", "demanda"})
RULE_PROOF = frozenset({"comprobante", "recibo", "abono", "cuota", "historial", "crediticio", "limite"})
RULE_URGENT = frozenset({"hoy", "ahora", "pronto"})

EXCLUSION_TELECOM_CORE = frozenset({
    "recarga", "paquete", "datos", "llamadas", "conexion", "internet",
    "wifi", "saldo", "compras", "debito", "factura", "facturas",
    "servicio", "corte", "suspendido", "plan", "planes",
})
EXCLUSION_TELECOM_BRANDS = frozenset({"claro", "tigo", "movistar", "energuate"})
EXCLUSION_UTILITY_TERMS = frozenset({"energia", "agua", "luz", "electricidad", "oficinavirtual"})
EXCLUSION_LOAN_CORE = frozenset({"prestamo", "microcredito", "cuota", "cuotas", "vencimiento", "vence", "vencido", "vencida"})
EXCLUSION_COLLECTION_CORE = frozenset({"cobranza", "juridico", "embargo", "demanda", "expediente", "convenio", "legal"})
EXCLUSION_BANK_CORE = frozenset({"credito", "banco", "tarjeta", "cuenta", "saldo", "debito"})
EXCLUSION_LOAN_APP_CONTEXT = frozenset({"credito", "cuota", "cuotas", "vence", "vencimiento", "mora", "atraso", "atrasos"})
EXCLUSION_BANK_STATEMENT_CONTEXT = frozenset({"pago", "pagos", "corte", "cobros", "cargos", "tc", "tarjeta", "banco"})
EXCLUSION_CARD_FEEDBACK_CONTEXT = frozenset({"califica", "encuesta", "opinion", "experiencia", "satisfaccion", "evaluacion", "solicitudes"})
EXCLUSION_BANK_TRANSFER_CONTEXT = frozenset({"ach", "transf", "autorizacion", "localidad", "acreditado"})
EXCLUSION_LOAN_HARD_CONTEXT = frozenset({"prestamo", "cuota", "mora", "vencido", "vencimiento", "cobranza"})

SEGMENT_LOAN_CORE = frozenset({"prestamo", "microcredito", "desembolso", "cuota", "mora", "vencido", "vencida"})
SEGMENT_COLLECTION_CORE = frozenset({"cobranza", "juridico", "embargo", "demanda", "expediente", "convenio", "legal"})
SEGMENT_BANK_CORE = frozenset({"credito", "banco", "tarjeta", "cuenta", "saldo", "debito"})
SEGMENT_OTP_STRICT = frozenset({"codigo", "token", "verificacion", "registro", "clave"})
SEGMENT_MARKETING_CORE = frozenset({
    "aprobado", "aprobada", "solicitud", "monto", "disponible", "financiamiento",
    "desembolso", "segundos", "tramites", "directo", "pedi", "necesitas",
})
SEGMENT_LOAN_SERVICE_CONTEXT = frozenset({
    "pago", "pagos", "vence", "vencimiento", "vencido", "mora",
    "prestamo", "cuota", "reembolso", "mantenimiento",
})
SEGMENT_EXCLUSION_HITS = frozenset({
    "telecom_or_utility_context",
    "card_or_recharge_context",
    "billing_service_context",
    "credito_card_context",
})
SEGMENT_BANK_TRANSFER_CONTEXT = frozenset({"ach", "transf", "autorizacion", "localidad", "transferencia", "deposito", "acreditado"})
SEGMENT_REPAYMENT_OR_COLLECTION_CONTEXT = frozenset({"cuota", "pago", "pagos", "vence", "vencimiento", "mora", "prestamo", "cobranza"})

SCORE_REQUIRED_LOAN_SIGNAL = frozenset({"prestamo", "credito", "microcredito"})
SCORE_REQUIRED_COLLECTION_SIGNAL = frozenset({"cobranza", "juridico", "embargo", "demanda", "expediente", "convenio"})


@lru_cache(maxsize=200000)
def tokenize(text: str) -> frozenset[str]:
    # 分词后直接去掉 ignore 里的停用词，只保留对识别有帮助的 token。
    normalized = normalize_sms_text(text)
    if not normalized:
        return frozenset()
    return frozenset(normalized.split()) - NORMALIZED_KEYWORDS["ignore"]


def rule_based_boost(tokens: set[str]) -> tuple[float, list[str]]:
    # 组合规则加分。
    # 单个关键词有时不够稳定，所以这里额外捕捉几组高置信共现模式。
    # 例如：
    # - 借贷核心词 + 逾期词
    # - 借贷核心词 + 还款凭证/额度词
    # - 审批/放款词 + 借贷核心词
    # - 催收/法务词
    score = 0.0
    hits: list[str] = []

    if tokens & RULE_LOAN_CORE and tokens & RULE_OVERDUE:
        score += 4.0
        hits.append("loan+overdue")
    if tokens & RULE_LOAN_CORE and tokens & RULE_PROOF:
        score += 3.0
        hits.append("loan+proof_or_limit")
    if tokens & RULE_APPROVAL and (tokens & RULE_LOAN_CORE or {"linea", "financiamiento"} & tokens):
        score += 3.5
        hits.append("approval_or_disbursement")
    if tokens & RULE_COLLECTION:
        score += 3.5
        hits.append("collection_or_legal")
    if {"paga", "pagar", "pague"} & tokens and RULE_URGENT & tokens and tokens & RULE_LOAN_CORE:
        score += 2.0
        hits.append("urgent_payment")

    return score, hits


def exclusion_rule_adjustment(tokens: set[str]) -> tuple[float, list[str]]:
    # 排除场景修正。
    # 这部分不是为了识别贷款，而是为了压低容易误判的上下文，
    # 尤其是电信、公共事业、信用卡充值、普通账单这些场景。
    #
    # 设计思路是：
    # - 如果明显像电信/水电/话费类短信，就大幅扣分
    # - 如果更像信用卡、银行账户通知，就避免被当成现金贷
    # - 但对带有平台痕迹和贷后语境的“服务类”短信要留保护口
    score = 0.0
    hits: list[str] = []

    platform_hints = NORMALIZED_KEYWORDS["brand_hints"]

    telecom_hit_count = (
        len(tokens & EXCLUSION_TELECOM_CORE)
        + len(tokens & EXCLUSION_TELECOM_BRANDS)
        + len(tokens & EXCLUSION_UTILITY_TERMS)
    )
    has_loan_signal = bool(tokens & EXCLUSION_LOAN_CORE)
    has_collection_signal = bool(tokens & EXCLUSION_COLLECTION_CORE)
    has_finance_signal = bool(tokens & EXCLUSION_BANK_CORE)
    has_platform_hint = bool(tokens & platform_hints)
    protected_loan_service_context = bool(tokens & EXCLUSION_LOAN_APP_CONTEXT) and (has_platform_hint or bool({"credito"} & tokens))
    has_card_feedback_context = bool(tokens & EXCLUSION_CARD_FEEDBACK_CONTEXT)
    has_bank_transfer_context = bool(tokens & EXCLUSION_BANK_TRANSFER_CONTEXT)

    if telecom_hit_count >= 2 and not has_loan_signal and not has_collection_signal:
        score -= 4.0
        hits.append("telecom_or_utility_context")
    if {"tarjeta"} & tokens and {"recarga"} & tokens and not has_loan_signal:
        score -= 3.0
        hits.append("card_or_recharge_context")
    if (
        {"factura", "facturas", "servicio", "corte"} & tokens
        and not has_loan_signal
        and not has_collection_signal
        and not protected_loan_service_context
        and not (tokens & EXCLUSION_BANK_STATEMENT_CONTEXT and (tokens & {"banco", "tarjeta", "tc", "credito"}))
    ):
        score -= 2.5
        hits.append("billing_service_context")
    if {"credito"} & tokens and {"tarjeta"} & tokens and {"recarga"} & tokens and not has_loan_signal:
        score -= 2.0
        hits.append("credito_card_context")
    if {"credito"} & tokens and {"banco", "tarjeta", "cuenta"} & tokens and not has_loan_signal and not has_collection_signal:
        score += 1.0
        hits.append("banking_context")
    if has_finance_signal and not has_loan_signal and not has_collection_signal:
        score += 0.5
        hits.append("generic_finance_context")
    if {"tarjeta", "credito"} <= tokens and has_card_feedback_context and not (tokens & EXCLUSION_LOAN_HARD_CONTEXT):
        score -= 4.0
        hits.append("card_feedback_context")
    if has_bank_transfer_context and (tokens & EXCLUSION_BANK_CORE) and not has_loan_signal and not has_collection_signal:
        score -= 3.0
        hits.append("bank_transfer_context")

    return score, hits


def determine_finance_segment(tokens: set[str], result_label: str, exclusion_hits: list[str]) -> str:
    # 在总标签之外补一个更细的金融类别。
    # 这一步主要服务人工复核与后续分析，帮助区分：
    # - 贷款/放款
    # - 催收/法务
    # - 银行账户类通知
    # - OTP/安全验证
    # - 一般金融营销
    # - 已命中排除规则的非目标场景
    #
    # 注意这里输出的是“细分类别”，不是 score。
    # 也就是说，一条短信即使进入输出结果，仍然可以被进一步标成
    # collection_like / finance_otp_like / bank_or_finance_like 等不同子类。
    platform_hints = NORMALIZED_KEYWORDS["brand_hints"]

    if any(hit in exclusion_hits for hit in SEGMENT_EXCLUSION_HITS):
        return "excluded_non_finance"
    if tokens & SEGMENT_COLLECTION_CORE:
        return "collection_like"
    if tokens & SEGMENT_BANK_TRANSFER_CONTEXT and tokens & SEGMENT_BANK_CORE and not (tokens & SEGMENT_LOAN_CORE) and not (tokens & SEGMENT_COLLECTION_CORE):
        return "bank_or_finance_like"
    if tokens & SEGMENT_LOAN_SERVICE_CONTEXT and (tokens & SEGMENT_LOAN_CORE or tokens & platform_hints):
        return "loan_like"
    if (
        (tokens & SEGMENT_OTP_STRICT or ({"seguridad"} & tokens and {"registrarse", "registro"} & tokens))
        and not (tokens & SEGMENT_REPAYMENT_OR_COLLECTION_CONTEXT)
        and (tokens & SEGMENT_BANK_CORE or tokens & platform_hints)
    ):
        return "finance_otp_like"
    if (tokens & SEGMENT_MARKETING_CORE and (tokens & SEGMENT_BANK_CORE or tokens & platform_hints)) or (tokens & platform_hints and {"credito", "cuenta"} & tokens):
        return "loan_marketing_like"
    if tokens & SEGMENT_LOAN_CORE or result_label == "cash_loan_like":
        return "loan_like"
    if tokens & SEGMENT_BANK_CORE:
        return "bank_or_finance_like"
    return "other_finance_like"


def score_sms(text: str, loan_threshold: float = 6.0, min_output_score: float = 3.0) -> ClassificationResult:
    # 单条短信的总评分入口。
    # 这里会汇总：
    # - 强/中/品牌/噪音关键词命中
    # - 组合规则加分
    # - 排除场景修正
    # 最终输出统一的 ClassificationResult。
    #
    # 标签规则保持两层：
    # - score >= loan_threshold        -> cash_loan_like
    # - min_output_score <= score <   -> suspicious
    # 低于输出阈值的短信不会进入最终结果文件。
    normalized_text = normalize_sms_text(text)
    tokens = tokenize(text)

    strong_hits = sorted(tokens & NORMALIZED_KEYWORDS["strong"])
    medium_hits = sorted(tokens & NORMALIZED_KEYWORDS["medium"])
    brand_hits = sorted(tokens & NORMALIZED_KEYWORDS["brand_hints"])
    noise_hits = sorted(tokens & NORMALIZED_KEYWORDS["low_weight_noise"])

    score = 0.0
    score += len(strong_hits) * 3.0
    score += len(medium_hits) * 1.0
    score += len(brand_hits) * 2.0
    score -= len(noise_hits) * 0.35

    rule_score, rule_hits = rule_based_boost(tokens)
    score += rule_score
    exclusion_score, exclusion_hits = exclusion_rule_adjustment(tokens)
    score += exclusion_score

    if not (SCORE_REQUIRED_LOAN_SIGNAL & tokens) and not (SCORE_REQUIRED_COLLECTION_SIGNAL & tokens):
        score -= 1.5

    score = round(score, 2)

    if score >= loan_threshold:
        label = "cash_loan_like"
    elif score >= min_output_score:
        label = "suspicious"
    else:
        label = "ignored"

    finance_segment = determine_finance_segment(tokens, label, exclusion_hits)

    return ClassificationResult(
        normalized_text=normalized_text,
        label=label,
        finance_segment=finance_segment,
        score=score,
        strong_hits=strong_hits,
        medium_hits=medium_hits,
        brand_hits=brand_hits,
        noise_hits=noise_hits,
        rule_hits=rule_hits,
        exclusion_hits=exclusion_hits,
    )


def load_workbook_frames(input_file: str | Path) -> dict[str, pl.DataFrame]:
    # 使用 polars + calamine 一次性读取整个 workbook。
    # 这里单独拆成一个函数，是为了让“读 Excel”和“评分/输出”解耦：
    # - 便于后续替换读取实现
    # - 便于 collect_rows 和 build_filtered_rows_from_excel 复用
    # - 在大文件场景下比 openpyxl 逐行读取更快
    workbook = pl.read_excel(
        input_file,
        sheet_id=0,
        engine="calamine",
        infer_schema_length=1000,
        drop_empty_rows=True,
        drop_empty_cols=True,
    )
    if isinstance(workbook, pl.DataFrame):
        return {"Sheet1": workbook}
    return workbook


def collect_rows(input_file: str | Path) -> tuple[list[str], list[dict[str, object]]]:
    # 兼容旧流程的全量读入接口。
    # 会读取全部 sheet，整理出统一列名，并把每行转换成 row_dict。
    # 当前主流程已优先使用 build_filtered_rows_from_excel 做边读边评，
    # 这里保留主要是为了调试、对比和兼容历史调用方式。
    print(f"读取文件: {input_file}")
    workbook_frames = load_workbook_frames(input_file)

    all_columns: list[str] = []
    all_rows: list[dict[str, object]] = []
    total_sheets = len(workbook_frames)

    for sheet_idx, (sheet_name, df) in enumerate(workbook_frames.items(), start=1):
        print(f"读取 sheet {sheet_idx}/{total_sheets}: {sheet_name}")

        headers = [str(column).strip() for column in df.columns]
        if "content" not in headers:
            print(f"  跳过 sheet '{sheet_name}'，缺少列: ['content']")
            continue

        for header in headers:
            if header not in all_columns:
                all_columns.append(header)

        sheet_rows = 0
        for row_dict in df.iter_rows(named=True):
            row_dict["_sheet_name"] = sheet_name
            all_rows.append(row_dict)
            sheet_rows += 1

        print(f"  读取 {sheet_rows} 行")

    print(f"\n总共读取 {len(all_rows)} 行")
    return all_columns, all_rows


def build_filtered_rows(
    rows: list[dict[str, object]],
    loan_threshold: float,
    min_output_score: float,
) -> list[tuple[dict[str, object], ClassificationResult]]:
    # 对已经读入内存的短信逐条评分，只保留达到输出阈值的记录。
    # 这是旧版“先读全量，再统一评分”的处理路径。
    filtered_rows: list[tuple[dict[str, object], ClassificationResult]] = []
    total_rows = len(rows)
    score_cache: dict[str, ClassificationResult] = {}
    print("开始评分...")

    for idx, row_dict in enumerate(rows):
        if idx % 50000 == 0:
            print(f"已处理 {idx}/{total_rows} 行，当前保留 {len(filtered_rows)} 行...")

        content = str(row_dict.get("content") or "")
        result = score_cache.get(content)
        if result is None:
            result = score_sms(content, loan_threshold=loan_threshold, min_output_score=min_output_score)
            score_cache[content] = result
        if result.score >= min_output_score:
            filtered_rows.append((row_dict, result))

    print(f"评分完成，保留 {len(filtered_rows)} 行。")
    return filtered_rows


def build_filtered_rows_from_excel(
    input_file: str | Path,
    loan_threshold: float,
    min_output_score: float,
) -> tuple[list[str], list[tuple[dict[str, object], ClassificationResult]]]:
    # 当前推荐的主处理路径。
    # 直接从 Excel 边读取边评分边保留，避免先构造超大的 all_rows 列表，
    # 更适合海量短信场景下的速度和内存控制。
    print(f"读取文件: {input_file}")
    workbook_frames = load_workbook_frames(input_file)

    all_columns: list[str] = []
    filtered_rows: list[tuple[dict[str, object], ClassificationResult]] = []
    score_cache: dict[str, ClassificationResult] = {}
    total_sheets = len(workbook_frames)
    total_rows = 0

    print("开始评分...")
    for sheet_idx, (sheet_name, df) in enumerate(workbook_frames.items(), start=1):
        print(f"读取 sheet {sheet_idx}/{total_sheets}: {sheet_name}")

        headers = [str(column).strip() for column in df.columns]
        if "content" not in headers:
            print(f"  跳过 sheet '{sheet_name}'，缺少列: ['content']")
            continue

        for header in headers:
            if header not in all_columns:
                all_columns.append(header)

        sheet_rows = 0
        for row_dict in df.iter_rows(named=True):
            total_rows += 1
            sheet_rows += 1

            if total_rows % 50000 == 0:
                print(f"已处理 {total_rows} 行，当前保留 {len(filtered_rows)} 行，唯一内容缓存 {len(score_cache)} 条...")

            row_dict["_sheet_name"] = sheet_name
            content = str(row_dict.get("content") or "")
            result = score_cache.get(content)
            if result is None:
                result = score_sms(content, loan_threshold=loan_threshold, min_output_score=min_output_score)
                score_cache[content] = result
            if result.score >= min_output_score:
                filtered_rows.append((row_dict, result))

        print(f"  读取 {sheet_rows} 行")

    print(f"\n评分完成，共处理 {total_rows} 行，保留 {len(filtered_rows)} 行。")
    return all_columns, filtered_rows


def get_score_band(score: float) -> tuple[str, str]:
    # 将连续分数映射到便于人工理解的分段标签。
    if score < 3:
        return "S0", "低于保留线"
    if score < 5:
        return "S1", "弱金融相关"
    if score < 8:
        return "S2", "中等相关"
    if score < 12:
        return "S3", "高相关"
    return "S4", "极高相关"


def get_priority_and_bucket(finance_segment: str, score: float) -> tuple[str, str, str]:
    # 在 score 与 finance_segment 的基础上，进一步输出业务优先级和业务桶。
    # 这是 v3.1 相比基础识别脚本更偏业务使用的一层封装。
    if finance_segment == "excluded_non_finance" or score < 3:
        return "D", "排除或低优先级", "主要是运营商/公用事业/明显非目标短信，或证据太弱"
    if finance_segment in {"loan_like", "collection_like"} and score >= 8:
        return "A", "核心贷款催收", "贷款/催收证据较强，适合作为最高优先级池"
    if finance_segment in {"loan_like", "collection_like"} and score < 5:
        return "B-", "边界贷款样本", "类别偏贷款/催收，但总证据偏弱，建议单独复核"
    if finance_segment == "loan_marketing_like" and score < 5:
        return "C+", "低优先级贷款营销", "明显属于贷款营销，但证据较弱，不建议与排除类混放"
    if (
        finance_segment in {"loan_like", "collection_like"} and 5 <= score < 8
    ) or (
        finance_segment == "loan_marketing_like" and score >= 5
    ):
        return "B", "贷款相关", "与贷款业务高度相关，但不一定是强催收"
    if finance_segment in {"bank_or_finance_like", "finance_otp_like", "other_finance_like"} and score >= 3:
        return "C", "银行/金融服务", "银行、验证码、普通金融服务类，保留但优先级低于贷款主池"
    return "D", "排除或低优先级", "不满足主要金融分层口径"


def save_output_v3_1(
    output_file: str | Path,
    original_columns: list[str],
    filtered_rows,
    output_mode: str,
) -> None:
    # 生成 v3.1 的输出工作簿。
    # 这个版本不再输出全部识别中间字段，而是只保留人工复核最常用的：
    # - 来源 sheet
    # - 中文细分类别
    # - 优先级
    # - 业务桶
    wb = Workbook()
    ws = wb.active
    ws.title = "all_sheets_scored_v3_1"

    if output_mode == "compact":
        selected_columns = [col for col in COMPACT_INPUT_COLUMNS if col in original_columns]
    else:
        selected_columns = original_columns

    ws.append(selected_columns + RESULT_COLUMNS_V3_1)

    for row_dict, result in filtered_rows:
        # Keep these derived values available in code so the output columns can be
        # restored later without reworking the scoring pipeline.
        _score_band, _score_band_zh = get_score_band(result.score)
        priority_level, business_bucket, _ = get_priority_and_bucket(
            result.finance_segment,
            result.score,
        )
        finance_segment_zh = FINANCE_SEGMENT_ZH.get(result.finance_segment, "")
        original_values = [row_dict.get(col) for col in selected_columns]
        _rule_hits = ",".join(result.rule_hits)
        _exclusion_hits = ",".join(result.exclusion_hits)

        ws.append(
            original_values + [
                row_dict.get("_sheet_name", ""),
                finance_segment_zh,
                priority_level,
                business_bucket,
            ]
        )

    for column_cells in ws.columns:
        column_letter = column_cells[0].column_letter
        header = str(column_cells[0].value or "")
        if header == "content":
            width = 90
        elif header in {"finance_segment_zh", "business_bucket"}:
            width = 24
        else:
            width = min(max(len(header) + 4, 12), 28)
        ws.column_dimensions[column_letter].width = width

    wb.save(output_file)
    print(f"输出完成: {output_file}")


def process_all_sheets_v3_1(
    input_file: str | Path,
    output_file: str | Path,
    loan_threshold: float = 6.0,
    min_output_score: float = 3.0,
    output_mode: str = "compact",
) -> None:
    # v3.1 主流程入口。
    # 默认走“边读边评边保留”的快速路径，再统一导出结果。
    original_columns, filtered_rows = build_filtered_rows_from_excel(
        input_file,
        loan_threshold,
        min_output_score,
    )
    save_output_v3_1(output_file, original_columns, filtered_rows, output_mode)


def parse_args(argv: list[str]) -> argparse.Namespace:
    # 命令行参数解析。
    parser = argparse.ArgumentParser(
        description="基于 v2 规则并追加业务分层字段的金融短信筛选脚本 v3.1。"
    )
    parser.add_argument("input_file", nargs="?", default="input.xlsx")
    parser.add_argument("output_file", nargs="?", default="output_scored_all_sheets_v3_1.xlsx")
    parser.add_argument("--loan-threshold", type=float, default=6.0)
    parser.add_argument("--min-output-score", type=float, default=3.0)
    parser.add_argument(
        "--output-mode",
        choices=["compact", "full"],
        default="compact",
        help="compact 仅保留前五列加结果列；full 保留全部原始列加结果列。",
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> None:
    # 脚本入口。
    args = parse_args(argv or sys.argv[1:])
    process_all_sheets_v3_1(
        input_file=args.input_file,
        output_file=args.output_file,
        loan_threshold=args.loan_threshold,
        min_output_score=args.min_output_score,
        output_mode=args.output_mode,
    )


if __name__ == "__main__":
    main()
