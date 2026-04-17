from __future__ import annotations

import argparse
import contextlib
import io
import json
import math
import os
import re
import sys
import time
import warnings
from concurrent.futures import ProcessPoolExecutor, as_completed
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import numpy as np
import pandas as pd
import polars as pl
from sklearn.cluster import KMeans
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

os.environ.setdefault("OMP_NUM_THREADS", "1")
warnings.filterwarnings("ignore", category=FutureWarning)
warnings.filterwarnings("ignore", category=UserWarning)

CURRENT_DIR = Path(__file__).resolve().parent
if str(CURRENT_DIR) not in sys.path:
    sys.path.insert(0, str(CURRENT_DIR))

from tools.commontools import woe2, woe3, woe_more

# 分箱规则摘要：
# 1. 非 MA 特征先看去重值个数，去重 < 20 直接按原值展示。
# 2. 非 MA 的低值区分三类处理：离散整数保留单值，0-1 比例型按动态步长分箱，
#    1-10 连续浮点按小步长分箱；如果 0-10 单值箱全部 < 1%，则合并成一个低值箱。
# 3. 非 MA 的 >10 区间按长尾/普通分箱处理，最后再做连续稀疏尾箱合并。
# 4. MA 特征维持评分卡风格分箱，最大箱数不超过 20；当前只把步长的参考区间改为 P10-P90。

DEFAULT_SHEET_OUTPUT_CONFIG = {
    "woe_group_all": True,
    "APP": True,
    "APP_LX": True,
    "AUP": True,
    "DS": True,
    "DX": True,
    "DX_BANK": True,
    "DX_LS": True,
    "FD": True,
    "GL": True,
    "GZ": True,
    "HC_DX": True,
    "INX": True,
    "JC": True,
    "LOW_DIGIT_MISC": True,
    "MA": True,
    "MXG_GL": True,
    "RL_FLY": True,
    "RL_LOAN": True,
    "SB": True,
    "SJH": True,
    "TH": True,
    "TXL": True,
    "WIFI": True,
    "XW": True,
    "XYK": True,
    "YQ": True,
    "YYS": True,
}

SEGMENT_ORDER = ["主包首贷", "主包复贷", "多贷首贷", "多贷复贷"]
CONFIG_DIR = CURRENT_DIR / "config"
FEATURE_DESC_PATH = CONFIG_DIR / "模型分特征码汇总.xlsx"
DISPLAY_CONFIG_PATH = CONFIG_DIR / "feature_display_config.json"


def now_ts() -> float:
    return time.perf_counter()


def format_elapsed(start_ts: float) -> str:
    return f"{time.perf_counter() - start_ts:.2f}s"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="读取原始数据，生成 WOE 分箱结果和特征分组分析报表。"
    )
    parser.add_argument("input_file", help="输入文件路径，支持 xlsx/xls/csv/parquet")
    parser.add_argument(
        "status_source_pos",
        nargs="?",
        type=int,
        choices=[0, 1, 2, 3],
        help="可选位置参数，兼容旧调用：0->settleStatus, 1->settleStatus_1, 2->settleStatus_2, 3->settleStatus_3",
    )
    parser.add_argument(
        "--output-file",
        default=None,
        help="输出 Excel 路径，默认在输入文件同目录生成 *_feature_analysis.xlsx",
    )
    parser.add_argument(
        "--feature-prefixes",
        default=None,
        help="可选，按前缀筛选特征，多个前缀用逗号分隔，例如 MA,DX,YQ",
    )
    parser.add_argument(
        "--status-source",
        type=int,
        choices=[0, 1, 2, 3],
        default=0,
        help="起始状态列选择：0->settleStatus, 1->settleStatus_1, 2->settleStatus_2, 3->settleStatus_3",
    )
    parser.add_argument(
        "--sheet-config-file",
        default=None,
        help="可选，sheet 输出开关配置 JSON 路径，默认使用 config/feature_sheet_output_config.json",
    )
    parser.add_argument(
        "--display-config-file",
        default=None,
        help="可选，特征展示配置 JSON 路径，默认使用 config/feature_display_config.json",
    )
    parser.add_argument(
        "--feature-desc-file",
        default=None,
        help="可选，特征中文解释 Excel 路径，默认使用 config/模型分特征码汇总.xlsx",
    )
    args = parser.parse_args()
    if args.status_source_pos is not None:
        args.status_source = args.status_source_pos
    return args


def normalize_column_lookup(columns: Iterable[str]) -> Dict[str, str]:
    lookup: Dict[str, str] = {}
    for column in columns:
        if column is None:
            continue
        lookup[str(column).strip().lower()] = str(column)
    return lookup


def find_column(columns: Iterable[str], *candidates: str) -> Optional[str]:
    lookup = normalize_column_lookup(columns)
    for candidate in candidates:
        found = lookup.get(candidate.strip().lower())
        if found:
            return found
    return None


def derive_output_path(input_path: Path, output_file: Optional[str]) -> Path:
    if output_file:
        target = Path(output_file)
        target.parent.mkdir(parents=True, exist_ok=True)
        return target
    output_dir = CURRENT_DIR / "feature_analysis_output" / input_path.stem
    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir / "feature_analysis.xlsx"


def derive_segment_output_path(base_output_path: Path, segment_name: str) -> Path:
    return base_output_path.with_name(
        f"feature_analysis_{segment_name}{base_output_path.suffix}"
    )


def load_sheet_output_config(config_path: Path) -> Dict[str, bool]:
    config_path.parent.mkdir(parents=True, exist_ok=True)
    config_data = dict(DEFAULT_SHEET_OUTPUT_CONFIG)
    if config_path.exists():
        with config_path.open("r", encoding="utf-8") as file:
            raw_text = file.read()
        cleaned_lines = []
        for line in raw_text.splitlines():
            stripped = line.lstrip()
            if stripped.startswith("//"):
                comment_match = re.match(r'^\s*//\s*"([^"]+)"\s*:', line)
                if comment_match:
                    config_data[comment_match.group(1)] = False
                continue
            cleaned_lines.append(line)
        loaded = json.loads("\n".join(cleaned_lines))
        if isinstance(loaded, dict):
            for key, value in loaded.items():
                config_data[str(key)] = bool(value)
        return config_data
    with config_path.open("w", encoding="utf-8") as file:
        json.dump(config_data, file, ensure_ascii=False, indent=2)
    return config_data


def sync_new_sheet_config(
    config_path: Path,
    sheet_output_config: Dict[str, bool],
    detected_sheet_names: Iterable[str],
) -> Tuple[Dict[str, bool], List[str]]:
    updated = dict(sheet_output_config)
    new_sheets: List[str] = []
    has_change = False
    for sheet_name in detected_sheet_names:
        name = str(sheet_name).strip()
        if not name:
            continue
        if name not in updated:
            updated[name] = True
            has_change = True
            new_sheets.append(name)
    if has_change:
        if config_path.exists():
            raw_text = config_path.read_text(encoding="utf-8")
            trimmed = raw_text.rstrip()
            close_idx = trimmed.rfind("}")
            if close_idx != -1:
                prefix = trimmed[:close_idx].rstrip()
                if prefix.endswith("{"):
                    joiner = "\n"
                elif prefix.endswith(","):
                    joiner = "\n"
                else:
                    joiner = ",\n"
                insert_lines = ",\n".join([f'  "{name}": true' for name in sorted(new_sheets, key=natural_sort_key)])
                new_text = f"{prefix}{joiner}{insert_lines}\n{trimmed[close_idx:]}"
                if raw_text.endswith("\n"):
                    new_text += "\n"
                config_path.write_text(new_text, encoding="utf-8")
            else:
                with config_path.open("w", encoding="utf-8") as file:
                    json.dump(updated, file, ensure_ascii=False, indent=2)
        else:
            with config_path.open("w", encoding="utf-8") as file:
                json.dump(updated, file, ensure_ascii=False, indent=2)
    return updated, sorted(new_sheets, key=natural_sort_key)


def split_dataframe_by_segment(df: pd.DataFrame) -> Dict[str, pd.DataFrame]:
    appssid_col = find_column(df.columns, "appssid")
    loancount_col = find_column(df.columns, "loancount")
    if not appssid_col or not loancount_col:
        raise ValueError("数据缺少 appSsid 或 LOANCOUNT 字段，无法按主包/多贷、首贷/复贷分层。")

    appssid_text = df[appssid_col].fillna("").astype(str).str.strip()
    loancount = pd.to_numeric(df[loancount_col], errors="coerce")
    second_last_is_zero = appssid_text.str.len().ge(2) & appssid_text.str[-2].eq("0")
    package_type = np.where((appssid_text.str.len() > 3) | second_last_is_zero, "多贷", "主包")
    loan_type = np.where(loancount == 1, "首贷", np.where(loancount > 1, "复贷", ""))
    segment_series = pd.Series(package_type, index=df.index).astype(str) + pd.Series(loan_type, index=df.index).astype(str)

    segments: Dict[str, pd.DataFrame] = {}
    for segment_name in SEGMENT_ORDER:
        mask = segment_series == segment_name
        if mask.any():
            segments[segment_name] = df.loc[mask].copy()
    return segments


def filter_sheet_tables_by_config(
    grouped_feature_tables: Dict[str, List[Tuple[str, pd.DataFrame]]],
    sheet_output_config: Dict[str, bool],
) -> Dict[str, List[Tuple[str, pd.DataFrame]]]:
    filtered: Dict[str, List[Tuple[str, pd.DataFrame]]] = {}
    for sheet_name, feature_tables in grouped_feature_tables.items():
        if sheet_output_config.get(sheet_name, True):
            filtered[sheet_name] = feature_tables
    return filtered


def load_excel_with_polars(input_path: Path) -> pd.DataFrame:
    workbook = pl.read_excel(
        input_path,
        sheet_id=0,
        engine="calamine",
        infer_schema_length=1000,
        drop_empty_rows=True,
        drop_empty_cols=True,
    )
    if isinstance(workbook, dict):
        frames: List[pd.DataFrame] = []
        for sheet_name, frame in workbook.items():
            pdf = frame.to_pandas()
            pdf["_sheet_name"] = sheet_name
            frames.append(pdf)
        if not frames:
            return pd.DataFrame()
        return pd.concat(frames, ignore_index=True)
    pdf = workbook.to_pandas()
    pdf["_sheet_name"] = "sheet1"
    return pdf


def load_input_dataframe(input_file: str) -> pd.DataFrame:
    input_path = Path(input_file)
    suffix = input_path.suffix.lower()
    if suffix in {".xlsx", ".xlsm", ".xls"}:
        return load_excel_with_polars(input_path)
    if suffix == ".csv":
        for encoding in ("utf-8", "utf-8-sig", "gbk", "latin1"):
            try:
                return pd.read_csv(input_path, encoding=encoding)
            except Exception:
                continue
        raise ValueError(f"无法识别 CSV 编码: {input_file}")
    if suffix == ".parquet":
        return pd.read_parquet(input_path)
    raise ValueError(f"暂不支持的文件类型: {suffix}")


def load_feature_description_map(desc_path: Path) -> Dict[str, str]:
    if not desc_path.exists():
        return {}
    try:
        workbook = pl.read_excel(
            desc_path,
            sheet_id=0,
            engine="calamine",
            infer_schema_length=1000,
            drop_empty_rows=True,
            drop_empty_cols=True,
        )
        frame = next(iter(workbook.values())) if isinstance(workbook, dict) else workbook
        pdf = frame.to_pandas()
    except Exception:
        return {}

    code_col = find_column(pdf.columns, "特征码")
    meaning_col = find_column(pdf.columns, "含义")
    if not code_col or not meaning_col:
        return {}

    mapping: Dict[str, str] = {}
    for _, row in pdf[[code_col, meaning_col]].iterrows():
        code = str(row[code_col]).strip() if pd.notna(row[code_col]) else ""
        meaning = str(row[meaning_col]).strip() if pd.notna(row[meaning_col]) else ""
        if code and meaning:
            mapping[code] = meaning
    return mapping


def load_feature_display_config(config_path: Path) -> bool:
    config_path.parent.mkdir(parents=True, exist_ok=True)
    default_config = {"show_feature_desc": True}
    if not config_path.exists():
        with config_path.open("w", encoding="utf-8") as file:
            json.dump(default_config, file, ensure_ascii=False, indent=2)
        return True
    try:
        with config_path.open("r", encoding="utf-8") as file:
            config = json.load(file)
        return bool(config.get("show_feature_desc", True))
    except Exception:
        return True


def display_feature_name(
    feature_name: str,
    feature_desc_map: Dict[str, str],
    show_feature_desc: bool,
) -> str:
    if not show_feature_desc:
        return str(feature_name)
    meaning = feature_desc_map.get(str(feature_name).strip())
    if not meaning:
        return str(feature_name)
    return f"{feature_name}/{meaning}"


def normalize_settle_status(series: pd.Series) -> pd.Series:
    status = pd.to_numeric(series, errors="coerce")
    status = status.replace({4: 2, 5: 3})
    return status


def get_status_source_column(columns: Iterable[str], status_source: int) -> str:
    mapping = {
        0: ("settlestatus", "settleStatus"),
        1: ("settlestatus_1", "settleStatus_1"),
        2: ("settlestatus_2", "settleStatus_2"),
        3: ("settlestatus_3", "settleStatus_3"),
    }
    candidates = mapping.get(status_source, mapping[0])
    selected = find_column(columns, *candidates)
    if not selected:
        raise ValueError(
            f"未找到 status_source={status_source} 对应的字段，期望列名之一: {', '.join(candidates)}"
        )
    return selected


def ensure_target_columns(df: pd.DataFrame, status_source: int = 0) -> pd.DataFrame:
    result = df.copy()
    settle_col = get_status_source_column(result.columns, status_source=status_source)
    result["settleStatus_normalized"] = normalize_settle_status(result[settle_col])
    result["target"] = (pd.to_numeric(result["settleStatus_normalized"], errors="coerce") == 0).astype(int)
    result = result[result["target"].isin([0, 1])].copy()
    result = result[
        result["settleStatus_normalized"].isna()
        | result["settleStatus_normalized"].isin([0, 1, 2, 3])
    ].copy()
    return result


def clean_feature_frame(df: pd.DataFrame, feature_columns: List[str]) -> pd.DataFrame:
    cleaned = df[feature_columns + ["target"]].copy()
    for column in feature_columns:
        cleaned[column] = pd.to_numeric(cleaned[column], errors="coerce")
        cleaned[column] = cleaned[column].fillna(-999)
    return cleaned


def compute_single_feature_woe(
    feature_name: str,
    feature_values: List[float],
    target_values: List[int],
) -> Tuple[str, List[List[object]], List[float]]:
    item_df = pd.DataFrame({feature_name: feature_values, "target": target_values})
    unique_count = item_df[feature_name].nunique(dropna=True)
    if unique_count <= 1:
        return feature_name, [], []

    local_df_woe = pd.DataFrame(index=item_df.index)
    local_dvars: Dict[str, List[List[float]]] = {}
    local_scores: Dict[str, List[List[object]]] = {}
    x_values = item_df[feature_name].to_numpy()
    x_pred = [[value] for value in x_values]

    if unique_count > 4:
        with contextlib.redirect_stdout(io.StringIO()):
            result_flag, local_dvars, local_scores = woe_more(
                feature_name, item_df, local_df_woe, local_dvars, local_scores
            )
        if result_flag == 1:
            feature_woe_values = local_df_woe[feature_name].tolist() if feature_name in local_df_woe else []
            return feature_name, local_scores.get(feature_name, []), feature_woe_values

    if unique_count > 2:
        for random_state in range(10):
            y_pred = KMeans(n_clusters=3, random_state=random_state).fit_predict(x_pred)
            judge: List[List[object]] = []
            for cluster_id in sorted(set(y_pred)):
                cluster_mask = y_pred == cluster_id
                cluster_target = item_df.loc[cluster_mask, "target"]
                bad_rate = cluster_target.mean()
                cluster_values = x_values[cluster_mask]
                judge.append(
                    [
                        cluster_id,
                        bad_rate,
                        [float(cluster_values.min()), float(cluster_values.max())],
                    ]
                )
            judge = sorted(judge, key=lambda x: x[2])
            monotonic = (judge[1][1] - judge[0][1]) * (judge[2][1] - judge[1][1]) >= 0
            if not monotonic:
                continue

            with contextlib.redirect_stdout(io.StringIO()):
                woe_values, local_dvars, local_scores = woe3(
                    y_pred, feature_name, item_df, local_df_woe, local_dvars, local_scores
                )
            local_scores[feature_name] = []
            for row in judge:
                cluster_score = None
                for cluster_id, woe_value in sorted(woe_values, key=lambda x: x[0]):
                    if cluster_id == row[0]:
                        cluster_score = woe_value
                        break
                local_scores[feature_name].append([row[2], cluster_score])
            feature_woe_values = local_df_woe[feature_name].tolist() if feature_name in local_df_woe else []
            return feature_name, local_scores.get(feature_name, []), feature_woe_values

    with contextlib.redirect_stdout(io.StringIO()):
        _, local_dvars, local_scores = woe2(x_pred, feature_name, item_df, local_df_woe, local_dvars, local_scores)
    feature_woe_values = local_df_woe[feature_name].tolist() if feature_name in local_df_woe else []
    return feature_name, local_scores.get(feature_name, []), feature_woe_values


def build_tools_woe_bins(df: pd.DataFrame, feature_columns: List[str]) -> Tuple[pd.DataFrame, pd.DataFrame]:
    stage_start = now_ts()
    print(f"[woe] start feature_count={len(feature_columns)}")
    source = clean_feature_frame(df, feature_columns)
    print(f"[woe] clean_feature_frame_done elapsed={format_elapsed(stage_start)}")
    scores: Dict[str, List[List[object]]] = {}
    df_woe_columns: Dict[str, List[float]] = {}
    max_workers = max(1, min((os.cpu_count() or 2) - 1, 8))
    feature_payloads = [(item, source[item].tolist(), source["target"].tolist()) for item in feature_columns]
    print(f"[woe] process_pool_workers={max_workers} payload_count={len(feature_payloads)}")

    def handle_result(
        feature_name: str,
        feature_scores: List[List[object]],
        feature_woe_values: List[float],
    ) -> None:
        if feature_scores:
            scores[feature_name] = feature_scores
        if feature_woe_values:
            df_woe_columns[feature_name] = feature_woe_values

    total = len(feature_payloads)
    done = 0
    try:
        with ProcessPoolExecutor(max_workers=max_workers) as executor:
            futures = [executor.submit(compute_single_feature_woe, *payload) for payload in feature_payloads]
            for future in as_completed(futures):
                feature_name, feature_scores, feature_woe_values = future.result()
                done += 1
                if done % 100 == 0 or done == total:
                    print(f"[woe] progress={done}/{total} elapsed={format_elapsed(stage_start)}")
                handle_result(feature_name, feature_scores, feature_woe_values)
    except (PermissionError, OSError) as error:
        print(f"[woe] process_pool_unavailable={error} -> fallback=single_process")
        for payload in feature_payloads:
            feature_name, feature_scores, feature_woe_values = compute_single_feature_woe(*payload)
            done += 1
            if done % 100 == 0 or done == total:
                print(f"[woe] progress={done}/{total} elapsed={format_elapsed(stage_start)}")
            handle_result(feature_name, feature_scores, feature_woe_values)

    df_woe = pd.DataFrame(df_woe_columns, index=source.index) if df_woe_columns else pd.DataFrame(index=source.index)
    woe_group_all = build_woe_group_all_from_scores(source, scores)
    print(
        f"[woe] finished score_feature_count={len(scores)} "
        f"woe_group_rows={len(woe_group_all)} elapsed={format_elapsed(stage_start)}"
    )
    return woe_group_all, df_woe


def build_woe_group_all_from_scores(
    df: pd.DataFrame, scores: Dict[str, List[List[object]]]
) -> pd.DataFrame:
    rows: List[Dict[str, object]] = []
    total_bad = float(df["target"].sum())
    total_good = float((1 - df["target"]).sum())
    eps = 1e-9

    for feature_name, score_rules in scores.items():
        feature_df = df[[feature_name, "target"]].dropna().copy()
        if feature_df.empty:
            continue
        for value_range, woe_value in score_rules:
            lower = float(value_range[0])
            upper = float(value_range[1])
            mask = (feature_df[feature_name] >= lower) & (feature_df[feature_name] <= upper)
            bucket = feature_df.loc[mask]
            total_count = int(len(bucket))
            if total_count == 0:
                continue
            bad_count = int(bucket["target"].sum())
            good_count = total_count - bad_count
            bad_rate = bad_count / total_count if total_count else 0
            woe_final = math.log(
                ((bad_count + eps) / (good_count + eps)) / ((total_bad + eps) / (total_good + eps))
            )
            iv = (
                ((bad_count + eps) / (total_bad + eps) - (good_count + eps) / (total_good + eps))
                * woe_final
            )
            rows.append(
                {
                    "特征名": feature_name,
                    "分箱结果": f"({lower}, {upper}]",
                    "样本数": total_count,
                    "坏样本数": bad_count,
                    "好样本数": good_count,
                    "坏样本率": bad_rate,
                    "WOE": woe_value if woe_value is not None else woe_final,
                    "IV": iv,
                }
            )

    if not rows:
        return pd.DataFrame(
            columns=["特征名", "分箱结果", "样本数", "坏样本数", "好样本数", "坏样本率", "样本占比", "WOE", "IV", "IV总计", "KS整体"]
        )
    result = pd.DataFrame(rows)
    feature_total = result.groupby("特征名")["样本数"].transform("sum").replace(0, np.nan)
    result["样本占比"] = result["样本数"] / feature_total
    result["坏样本占比"] = result["坏样本数"] / (total_bad + eps)
    result["好样本占比"] = result["好样本数"] / (total_good + eps)

    result["区间下界"] = result["分箱结果"].astype(str).str.extract(r"[\(\[]\s*([^,]+),", expand=False)
    result["区间下界"] = pd.to_numeric(result["区间下界"], errors="coerce")

    result["IV总计"] = result.groupby("特征名")["IV"].transform("sum")

    ks_values: Dict[str, float] = {}
    for feature_name, feature_df in result.groupby("特征名", sort=False):
        ordered = feature_df.sort_values(["区间下界", "分箱结果"], kind="stable").copy()
        ordered["累计坏样本占比"] = ordered["坏样本占比"].cumsum()
        ordered["累计好样本占比"] = ordered["好样本占比"].cumsum()
        ks_values[feature_name] = float(
            (ordered["累计坏样本占比"] - ordered["累计好样本占比"]).abs().max()
        )
    result["KS整体"] = result["特征名"].map(ks_values)

    result = result.drop(columns=["坏样本占比", "好样本占比", "区间下界"])
    result = result[
        ["特征名", "分箱结果", "样本数", "坏样本数", "好样本数", "坏样本率", "样本占比", "WOE", "IV", "IV总计", "KS整体"]
    ]
    return result


def select_feature_columns(df: pd.DataFrame) -> List[str]:
    excluded = {
        "target",
        "settlestatus",
        "settlestatus_1",
        "settlestatus_2",
        "settlestatus_3",
        "settlestatus_normalized",
        "_sheet_name",
    }
    feature_columns: List[str] = []
    for column in df.columns:
        column_lower = str(column).strip().lower()
        if column_lower in excluded:
            continue
        if str(column).strip().upper() == "BUILD_NAME":
            continue
        numeric_series = pd.to_numeric(df[column], errors="coerce")
        null_ratio = numeric_series.isna().mean()
        if null_ratio > 0.9:
            continue
        numeric_ratio = numeric_series.notna().mean()
        unique_count = numeric_series.nunique(dropna=True)
        if unique_count <= 1:
            continue
        if numeric_ratio >= 0.6:
            feature_columns.append(column)
    return sorted(feature_columns)


def is_ma_like_feature(feature_name: str) -> bool:
    feature_upper = str(feature_name).strip().upper()
    return ((feature_upper.startswith("MA") and not feature_upper.startswith("MAIN"))
            or feature_upper.startswith("MODEL"))


def classify_feature_group(feature_name: str) -> str:
    if is_ma_like_feature(feature_name):
        return "MA"
    digits = re.findall(r"\d", feature_name)
    if len(digits) <= 1:
        return "LOW_DIGIT_MISC"
    match = re.match(r"([^0-9]+)", feature_name)
    if match:
        return match.group(1).strip("_").upper()
    return "OTHER"


def natural_sort_key(value: str) -> List[object]:
    parts = re.split(r"(\d+)", str(value))
    key: List[object] = []
    for part in parts:
        if not part:
            continue
        if part.isdigit():
            key.append(int(part))
        else:
            key.append(part.upper())
    return key


def filter_features_by_prefix(feature_columns: List[str], prefixes_text: Optional[str]) -> List[str]:
    if not prefixes_text:
        return feature_columns
    prefixes = [item.strip().upper() for item in prefixes_text.split(",") if item.strip()]
    if not prefixes:
        return feature_columns
    filtered: List[str] = []
    for feature_name in feature_columns:
        feature_upper = feature_name.upper()
        group_name = classify_feature_group(feature_name)
        if any(feature_upper.startswith(prefix) or group_name == prefix for prefix in prefixes):
            filtered.append(feature_name)
    return filtered


def pretty_step(raw_step: float) -> float:
    if raw_step <= 0 or math.isnan(raw_step):
        return 1.0
    magnitude = 10 ** math.floor(math.log10(raw_step))
    normalized = raw_step / magnitude
    if normalized <= 1:
        nice = 1
    elif normalized <= 2:
        nice = 2
    elif normalized <= 5:
        nice = 5
    else:
        nice = 10
    return nice * magnitude


def pretty_ma_step(raw_step: float) -> float:
    if raw_step <= 0 or math.isnan(raw_step):
        return 5.0
    rounded = round(raw_step / 5.0) * 5.0
    return max(5.0, rounded)


def build_ma_edges(series: pd.Series, max_bins: int = 20) -> List[float]:
    numeric = pd.to_numeric(series, errors="coerce").dropna()
    if numeric.empty:
        return []
    min_value = float(numeric.min())
    max_value = float(numeric.max())
    lower_ref = float(numeric.quantile(0.10))
    upper_ref = float(numeric.quantile(0.90))
    if upper_ref <= lower_ref:
        lower_ref = min_value
        upper_ref = max_value
    raw_step = (upper_ref - lower_ref) / max_bins if max_bins else 1
    step = max(pretty_ma_step(raw_step), 5)
    start = math.floor(min_value / step) * step
    while True:
        edges = []
        current = start
        upper_limit = max_value + step
        while current <= upper_limit:
            edges.append(round(current, 10))
            current += step
        if len(edges) - 1 <= max_bins:
            return edges
        next_step = pretty_ma_step(step * 1.2)
        if next_step <= step:
            next_step = step + 5
        step = next_step


def build_regular_edges(series: pd.Series) -> List[float]:
    numeric = pd.to_numeric(series, errors="coerce").dropna()
    if numeric.empty:
        return []
    min_value = float(numeric.min())
    max_value = float(numeric.max())
    robust_upper = max_value
    if len(numeric) >= 20:
        p95 = float(numeric.quantile(0.95))
        if p95 > min_value and max_value > p95 * 1.5:
            robust_upper = p95

    raw_step = (robust_upper - min_value) / 20
    step = max(pretty_step(raw_step), 1)
    start = math.floor(min_value / step) * step
    if start < min_value:
        start += step
    end = math.ceil(robust_upper / step) * step
    edges: List[float] = []
    current = start
    if current > min_value:
        edges.append(round(min_value, 10))
    while current <= end + step:
        edges.append(round(current, 10))
        current += step
    return sorted(set(edges))


def is_long_tail_distribution(series: pd.Series) -> bool:
    numeric = pd.to_numeric(series, errors="coerce").dropna()
    if len(numeric) < 20:
        return False
    p50 = float(numeric.quantile(0.50))
    p75 = float(numeric.quantile(0.75))
    p95 = float(numeric.quantile(0.95))
    max_value = float(numeric.max())
    if p50 <= 0:
        return max_value > 100
    return (max_value > p95 * 3) and (p95 > p50 * 2 or p75 > p50 * 1.5)


def build_long_tail_edges(series: pd.Series) -> List[float]:
    numeric = pd.to_numeric(series, errors="coerce").dropna()
    if numeric.empty:
        return []

    min_value = float(numeric.min())
    if min_value > 10 and not float(min_value).is_integer():
        min_value = float(math.ceil(min_value))
    max_value = float(numeric.max())
    candidate_edges = [
        11, 20, 30, 40, 50, 80, 100, 150, 200, 300, 500, 800, 1000,
        1500, 2000, 3000, 5000, 8000, 10000, 15000, 20000, 30000,
        50000, 80000, 100000, 150000, 200000, 300000, 500000,
        800000, 1000000, 1500000, 2000000, 3000000, 5000000,
        8000000, 10000000, 20000000, 50000000, 100000000,
        200000000, 500000000, 1000000000,
    ]

    edges = [min_value]
    for edge in candidate_edges:
        if min_value < edge < max_value:
            edges.append(float(edge))
    edges.append(max_value)
    edges = sorted(set(edges))
    if len(edges) < 2:
        return [min_value, max_value]
    return edges


def is_discrete_low_integer_like(series: pd.Series, tolerance: float = 1e-8) -> bool:
    numeric = pd.to_numeric(series, errors="coerce").dropna()
    if numeric.empty:
        return False
    if numeric.nunique(dropna=True) > 15:
        return False
    integer_like = (numeric - numeric.round()).abs() <= tolerance
    return bool(integer_like.all())


def build_fixed_width_edges(
    series: pd.Series,
    step: float,
    start: Optional[float] = None,
    end: Optional[float] = None,
) -> List[float]:
    numeric = pd.to_numeric(series, errors="coerce").dropna()
    if numeric.empty or step <= 0:
        return []

    min_value = float(numeric.min())
    max_value = float(numeric.max())
    if start is None:
        start = math.floor(min_value / step) * step
    if end is None:
        end = math.ceil(max_value / step) * step

    edges: List[float] = []
    current = float(start)
    upper_limit = float(end)
    while current <= upper_limit + step * 1e-9:
        edges.append(round(current, 10))
        current += step

    if not edges:
        edges = [round(min_value, 10), round(max_value, 10)]

    edges = sorted(set(edges))
    if edges[0] > min_value:
        edges.insert(0, round(min_value, 10))
    return edges


def choose_ratio_bin_step(series: pd.Series, target_bins: int = 15) -> float:
    numeric = pd.to_numeric(series, errors="coerce").dropna()
    if numeric.empty:
        return 0.1
    positive = numeric[numeric > 0]
    base = positive if not positive.empty else numeric
    min_value = float(base.min())
    max_value = float(base.max())
    if max_value <= min_value:
        return max(pretty_step(max_value if max_value > 0 else 0.001), 0.001)

    raw_step = (max_value - min_value) / max(target_bins, 1)
    return max(pretty_step(raw_step), 0.001)


def should_split_zero_from_ratio(series: pd.Series, zero_ratio_threshold: float = 0.1) -> bool:
    numeric = pd.to_numeric(series, errors="coerce").dropna()
    if numeric.empty:
        return False
    zero_ratio = float((numeric == 0).mean())
    return zero_ratio >= zero_ratio_threshold


def apply_interval_bins(
    values: pd.Series,
    edges: List[float],
    step_hint: float,
    fallback_prefix: str,
) -> pd.Series:
    numeric = pd.to_numeric(values, errors="coerce")
    non_null = numeric.dropna()
    if non_null.empty:
        return pd.Series([""] * len(values), index=values.index, dtype="object")

    sorted_edges = sorted(set(edges))
    if len(sorted_edges) < 2:
        return numeric.apply(lambda x: "" if pd.isna(x) else format_number(float(x)))

    actual_max = float(non_null.max())
    epsilon = max(abs(step_hint) * 0.001, 1e-9)
    if sorted_edges[-1] < actual_max:
        sorted_edges.append(actual_max + epsilon)
    elif sorted_edges[-1] == actual_max:
        sorted_edges[-1] = actual_max + epsilon

    label_gap = step_hint / 10 if step_hint > 0 else 0
    labels: List[str] = []
    for idx in range(len(sorted_edges) - 1):
        left = sorted_edges[idx]
        right = sorted_edges[idx + 1]
        if idx == len(sorted_edges) - 2:
            right_label = actual_max
        else:
            right_label = right - label_gap if label_gap > 0 and right > left else right
        labels.append(f"{format_number(left)}-{format_number(right_label)}")

    if len(set(labels)) != len(labels):
        labels = [
            f"{format_number(sorted_edges[idx])}-{format_number(sorted_edges[idx + 1])}"
            for idx in range(len(sorted_edges) - 1)
        ]
    if len(set(labels)) != len(labels):
        labels = [f"{fallback_prefix}_{idx + 1}" for idx in range(len(sorted_edges) - 1)]

    return pd.cut(
        numeric,
        bins=sorted_edges,
        labels=labels,
        include_lowest=True,
        right=False,
        duplicates="drop",
        ordered=False,
    ).astype("object").fillna("")


def format_number(value: float) -> str:
    if pd.isna(value):
        return ""
    if float(value).is_integer():
        return str(int(value))
    return f"{value:.4f}".rstrip("0").rstrip(".")


def parse_bin_sort_key(bin_value: object) -> Tuple[int, float, float, str]:
    text = "" if pd.isna(bin_value) else str(bin_value).strip()
    if text == "总计":
        return (3, float("inf"), float("inf"), text)
    if text == "-9999":
        return (0, float("-inf"), float("-inf"), text)
    if re.fullmatch(r"-?\d+(?:\.\d+)?", text):
        number = float(text)
        return (1, number, number, text)
    interval_match = re.fullmatch(r"(-?\d+(?:\.\d+)?)\s*-\s*(-?\d+(?:\.\d+)?)", text)
    if interval_match:
        left = float(interval_match.group(1))
        right = float(interval_match.group(2))
        return (2, left, right, text)
    return (2, float("inf"), float("inf"), text)


def merge_sparse_tail_bins(pivot: pd.DataFrame) -> pd.DataFrame:
    if pivot.empty:
        return pivot

    work = pivot.copy()
    work["总计_tmp"] = work[[0, 1, 2, 3]].sum(axis=1)
    interval_info = work["bin"].apply(parse_bin_sort_key)
    work["_bin_type"] = interval_info.apply(lambda x: x[0])
    work["_bin_left"] = interval_info.apply(lambda x: x[1])
    total_count = float(work["总计_tmp"].sum())
    if total_count <= 0:
        return pivot

    work["_bin_sort_key"] = work["bin"].apply(parse_bin_sort_key)
    interval_rows = work[(work["_bin_type"] == 2) & (work["_bin_left"] > 10)].copy()
    if interval_rows.shape[0] <= 1:
        return pivot

    interval_rows = interval_rows.sort_values("_bin_sort_key", kind="stable")
    sparse_tail_indices: List[int] = []
    cumulative_ratio = 0.0
    for idx in reversed(interval_rows.index.tolist()):
        row_total = float(work.loc[idx, "总计_tmp"])
        row_ratio = row_total / total_count
        if row_ratio >= 0.01:
            break
        if cumulative_ratio + row_ratio > 0.05:
            break
        sparse_tail_indices.append(idx)
        cumulative_ratio += row_ratio

    if not sparse_tail_indices:
        return pivot

    merge_indices = sorted(sparse_tail_indices)
    merged = work.loc[merge_indices, [0, 1, 2, 3, "总计_tmp", "_bin_left"]].copy()
    merged_left = float(merged["_bin_left"].min())
    merged_label = f"{format_number(merged_left)}+"

    keep = work.drop(index=merge_indices).copy()
    new_row = {
        "bin": merged_label,
        0: float(merged[0].sum()),
        1: float(merged[1].sum()),
        2: float(merged[2].sum()),
        3: float(merged[3].sum()),
    }
    keep = pd.concat([keep[[c for c in keep.columns if c in ["bin", 0, 1, 2, 3]]], pd.DataFrame([new_row])], ignore_index=True)
    keep["_bin_sort_key"] = keep["bin"].apply(parse_bin_sort_key)
    keep = keep.sort_values("_bin_sort_key", kind="stable").drop(columns=["_bin_sort_key"]).reset_index(drop=True)
    return keep


def merge_sparse_low_integer_bins(pivot: pd.DataFrame) -> pd.DataFrame:
    if pivot.empty:
        return pivot

    work = pivot.copy()
    work["总计_tmp"] = work[[0, 1, 2, 3]].sum(axis=1)
    total_count = float(work["总计_tmp"].sum())
    if total_count <= 0:
        return pivot

    bin_info = work["bin"].apply(parse_bin_sort_key)
    work["_bin_type"] = bin_info.apply(lambda x: x[0])
    work["_bin_left"] = bin_info.apply(lambda x: x[1])
    work["_bin_right"] = bin_info.apply(lambda x: x[2])

    low_rows = work[
        (work["_bin_type"] == 1)
        & (work["_bin_left"] >= 0)
        & (work["_bin_left"] <= 10)
        & (work["_bin_left"] == work["_bin_right"])
    ].copy()

    if low_rows.empty:
        return pivot

    if not ((low_rows["总计_tmp"] / total_count) < 0.01).all():
        return pivot

    merge_indices = low_rows.index.tolist()
    merged_left = int(low_rows["_bin_left"].min())
    merged_right = int(low_rows["_bin_right"].max())
    merged_label = f"{merged_left}-{merged_right}" if merged_left != merged_right else str(merged_left)

    keep = work.drop(index=merge_indices).copy()
    new_row = {
        "bin": merged_label,
        0: float(low_rows[0].sum()),
        1: float(low_rows[1].sum()),
        2: float(low_rows[2].sum()),
        3: float(low_rows[3].sum()),
    }
    keep = pd.concat([keep[[c for c in keep.columns if c in ["bin", 0, 1, 2, 3]]], pd.DataFrame([new_row])], ignore_index=True)
    keep["_bin_sort_key"] = keep["bin"].apply(parse_bin_sort_key)
    keep = keep.sort_values("_bin_sort_key", kind="stable").drop(columns=["_bin_sort_key"]).reset_index(drop=True)
    return keep


def assign_feature_bins(feature_name: str, series: pd.Series) -> pd.Series:
    numeric = pd.to_numeric(series, errors="coerce")
    if is_ma_like_feature(feature_name):
        edges = build_ma_edges(numeric)
        if not edges:
            return pd.Series([pd.NA] * len(series), index=series.index)

        sorted_edges = sorted(set(edges))
        if len(sorted_edges) < 2:
            return numeric.apply(lambda x: "" if pd.isna(x) else format_number(float(x)))

        labels: List[str] = []
        for idx in range(len(sorted_edges) - 1):
            left = sorted_edges[idx]
            right = sorted_edges[idx + 1]
            right_label = right - 1 if float(right).is_integer() and right > left else right
            labels.append(f"{format_number(left)}-{format_number(right_label)}")
        if len(set(labels)) != len(labels):
            labels = [
                f"{format_number(sorted_edges[idx])}-{format_number(sorted_edges[idx + 1])}"
                for idx in range(len(sorted_edges) - 1)
            ]
        if len(set(labels)) != len(labels):
            labels = [f"bin_{idx + 1}" for idx in range(len(sorted_edges) - 1)]
        return pd.cut(
            numeric,
            bins=sorted_edges,
            labels=labels,
            include_lowest=True,
            right=False,
            duplicates="drop",
            ordered=False,
        ).astype("object").fillna("")

    non_null = numeric.dropna()
    unique_values = sorted(non_null.unique().tolist())
    if len(unique_values) < 20:
        return numeric.apply(lambda x: "" if pd.isna(x) else format_number(float(x)))

    result = pd.Series([""] * len(series), index=series.index, dtype="object")

    negative_mask = numeric.notna() & (numeric < 0)
    if negative_mask.any():
        result.loc[negative_mask] = numeric.loc[negative_mask].apply(lambda x: format_number(float(x)))

    low_mask = numeric.notna() & (numeric >= 0) & (numeric <= 10)
    low_values = numeric.loc[low_mask]
    if not low_values.empty:
        if is_discrete_low_integer_like(low_values):
            result.loc[low_mask] = low_values.apply(lambda x: format_number(float(x)))
        else:
            ratio_mask = numeric.notna() & (numeric >= 0) & (numeric <= 1)
            ratio_values = numeric.loc[ratio_mask]
            if not ratio_values.empty:
                if ratio_values.nunique(dropna=True) <= 10 and is_discrete_low_integer_like(ratio_values):
                    result.loc[ratio_mask] = ratio_values.apply(lambda x: format_number(float(x)))
                else:
                    ratio_step = choose_ratio_bin_step(ratio_values)
                    if should_split_zero_from_ratio(ratio_values):
                        zero_mask = ratio_mask & (numeric == 0)
                        result.loc[zero_mask] = "0"
                        positive_ratio_mask = ratio_mask & (numeric > 0)
                        positive_ratio_values = numeric.loc[positive_ratio_mask]
                        if not positive_ratio_values.empty:
                            ratio_start = math.floor(float(positive_ratio_values.min()) / ratio_step) * ratio_step
                            if ratio_start <= 0:
                                ratio_start = ratio_step
                            ratio_end = max(ratio_start, math.ceil(float(positive_ratio_values.max()) / ratio_step) * ratio_step)
                            ratio_edges = build_fixed_width_edges(
                                positive_ratio_values,
                                step=ratio_step,
                                start=ratio_start,
                                end=ratio_end,
                            )
                            result.loc[positive_ratio_mask] = apply_interval_bins(
                                positive_ratio_values,
                                ratio_edges,
                                ratio_step,
                                "ratio_bin",
                            )
                    else:
                        ratio_end = max(ratio_step, math.ceil(float(ratio_values.max()) / ratio_step) * ratio_step)
                        ratio_edges = build_fixed_width_edges(ratio_values, step=ratio_step, start=0.0, end=ratio_end)
                        result.loc[ratio_mask] = apply_interval_bins(ratio_values, ratio_edges, ratio_step, "ratio_bin")

            low_float_mask = numeric.notna() & (numeric > 1) & (numeric <= 10)
            low_float_values = numeric.loc[low_float_mask]
            if not low_float_values.empty:
                if is_discrete_low_integer_like(low_float_values):
                    result.loc[low_float_mask] = low_float_values.apply(lambda x: format_number(float(x)))
                else:
                    start = max(1.0, math.floor(float(low_float_values.min())))
                    low_float_edges = build_fixed_width_edges(low_float_values, step=1.0, start=start, end=10.0)
                    result.loc[low_float_mask] = apply_interval_bins(low_float_values, low_float_edges, 1.0, "low_bin")

    high_mask = numeric.notna() & (numeric > 10)
    high_values = numeric.loc[high_mask]
    if high_values.empty:
        return result

    if is_long_tail_distribution(high_values):
        edges = build_long_tail_edges(high_values)
    else:
        edges = build_regular_edges(high_values)
    if not edges:
        result.loc[high_mask] = high_values.apply(lambda x: format_number(float(x)))
        return result

    sorted_edges = sorted(set(edges))
    if len(sorted_edges) < 2:
        result.loc[high_mask] = high_values.apply(lambda x: format_number(float(x)))
        return result
    high_max = float(high_values.max())
    step_hint = sorted_edges[-1] - sorted_edges[-2] if len(sorted_edges) >= 2 else 1.0
    if sorted_edges[-1] < high_max:
        step_hint = step_hint if step_hint > 0 else 1.0
        sorted_edges.append(high_max + max(step_hint * 0.001, 1e-9))
    elif sorted_edges[-1] == high_max:
        step_hint = step_hint if step_hint > 0 else 1.0
        sorted_edges[-1] = high_max + max(step_hint * 0.001, 1e-9)

    labels: List[str] = []
    for idx in range(len(sorted_edges) - 1):
        left = sorted_edges[idx]
        right = sorted_edges[idx + 1]
        if idx == len(sorted_edges) - 2:
            right_label = high_max
        else:
            right_label = right - 1 if float(right).is_integer() and right > left else right
        labels.append(f"{format_number(left)}-{format_number(right_label)}")
    if len(set(labels)) != len(labels):
        labels = [
            f"{format_number(sorted_edges[idx])}-{format_number(sorted_edges[idx + 1])}"
            for idx in range(len(sorted_edges) - 1)
        ]
    if len(set(labels)) != len(labels):
        labels = [f"high_bin_{idx + 1}" for idx in range(len(sorted_edges) - 1)]

    result.loc[high_mask] = pd.cut(
        high_values,
        bins=sorted_edges,
        labels=labels,
        include_lowest=True,
        right=False,
        duplicates="drop",
        ordered=False,
    ).astype("object").fillna("")
    return result


def summarize_feature(df: pd.DataFrame, feature_name: str) -> pd.DataFrame:
    work = df[[feature_name, "settleStatus_normalized"]].copy()
    work["bin"] = assign_feature_bins(feature_name, work[feature_name])
    work = work[(work["bin"] != "") & work["settleStatus_normalized"].notna()].copy()
    if work.empty:
        return pd.DataFrame()

    work["settleStatus_normalized"] = pd.to_numeric(work["settleStatus_normalized"], errors="coerce")
    pivot = (
        work.pivot_table(
            index="bin",
            columns="settleStatus_normalized",
            values=feature_name,
            aggfunc="count",
            fill_value=0,
        )
        .reset_index()
        .rename_axis(None, axis=1)
    )

    for column in [0, 1, 2, 3]:
        if column not in pivot.columns:
            pivot[column] = 0

    pivot = pivot[["bin", 0, 1, 2, 3]]
    pivot["_bin_sort_key"] = pivot["bin"].apply(parse_bin_sort_key)
    pivot = pivot.sort_values("_bin_sort_key", kind="stable").drop(columns=["_bin_sort_key"]).reset_index(drop=True)
    pivot = merge_sparse_low_integer_bins(pivot)
    pivot = merge_sparse_tail_bins(pivot)
    pivot["总计"] = pivot[[0, 1, 2, 3]].sum(axis=1)
    total_count = pivot["总计"].sum()
    pivot["首逾率"] = (pivot[0] + pivot[3]) / pivot["总计"].replace(0, pd.NA)
    pivot["在逾率"] = pivot[0] / pivot["总计"].replace(0, pd.NA)
    pivot["占比"] = pivot["总计"] / total_count if total_count else 0

    total_row = {
        "bin": "总计",
        0: int(pivot[0].sum()),
        1: int(pivot[1].sum()),
        2: int(pivot[2].sum()),
        3: int(pivot[3].sum()),
        "总计": int(pivot["总计"].sum()),
    }
    total_row["首逾率"] = ((total_row[0] + total_row[3]) / total_row["总计"]) if total_row["总计"] else 0
    total_row["在逾率"] = (total_row[0] / total_row["总计"]) if total_row["总计"] else 0
    total_row["占比"] = 1.0 if total_row["总计"] else 0

    pivot = pd.concat([pivot, pd.DataFrame([total_row])], ignore_index=True)
    pivot = pivot.rename(columns={"bin": feature_name})
    return pivot


def compute_single_feature_summary(
    feature_name: str,
    feature_values: List[float],
    settle_status_values: List[object],
) -> Tuple[str, pd.DataFrame]:
    df = pd.DataFrame(
        {
            feature_name: pd.to_numeric(pd.Series(feature_values), errors="coerce"),
            "settleStatus_normalized": pd.to_numeric(pd.Series(settle_status_values), errors="coerce"),
        }
    )
    return feature_name, summarize_feature(df, feature_name)


def compute_group_feature_tables(
    group_name: str,
    feature_payloads: List[Tuple[str, List[float]]],
    settle_status_values: List[object],
) -> Tuple[str, List[Tuple[str, pd.DataFrame]]]:
    feature_tables: List[Tuple[str, pd.DataFrame]] = []
    for feature_name, feature_values in feature_payloads:
        _, summary_df = compute_single_feature_summary(feature_name, feature_values, settle_status_values)
        if not summary_df.empty:
            feature_tables.append((feature_name, summary_df))
    feature_tables = sorted(feature_tables, key=lambda item: natural_sort_key(item[0]))
    return group_name, feature_tables


def build_feature_sheet_data(
    df: pd.DataFrame,
    feature_columns: List[str],
    feature_desc_map: Dict[str, str],
    show_feature_desc: bool,
) -> Dict[str, List[Tuple[str, pd.DataFrame]]]:
    stage_start = now_ts()
    print(f"[summary] start feature_count={len(feature_columns)}")
    grouped: Dict[str, List[Tuple[str, pd.DataFrame]]] = {}
    total = len(feature_columns)
    for index, feature_name in enumerate(feature_columns, start=1):
        summary_df = summarize_feature(df, feature_name)
        if not summary_df.empty:
            display_name = display_feature_name(feature_name, feature_desc_map, show_feature_desc)
            summary_df = summary_df.rename(columns={feature_name: display_name})
            group_name = classify_feature_group(feature_name)
            grouped.setdefault(group_name, []).append((feature_name, summary_df))
        if index % 100 == 0 or index == total:
            print(f"[summary] progress={index}/{total} elapsed={format_elapsed(stage_start)}")

    for group_name in grouped:
        grouped[group_name] = sorted(grouped[group_name], key=lambda item: natural_sort_key(item[0]))
    print(
        f"[summary] finished group_count={len(grouped)} "
        f"elapsed={format_elapsed(stage_start)}"
    )
    return dict(sorted(grouped.items(), key=lambda item: natural_sort_key(item[0])))


def write_dataframe_to_sheet(
    sheet,
    df: pd.DataFrame,
    start_row: int,
    start_col: int,
    header_fill: Optional[PatternFill] = None,
    left_align_first_header: bool = False,
) -> Tuple[int, int]:
    headers = list(df.columns)
    for col_offset, header in enumerate(headers):
        cell = sheet.cell(row=start_row, column=start_col + col_offset, value=header)
        cell.font = Font(bold=True)
        if col_offset == 0 and left_align_first_header:
            cell.alignment = Alignment(horizontal="left", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        if col_offset == 0:
            cell.number_format = "@"
        if header_fill:
            cell.fill = header_fill

    values = df.to_numpy(dtype=object)
    for row_offset, row in enumerate(values, start=1):
        for col_offset, header in enumerate(headers):
            value = row[col_offset]
            cell = sheet.cell(row=start_row + row_offset, column=start_col + col_offset)
            if col_offset == 0:
                cell.value = "" if pd.isna(value) else str(value)
                cell.number_format = "@"
                cell.quotePrefix = True
                if left_align_first_header:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            elif header in {"首逾率", "在逾率", "占比", "样本占比", "坏样本率"} and value != "" and pd.notna(value):
                cell.value = float(value)
                cell.number_format = "0.00%"
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.value = value
                cell.alignment = Alignment(horizontal="center", vertical="center")
    return df.shape[0] + 1, df.shape[1]


def write_feature_group_sheet(
    workbook: Workbook,
    sheet_name: str,
    feature_tables: List[Tuple[str, pd.DataFrame]],
    show_feature_desc: bool,
) -> None:
    sheet = workbook.create_sheet(title=sheet_name[:31])
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    current_row = 1
    table_width = 9
    gap_cols = 1
    gap_rows = 3

    for index in range(0, len(feature_tables), 3):
        chunk = feature_tables[index : index + 3]
        max_height = 0
        for chunk_index, (_, table_df) in enumerate(chunk):
            start_col = 1 + chunk_index * (table_width + gap_cols)
            height, _ = write_dataframe_to_sheet(
                sheet,
                table_df,
                start_row=current_row,
                start_col=start_col,
                header_fill=header_fill,
                left_align_first_header=True,
            )
            max_height = max(max_height, height)
        current_row += max_height + gap_rows

    for column_cells in sheet.columns:
        column_letter = column_cells[0].column_letter
        sheet.column_dimensions[column_letter].width = 8


def write_output_workbook(
    output_path: Path,
    woe_group_all: pd.DataFrame,
    grouped_feature_tables: Dict[str, List[Tuple[str, pd.DataFrame]]],
    show_feature_desc: bool,
) -> None:
    stage_start = now_ts()
    print(
        f"[excel] start output={output_path} "
        f"sheet_count={1 + len(grouped_feature_tables)}"
    )
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    woe_sheet = workbook.create_sheet(title="woe_group_all")
    write_dataframe_to_sheet(
        woe_sheet,
        woe_group_all,
        start_row=1,
        start_col=1,
        header_fill=PatternFill(fill_type="solid", fgColor="FCE4D6"),
    )
    for column_cells in woe_sheet.columns:
        column_letter = column_cells[0].column_letter
        woe_sheet.column_dimensions[column_letter].width = 16
    print(f"[excel] wrote_sheet=woe_group_all elapsed={format_elapsed(stage_start)}")

    total_groups = len(grouped_feature_tables)
    for index, (sheet_name, feature_tables) in enumerate(grouped_feature_tables.items(), start=1):
        write_feature_group_sheet(
            workbook,
            sheet_name,
            feature_tables,
            show_feature_desc=show_feature_desc,
        )
        print(
            f"[excel] wrote_sheet={sheet_name} "
            f"progress={index}/{total_groups} elapsed={format_elapsed(stage_start)}"
        )

    workbook.save(output_path)
    print(f"[excel] saved elapsed={format_elapsed(stage_start)}")


def process_single_segment(
    segment_name: str,
    df: pd.DataFrame,
    output_path: Path,
    feature_prefixes: Optional[str],
    sheet_output_config: Dict[str, bool],
    config_path: Path,
    feature_desc_map: Dict[str, str],
    show_feature_desc: bool,
) -> None:
    print(f"[segment] processing={segment_name} rows={len(df)} output={output_path}")

    stage_start = now_ts()
    print("[main] select_features start")
    feature_columns = select_feature_columns(df)
    feature_columns = filter_features_by_prefix(feature_columns, feature_prefixes)
    print(f"[main] select_features done feature_count={len(feature_columns)} elapsed={format_elapsed(stage_start)}")
    if not feature_columns:
        print(f"[segment] skip={segment_name} reason=未识别到可分析的数值特征列")
        return

    print(f"loaded_rows={len(df)}")
    print(f"feature_count={len(feature_columns)}")

    stage_start = now_ts()
    woe_group_all, _ = build_tools_woe_bins(df, feature_columns)
    if not woe_group_all.empty:
        woe_group_all["特征名"] = woe_group_all["特征名"].apply(
            lambda x: display_feature_name(str(x), feature_desc_map, show_feature_desc)
        )
    print(f"[main] woe_done rows={len(woe_group_all)} elapsed={format_elapsed(stage_start)}")

    stage_start = now_ts()
    grouped_feature_tables = build_feature_sheet_data(
        df,
        feature_columns,
        feature_desc_map,
        show_feature_desc,
    )
    sheet_output_config, new_sheets = sync_new_sheet_config(
        config_path=config_path,
        sheet_output_config=sheet_output_config,
        detected_sheet_names=["woe_group_all", *grouped_feature_tables.keys()],
    )
    if new_sheets:
        print(f"[config] segment={segment_name} new_sheets_added={new_sheets}")
    else:
        print(f"[config] segment={segment_name} new_sheets_added=[]")
    grouped_feature_tables = filter_sheet_tables_by_config(grouped_feature_tables, sheet_output_config)
    print(f"[main] summary_done group_count={len(grouped_feature_tables)} elapsed={format_elapsed(stage_start)}")

    stage_start = now_ts()
    write_output_workbook(
        output_path,
        woe_group_all,
        grouped_feature_tables,
        show_feature_desc=show_feature_desc,
    )
    print(f"[main] excel_done elapsed={format_elapsed(stage_start)}")
    print(f"output_file={output_path}")


def main() -> None:
    total_start = now_ts()
    args = parse_args()
    input_path = Path(args.input_file)
    output_path = derive_output_path(input_path, args.output_file)
    CONFIG_DIR.mkdir(parents=True, exist_ok=True)
    config_path = Path(args.sheet_config_file) if args.sheet_config_file else (CONFIG_DIR / "feature_sheet_output_config.json")
    display_config_path = Path(args.display_config_file) if args.display_config_file else DISPLAY_CONFIG_PATH
    feature_desc_path = Path(args.feature_desc_file) if args.feature_desc_file else FEATURE_DESC_PATH
    sheet_output_config = load_sheet_output_config(config_path)
    show_feature_desc = load_feature_display_config(display_config_path)
    feature_desc_map = load_feature_description_map(feature_desc_path)
    print(f"[config] feature_desc_count={len(feature_desc_map)} from={feature_desc_path}")
    print(f"[config] show_feature_desc={show_feature_desc} from={display_config_path}")

    stage_start = now_ts()
    print(f"[main] load_input start file={input_path}")
    df = load_input_dataframe(args.input_file)
    print(f"[main] load_input done rows={len(df)} cols={len(df.columns)} elapsed={format_elapsed(stage_start)}")

    stage_start = now_ts()
    print("[main] normalize_target start")
    df = ensure_target_columns(df, status_source=args.status_source)
    print(f"[main] normalize_target done rows={len(df)} elapsed={format_elapsed(stage_start)}")
    status_mapping = {
        0: "settleStatus",
        1: "settleStatus_1",
        2: "settleStatus_2",
        3: "settleStatus_3",
    }
    print(f"[main] status_source={args.status_source} column={status_mapping.get(args.status_source, 'settleStatus')}")

    segments = split_dataframe_by_segment(df)
    available_segments = [segment_name for segment_name in SEGMENT_ORDER if segment_name in segments]
    print(f"[segment] available={available_segments}")

    for segment_name in available_segments:
        segment_output_path = derive_segment_output_path(output_path, segment_name)
        process_single_segment(
            segment_name=segment_name,
            df=segments[segment_name],
            output_path=segment_output_path,
            feature_prefixes=args.feature_prefixes,
            sheet_output_config=sheet_output_config,
            config_path=config_path,
            feature_desc_map=feature_desc_map,
            show_feature_desc=show_feature_desc,
        )

    print(f"[main] total_elapsed={format_elapsed(total_start)}")


if __name__ == "__main__":
    main()
