from __future__ import annotations

import locale
import math
import shutil
import subprocess
import tempfile
from io import BytesIO
from pathlib import Path

import joblib
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image as PILImage
from sklearn.metrics import roc_auc_score, roc_curve

try:
    import sklearn2pmml as sklearn2pmml_pkg
    from sklearn2pmml import PMMLPipeline
    from sklearn_pandas import DataFrameMapper
    PMML_AVAILABLE = True
except ImportError:
    PMML_AVAILABLE = False


def _configure_matplotlib():
    plt.rcParams['font.sans-serif'] = ['Microsoft YaHei', 'SimHei', 'Arial Unicode MS', 'DejaVu Sans']
    plt.rcParams['axes.unicode_minus'] = False
    plt.rcParams['font.size'] = 12
    plt.rcParams['axes.titlesize'] = 16
    plt.rcParams['axes.labelsize'] = 14
    plt.rcParams['xtick.labelsize'] = 11
    plt.rcParams['ytick.labelsize'] = 11
    plt.rcParams['legend.fontsize'] = 12


CHART_SIZES = {
    'roc': (880, 660),
    'trend': (880, 660),
    'psi': (880, 660),
    'bad_rate': (880, 660),
    'feature_importance': (900, 1500),
}


def _calc_ks_value(y_true, y_pred):
    fpr, tpr, _ = roc_curve(y_true, y_pred)
    return float(np.max(np.abs(tpr - fpr)))


def plot_roc_curve(y_true_dict, y_pred_dict, save_path):
    _configure_matplotlib()
    plt.figure(figsize=(10, 8))
    colors = {'train': 'blue', 'test': 'green', 'oot': 'red'}
    labels = {'train': '训练集', 'test': '测试集', 'oot': '验证集'}
    for name in ['train', 'test', 'oot']:
        if name in y_true_dict and name in y_pred_dict:
            fpr, tpr, _ = roc_curve(y_true_dict[name], y_pred_dict[name])
            auc = roc_auc_score(y_true_dict[name], y_pred_dict[name])
            ks = _calc_ks_value(y_true_dict[name], y_pred_dict[name])
            plt.plot(fpr, tpr, lw=2, color=colors[name], label=f"{labels[name]} (AUC={auc:.4f}, KS={ks:.4f})")
    plt.plot([0, 1], [0, 1], 'k--', lw=1)
    plt.xlabel('假阳性率')
    plt.ylabel('真正率')
    plt.title('训练集/测试集/验证集 ROC 曲线')
    plt.legend(loc='lower right', fontsize=9)
    plt.grid(True, alpha=0.3)
    plt.tight_layout()
    plt.savefig(save_path, dpi=150, bbox_inches='tight')
    plt.close()
    return save_path


def plot_feature_importance(var_list, importance_values, save_path):
    _configure_matplotlib()
    importance_df = pd.DataFrame({'var': var_list, 'importance': importance_values}).sort_values('importance', ascending=True)
    plt.figure(figsize=(10, max(6, len(importance_df) * 0.28)))
    plt.barh(importance_df['var'], importance_df['importance'], color='#6ea8fe')
    plt.xlabel('重要性')
    plt.ylabel('特征')
    plt.title('特征重要性排序')
    plt.grid(True, axis='x', alpha=0.3)
    plt.tight_layout()
    plt.savefig(save_path, dpi=150, bbox_inches='tight')
    plt.close()
    return save_path


def _safe_sklearn2pmml(pipeline, pmml_path):
    original_popen = sklearn2pmml_pkg.Popen

    def _encoding_safe_popen(*args, **kwargs):
        if kwargs.get('universal_newlines') and 'encoding' not in kwargs and 'text' not in kwargs:
            kwargs['encoding'] = locale.getencoding() or 'gbk'
            kwargs.setdefault('errors', 'replace')
        return subprocess.Popen(*args, **kwargs)

    sklearn2pmml_pkg.Popen = _encoding_safe_popen
    try:
        sklearn2pmml_pkg.sklearn2pmml(pipeline, str(pmml_path))
    finally:
        sklearn2pmml_pkg.Popen = original_popen


def export_pmml(trainer, output_dir, prefix, model_type):
    if not PMML_AVAILABLE:
        print('警告: sklearn2pmml 未安装，跳过 PMML 导出')
        return
    try:
        mapper = DataFrameMapper([(feature, None) for feature in trainer.var_list])
        pipeline = PMMLPipeline([('mapper', mapper), ('classifier', trainer.model)])
        pmml_path = output_dir / f'{prefix}_{model_type}_model.pmml'
        _safe_sklearn2pmml(pipeline, pmml_path)
        print(f'PMML 文件已保存: {pmml_path}')
    except Exception as e:
        print(f'PMML 导出失败: {e}')


def export_scorecard_excel(trainer, output_dir, prefix):
    scorecard_params = trainer.config['scorecard'].get('scorecard_params', {'pdo': 20, 'base_score': 500, 'base_odds': 20})
    pdo = scorecard_params['pdo']
    base_score = scorecard_params['base_score']
    base_odds = scorecard_params['base_odds']
    p = pdo / math.log(2)
    q = base_score - pdo / math.log(2) * math.log(base_odds)
    excel_path = output_dir / f'{prefix}_scorecard.xlsx'

    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        all_bins = [trainer.woe_binner.bin_info[var].copy() for var in trainer.var_list if var in trainer.woe_binner.bin_info]
        if all_bins:
            pd.concat(all_bins, ignore_index=True).to_excel(writer, sheet_name='分箱详情', index=False)
        if hasattr(trainer.model, 'coef_'):
            coef_dict = dict(zip(trainer.var_list, trainer.model.coef_[0]))
            intercept_arr = trainer.model.intercept_
            intercept = float(intercept_arr[0]) if np.ndim(intercept_arr) > 0 else float(intercept_arr)
            rows = []
            for var in trainer.var_list:
                if var not in trainer.woe_binner.bin_info:
                    continue
                for _, row in trainer.woe_binner.bin_info[var].iterrows():
                    rows.append({
                        'VAR': var,
                        'bin_range': f"({row['LL']}, {row['UL']}]" if row['LL'] != row['UL'] else str(row['LL']),
                        'woe': round(row['woe'], 4),
                        'coef': round(coef_dict.get(var, 0), 4),
                        'score': round(-(coef_dict.get(var, 0) * row['woe'] * p)),
                        'count': row['count'],
                        'bad_rate': round(row['bad_rate'], 4),
                    })
            pd.DataFrame(rows).to_excel(writer, sheet_name='评分卡', index=False)
            base_score_val = round(q + intercept * p)
            pd.DataFrame({'项目': ['基础分', 'PDO', '基准分', '基准Odds', '截距', '变量数'], '取值': [base_score_val, pdo, base_score, base_odds, round(intercept, 4), len(trainer.var_list)]}).to_excel(writer, sheet_name='评分卡参数', index=False)
    print(f'评分卡已保存: {excel_path}')


def export_training_report(trainer, output_dir, prefix, y_true_dict, y_pred_dict):
    export_oot_report(trainer, output_dir, prefix, y_true_dict=y_true_dict, y_pred_dict=y_pred_dict)
    excel_path = output_dir / f'{prefix}_training_report.xlsx'
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        pd.DataFrame({
            '项目': ['模型类型', '训练时间', '变量数', '训练集AUC', '训练集KS', '测试集AUC', '测试集KS', '验证集AUC', '验证集KS'],
            '取值': [
                trainer.train_log.get('model_type', ''),
                f"{trainer.train_log.get('start_time', '')} ~ {trainer.train_log.get('end_time', '')}",
                len(trainer.var_list or []),
                trainer.train_log.get('metrics', {}).get('train_auc', 0),
                trainer.train_log.get('metrics', {}).get('train_ks', 0),
                trainer.train_log.get('metrics', {}).get('test_auc', 0),
                trainer.train_log.get('metrics', {}).get('test_ks', 0),
                trainer.train_log.get('metrics', {}).get('oot_auc', 0),
                trainer.train_log.get('metrics', {}).get('oot_ks', 0),
            ],
        }).to_excel(writer, sheet_name='模型概览', index=False)
    print(f'训练报告已保存: {excel_path}')


def _probability_to_score(pred_values, trainer):
    params = trainer.config.get('scorecard', {}).get('scorecard_params', {'pdo': 20, 'base_score': 500, 'base_odds': 20})
    pdo = params.get('pdo', 20)
    base_score = params.get('base_score', 500)
    base_odds = params.get('base_odds', 20)
    factor = pdo / math.log(2)
    offset = base_score - factor * math.log(base_odds)
    pred = np.clip(np.asarray(pred_values, dtype=float), 1e-6, 1 - 1e-6)
    odds = pred / (1 - pred)
    return np.round(offset - factor * np.log(odds), 0)


def _build_eval_frame(base_df, y_name, pred_values, dataset_name, date_col, trainer):
    df = base_df.copy().reset_index(drop=True)
    df['pred'] = pred_values
    df['dataset'] = dataset_name
    df['score'] = _probability_to_score(df['pred'], trainer)
    if date_col in df.columns:
        df[date_col] = pd.to_datetime(df[date_col], errors='coerce')
    return df


def _make_score_bins(dev_scores, oot_scores, score_bin_width=10):
    all_scores = np.concatenate([np.asarray(dev_scores, dtype=float), np.asarray(oot_scores, dtype=float)])
    width = max(int(score_bin_width), 1)
    start = math.floor(float(np.min(all_scores)) / width) * width
    end = math.ceil(float(np.max(all_scores)) / width) * width
    if start == end:
        end = start + width
    edges = np.arange(start, end + width, width, dtype=float)
    edges[0] = -np.inf
    edges[-1] = np.inf
    return edges


def _calc_psi(dev_series, oot_series, edges):
    dev_bins = pd.cut(dev_series, bins=edges, include_lowest=True)
    oot_bins = pd.cut(oot_series, bins=edges, include_lowest=True)
    dev_dist = dev_bins.value_counts(normalize=True, sort=False)
    oot_dist = oot_bins.value_counts(normalize=True, sort=False)
    psi_df = pd.DataFrame({'baseline_pct': dev_dist, 'validation_pct': oot_dist}).fillna(0)
    psi_df['baseline_pct'] = psi_df['baseline_pct'].clip(lower=1e-6)
    psi_df['validation_pct'] = psi_df['validation_pct'].clip(lower=1e-6)
    psi_df['psi'] = (psi_df['validation_pct'] - psi_df['baseline_pct']) * np.log(psi_df['validation_pct'] / psi_df['baseline_pct'])
    return psi_df.reset_index().rename(columns={'index': 'score_band'})


def _format_date_range(series):
    valid = pd.to_datetime(series, errors='coerce').dropna()
    if valid.empty:
        return ''
    return f'{valid.min().strftime("%Y-%m-%d")} ~ {valid.max().strftime("%Y-%m-%d")}'


def _build_strategy_threshold_report(eval_df, y_name, score_edges):
    thresholds = sorted(set(int(edge) for edge in score_edges[1:-1] if np.isfinite(edge)))
    total_count = len(eval_df)
    total_bad = max(float(eval_df[y_name].sum()), 1.0)
    total_good = max(float((eval_df[y_name] == 0).sum()), 1.0)
    rows = []
    for threshold in thresholds:
        approved = eval_df[eval_df['score'] >= threshold]
        rejected = eval_df[eval_df['score'] < threshold]
        rows.append({
            'score_threshold': threshold,
            'approve_rate': len(approved) / total_count if total_count else 0.0,
            'reject_rate': len(rejected) / total_count if total_count else 0.0,
            'approved_bad_rate': approved[y_name].mean() if len(approved) else np.nan,
            'rejected_bad_rate': rejected[y_name].mean() if len(rejected) else np.nan,
            'bad_capture_rate': rejected[y_name].sum() / total_bad if len(rejected) else 0.0,
            'good_pass_rate': ((approved[y_name] == 0).sum() / total_good) if len(approved) else 0.0,
            'approved_count': len(approved),
            'rejected_count': len(rejected),
        })
    return pd.DataFrame(rows)


def _safe_image(fig, target_size=None):
    bio = BytesIO()
    fig.savefig(bio, format='png', dpi=220, bbox_inches='tight', facecolor='white')
    plt.close(fig)
    bio.seek(0)
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
    tmp.write(bio.read())
    tmp.close()
    if target_size:
        img = PILImage.open(tmp.name)
        img = img.resize(target_size, PILImage.Resampling.LANCZOS)
        img.save(tmp.name)
    return tmp.name


def _add_sized_image(ws, image_path, cell):
    img = XLImage(image_path)
    ws.add_image(img, cell)


def _build_oot_workbook():
    wb = Workbook()
    default_ws = wb.active
    default_ws.title = '上线建议说明'
    for sheet_name in [
        '上线判定阈值',
        '模型概览',
        '分数分布PSI',
        '验证集分数分段表现',
        '策略阈值分析',
        '入模变量及重要性',
        '变量稳定性筛选',
        '多时间窗验证',
        '候选参数稳定性复核',
        '树模型候选特征筛选过程',
        '逻辑回归候选特征筛选过程',
        '评估图表',
    ]:
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)
    return wb


def _apply_two_column_number_formats(ws):
    percent_keywords = ('坏账率', '通过率', '拒绝率', '拦截率', '好客户通过率')
    decimal_keywords = ('AUC', 'KS', 'PSI', '变化', '截距')
    integer_keywords = ('样本量', '变量数', '随机种子', '基础分', 'PDO', '基准Odds', '分数分段宽度')
    for row in range(2, ws.max_row + 1):
        label = ws.cell(row, 1).value
        value_cell = ws.cell(row, 2)
        if not isinstance(label, str) or not isinstance(value_cell.value, (int, float, np.integer, np.floating)):
            continue
        if any(keyword in label for keyword in percent_keywords):
            value_cell.number_format = '0.00%'
        elif any(keyword in label for keyword in decimal_keywords):
            value_cell.number_format = '0.0000'
        elif any(keyword in label for keyword in integer_keywords):
            value_cell.number_format = '0'


def _apply_header_number_formats(ws, header_row):
    header_map = {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(header_row, col).value
        if isinstance(header, str) and header:
            header_map[col] = header
    percent_keywords = ('占比', '坏账率', '通过率', '拒绝率', '拦截率')
    integer_keywords = ('样本量', '样本数', '排序', '阈值')
    decimal_keywords = ('AUC', 'KS', 'PSI', '重要性', '得分', '标准差', '平均分', 'WOE', 'coef', 'score')
    for row in range(header_row + 1, ws.max_row + 1):
        for col, header in header_map.items():
            cell = ws.cell(row, col)
            if not isinstance(cell.value, (int, float, np.integer, np.floating)):
                continue
            if any(keyword in header for keyword in percent_keywords):
                cell.number_format = '0.00%'
            elif any(keyword in header for keyword in integer_keywords):
                cell.number_format = '0'
            elif any(keyword in header for keyword in decimal_keywords):
                cell.number_format = '0.0000'


def _apply_workbook_number_formats(wb):
    for sheet_name in ['上线建议说明', '模型概览']:
        if sheet_name in wb.sheetnames:
            _apply_two_column_number_formats(wb[sheet_name])
    if '分数分布PSI' in wb.sheetnames:
        ws = wb['分数分布PSI']
        for row in range(2, ws.max_row + 1):
            ws.cell(row, 2).number_format = '0.00%'
            ws.cell(row, 3).number_format = '0.00%'
            ws.cell(row, 4).number_format = '0.0000'
    if '验证集分数分段表现' in wb.sheetnames:
        ws = wb['验证集分数分段表现']
        for row in range(2, ws.max_row + 1):
            ws.cell(row, 2).number_format = '0'
            ws.cell(row, 3).number_format = '0'
            ws.cell(row, 4).number_format = '0.00%'
            ws.cell(row, 5).number_format = '0.00'
            ws.cell(row, 6).number_format = '0.00%'
            ws.cell(row, 7).number_format = '0.00%'
            ws.cell(row, 8).number_format = '0'
            ws.cell(row, 9).number_format = '0.00%'
    if '策略阈值分析' in wb.sheetnames:
        ws = wb['策略阈值分析']
        for row in range(2, ws.max_row + 1):
            ws.cell(row, 1).number_format = '0'
            for col in range(2, 8):
                ws.cell(row, col).number_format = '0.00%'
            ws.cell(row, 8).number_format = '0'
            ws.cell(row, 9).number_format = '0'
    if '入模变量及重要性' in wb.sheetnames:
        ws = wb['入模变量及重要性']
        for row in range(2, ws.max_row + 1):
            ws.cell(row, 2).number_format = '0.0000'
            for col in range(3, 6):
                ws.cell(row, col).number_format = '0'
    header_rows = {
        '变量稳定性筛选': 9,
        '多时间窗验证': 6,
        '候选参数稳定性复核': 6,
        '树模型候选特征筛选过程': 13,
        '逻辑回归候选特征筛选过程': 13,
    }
    for sheet_name, header_row in header_rows.items():
        if sheet_name in wb.sheetnames:
            _apply_header_number_formats(wb[sheet_name], header_row)


def _prepare_eval_chart_sheet(ws):
    ws._images = []
    for row in range(1, max(ws.max_row, 400) + 1):
        ws.cell(row, 1).value = None
    ws['A1'] = '模型评估图表'


def _estimate_row_span(image_path, default_row_height=20):
    with PILImage.open(image_path) as img:
        height_px = img.height
    row_height_px = default_row_height * 96 / 72
    return max(int(math.ceil(height_px / row_height_px)), 1)


def _append_chart_block(ws, start_row, title, image_path, spacing_rows=10):
    ws.cell(start_row, 1).value = title
    image_row = start_row + 1
    _add_sized_image(ws, image_path, f'A{image_row}')
    row_span = _estimate_row_span(image_path)
    return image_row + row_span + spacing_rows


def _ensure_sheet(wb, sheet_name):
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name)
    return wb[sheet_name]


def _style_workbook(workbook_path):
    wb = load_workbook(workbook_path)
    body_font = Font(name='微软雅黑', size=10)
    header_font = Font(name='微软雅黑', size=10, bold=True)
    header_fill = PatternFill(fill_type='solid', fgColor='EAF2F8')
    section_fill = PatternFill(fill_type='solid', fgColor='D9EAF7')
    for ws in wb.worksheets:
        for row_idx, row in enumerate(ws.iter_rows(), start=1):
            for cell in row:
                cell.font = header_font if row_idx == 1 else body_font
                cell.alignment = Alignment(horizontal='left', vertical='center')
        if ws.max_row >= 1:
            for cell in ws[1]:
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
        if ws.title in {'变量稳定性筛选', '多时间窗验证', '候选参数稳定性复核'}:
            ws.freeze_panes = 'A6'
        elif ws.title in {'树模型候选特征筛选过程', '逻辑回归候选特征筛选过程'}:
            ws.freeze_panes = 'A12'
        else:
            ws.freeze_panes = 'A2'
        for row_idx in range(1, ws.max_row + 1):
            first_value = ws.cell(row=row_idx, column=1).value
            if isinstance(first_value, str) and first_value in {'说明', '步骤', '筛选阶段', '关键观察', '关键结论', '逻辑回归筛选结论', '口径说明'}:
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = section_fill
                    ws.cell(row=row_idx, column=col_idx).font = header_font
                    ws.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal='center', vertical='center')
        width_map = {
            '上线建议说明': {'A': 15, 'B': 29},
            '上线判定阈值': {'A': 12, 'B': 12, 'C': 13, 'D': 12},
            '模型概览': {'A': 12, 'B': 42},
            '分数分布PSI': {'A': 16, 'B': 22, 'C': 22, 'D': 23},
            '验证集分数分段表现': {'A': 16, 'B': 12, 'C': 12, 'D': 21, 'E': 19, 'F': 22, 'G': 22, 'H': 12, 'I': 16},
            '策略阈值分析': {'A': 12, 'B': 22, 'C': 21, 'D': 21, 'E': 20, 'F': 22, 'G': 21, 'H': 12, 'I': 12},
            '入模变量及重要性': {'A': 14, 'B': 22, 'C': 16, 'D': 12, 'E': 12},
            '变量稳定性筛选': {'A': 14, 'B': 42, 'C': 12, 'D': 12, 'E': 23, 'F': 12, 'G': 12},
            '多时间窗验证': {'A': 22, 'B': 42, 'C': 12, 'D': 12, 'E': 20, 'F': 20, 'G': 20, 'H': 20, 'I': 20, 'J': 20, 'K': 42},
            '候选参数稳定性复核': {'A': 12, 'B': 42, 'C': 20, 'D': 20, 'E': 20, 'F': 21, 'G': 20, 'H': 20, 'I': 20, 'J': 20, 'K': 42},
            '树模型候选特征筛选过程': {'A': 14, 'B': 42, 'C': 12, 'D': 12, 'E': 12, 'F': 12},
            '逻辑回归候选特征筛选过程': {'A': 14, 'B': 42, 'C': 12, 'D': 12, 'E': 12, 'F': 12},
            '评估图表': {'A': 20},
        }
        if ws.title in width_map:
            for col_letter, width in width_map[ws.title].items():
                ws.column_dimensions[col_letter].width = width
        else:
            for col_cells in ws.columns:
                width = min(max(len(str(cell.value)) if cell.value is not None else 0 for cell in col_cells) + 2, 42)
                ws.column_dimensions[col_cells[0].column_letter].width = width
    _apply_workbook_number_formats(wb)
    wb.save(workbook_path)


def export_oot_report(trainer, output_dir, prefix, y_true_dict, y_pred_dict):
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    eval_data = trainer.train_log.get('evaluation_data', {})
    y_name = eval_data.get('y_name', 'target')
    date_col = eval_data.get('date_col', 'date')
    train_eval = _build_eval_frame(eval_data['train'], y_name, eval_data['train_pred'], 'train', date_col, trainer)
    test_eval = _build_eval_frame(eval_data['test'], y_name, eval_data['test_pred'], 'test', date_col, trainer)
    oot_eval = _build_eval_frame(eval_data['oot'], y_name, eval_data['oot_pred'], 'oot', date_col, trainer)
    baseline_eval = pd.concat([train_eval, test_eval], ignore_index=True)

    train_auc = roc_auc_score(train_eval[y_name], train_eval['pred'])
    train_ks = _calc_ks_value(train_eval[y_name], train_eval['pred'])
    test_auc = roc_auc_score(test_eval[y_name], test_eval['pred'])
    test_ks = _calc_ks_value(test_eval[y_name], test_eval['pred'])
    oot_auc = roc_auc_score(oot_eval[y_name], oot_eval['pred'])
    oot_ks = _calc_ks_value(oot_eval[y_name], oot_eval['pred'])

    score_bin_width = trainer.config.get('scorecard', {}).get('scorecard_params', {}).get('score_bin_width', 10)
    score_edges = _make_score_bins(baseline_eval['score'], oot_eval['score'], score_bin_width=score_bin_width)
    psi_df = _calc_psi(baseline_eval['score'], oot_eval['score'], score_edges)
    psi_total = float(psi_df['psi'].sum())

    baseline_eval['score_band'] = pd.cut(baseline_eval['score'], bins=score_edges, include_lowest=True)
    oot_eval['score_band'] = pd.cut(oot_eval['score'], bins=score_edges, include_lowest=True)
    band_report = oot_eval.groupby('score_band', observed=False).agg(
        sample_count=(y_name, 'size'),
        bad_count=(y_name, 'sum'),
        bad_rate=(y_name, 'mean'),
        avg_score=('score', 'mean'),
    ).reset_index()
    band_report['sample_pct'] = band_report['sample_count'] / max(len(oot_eval), 1)
    band_report['cum_sample_pct'] = band_report['sample_pct'].cumsum()
    band_report['cum_bad_count'] = band_report['bad_count'].cumsum()
    total_bad_oot = max(float(oot_eval[y_name].sum()), 1.0)
    band_report['cum_bad_pct'] = band_report['cum_bad_count'] / total_bad_oot
    strategy_threshold_df = _build_strategy_threshold_report(oot_eval, y_name, score_edges)

    auc_delta = oot_auc - test_auc
    ks_delta = oot_ks - test_ks
    if oot_auc >= 0.65 and oot_ks >= 0.25 and psi_total < 0.10:
        stability_level = '稳定'
        launch_decision = '建议上线'
        action = '建议结合审批策略、拒绝率和收益测算后决定是否上线'
        risk_note = '训练集、测试集和验证集表现整体可接受，可按常规流程推进'
    elif oot_auc >= 0.60 and oot_ks >= 0.20 and psi_total < 0.25:
        stability_level = '需关注'
        launch_decision = '有条件上线'
        action = '建议补充策略模拟、分层坏账率和样本稳定性复核后再决策'
        risk_note = '模型效果尚可，但需重点关注验证集衰减和样本漂移'
    else:
        stability_level = '不稳定'
        launch_decision = '建议拒绝'
        action = '建议继续优化变量、样本窗口和模型结构后再评估'
        risk_note = '验证集效果或稳定性不足，不建议直接上线'

    decision_df = pd.DataFrame({
        '项目': [
            '模型类型', '训练集样本量', '测试集样本量', '验证集样本量',
            '训练集坏账率', '测试集坏账率', '验证集坏账率',
            '训练集AUC', '训练集KS', '测试集AUC', '测试集KS', '验证集AUC', '验证集KS',
            '验证集相对测试集AUC变化', '验证集相对测试集KS变化',
            '分数PSI', '稳定性等级', '上线结论', '建议动作', '风险提示'
        ],
        '取值': [
            trainer.train_log.get('model_type', ''),
            len(train_eval),
            len(test_eval),
            len(oot_eval),
            float(train_eval[y_name].mean()) if len(train_eval) else np.nan,
            float(test_eval[y_name].mean()) if len(test_eval) else np.nan,
            float(oot_eval[y_name].mean()) if len(oot_eval) else np.nan,
            train_auc, train_ks, test_auc, test_ks, oot_auc, oot_ks,
            auc_delta, ks_delta,
            psi_total, stability_level, launch_decision, action, risk_note,
        ],
    })
    criteria_df = pd.DataFrame({'指标': ['验证集AUC', '验证集KS', 'Score PSI'], '建议通过': ['>= 0.65', '>= 0.25', '< 0.10'], '有条件通过': ['0.60 - 0.65', '0.20 - 0.25', '0.10 - 0.25'], '建议拒绝': ['< 0.60', '< 0.20', '>= 0.25']})

    model_summary_df = pd.DataFrame({
        '项目': ['模型类型', '训练时间', '随机种子', '开发样本日期范围', '验证集日期范围', '基础分', 'PDO', '基准Odds', '分数分段宽度', '入模变量数'],
        '取值': [trainer.train_log.get('model_type', ''), f"{trainer.train_log.get('start_time', '')} ~ {trainer.train_log.get('end_time', '')}", trainer.config.get('seed', ''), _format_date_range(baseline_eval[date_col]) if date_col in baseline_eval.columns else '', _format_date_range(oot_eval[date_col]) if date_col in oot_eval.columns else '', trainer.config.get('scorecard', {}).get('scorecard_params', {}).get('base_score', 500), trainer.config.get('scorecard', {}).get('scorecard_params', {}).get('pdo', 20), trainer.config.get('scorecard', {}).get('scorecard_params', {}).get('base_odds', 20), score_bin_width, len(trainer.var_list or [])],
    })
    logistic_prefilter_df = trainer.train_log.get('logistic_tree_prefilter', pd.DataFrame()).copy()
    if trainer.train_log.get('model_type') == 'logistic' and not logistic_prefilter_df.empty:
        model_summary_df = pd.concat([model_summary_df, pd.DataFrame({'项目': ['逻辑回归原始数值特征数', '逻辑回归预筛后候选数', '逻辑回归最终入模变量数'], '取值': [trainer.train_log.get('tree_feature_screening', {}).get('candidate_feature_count', ''), len(logistic_prefilter_df), len(trainer.var_list or [])]})], ignore_index=True)

    feature_importance_df = trainer.train_log.get('feature_importance', pd.DataFrame()).copy()
    if not feature_importance_df.empty:
        feature_importance_df = feature_importance_df.rename(columns={'var': '变量名', 'importance': '特征重要性'})
    else:
        feature_importance_df = pd.DataFrame({'变量名': trainer.var_list or []})
    positive_rank_map = {}
    stability_rank_map = {}
    tree_candidates = trainer.train_log.get('tree_feature_candidates', pd.DataFrame()).copy()
    if not tree_candidates.empty and 'var' in tree_candidates.columns and 'candidate_rank' in tree_candidates.columns:
        positive_rank_map = tree_candidates.set_index('var')['candidate_rank'].to_dict()
    feature_stability_df = trainer.train_log.get('feature_stability', pd.DataFrame()).copy()
    if not feature_stability_df.empty and 'var' in feature_stability_df.columns and 'stability_rank' in feature_stability_df.columns:
        stability_rank_map = feature_stability_df.set_index('var')['stability_rank'].to_dict()
    if '变量名' in feature_importance_df.columns:
        feature_importance_df['candidate_rank'] = feature_importance_df['变量名'].map(positive_rank_map)
        feature_importance_df['正重要性排序'] = feature_importance_df['变量名'].map(positive_rank_map)
        feature_importance_df['稳定性排序'] = feature_importance_df['变量名'].map(stability_rank_map)

    candidate_process_before_df = logistic_prefilter_df.copy() if trainer.train_log.get('model_type') == 'logistic' and not logistic_prefilter_df.empty else tree_candidates.copy()
    if not candidate_process_before_df.empty:
        if 'candidate_rank' not in candidate_process_before_df.columns:
            candidate_process_before_df['candidate_rank'] = range(1, len(candidate_process_before_df) + 1)
        candidate_process_df = candidate_process_before_df.rename(columns={'var': '变量名', 'importance': '特征重要性', 'candidate_rank': '正重要性排序'}).copy()
        candidate_process_df['是否通过稳定性筛选'] = candidate_process_df['变量名'].map(
            feature_stability_df.set_index('var')['stability_pass'].to_dict() if not feature_stability_df.empty and 'var' in feature_stability_df.columns and 'stability_pass' in feature_stability_df.columns else {}
        )
        candidate_process_df['稳定性排序'] = candidate_process_df['变量名'].map(stability_rank_map)
        candidate_process_df['是否最终入模'] = candidate_process_df['变量名'].isin(set(trainer.var_list or []))
    else:
        candidate_process_df = pd.DataFrame(columns=['变量名', '特征重要性', '正重要性排序', '是否通过稳定性筛选', '稳定性排序', '是否最终入模'])
    candidate_sheet_name = '逻辑回归候选特征筛选过程' if trainer.train_log.get('model_type') == 'logistic' and not logistic_prefilter_df.empty else '树模型候选特征筛选过程'
    candidate_summary_df = pd.DataFrame({
        '步骤': [
            '候选变量范围',
            '候选变量过滤',
            '预筛选模型',
            '稳定性筛选',
            '重要性过滤',
            '数量截断',
            '最终入模原因',
            '随机种子',
            '结果说明',
        ],
        '说明': [
            f"原始样本字段共 {trainer.train_log.get('tree_feature_screening', {}).get('raw_column_count', '')} 个；进入树模型筛选前识别出 {trainer.train_log.get('tree_feature_screening', {}).get('candidate_feature_count', '')} 个数值型候选特征。",
            '剔除目标变量、主键、日期字段、显式排除变量，以及无法转成数值型的字段。',
            f"使用 {trainer.train_log.get('model_type', '') if trainer.train_log.get('model_type') != 'logistic' else 'lightgbm/random_forest'} 预筛选模型对全部候选特征先跑一遍重要性。",
            '变量稳定性筛选规则：变量PSI不超过 0.25，且开发样本与验证样本的缺失率差值不超过 0.05。',
            f"仅保留重要性大于 0 的候选特征，共 {len(candidate_process_before_df)} 个。",
            f"在通过稳定性筛选的变量中，再按重要性从高到低排序，结合配置上限 max_tree_features={trainer.config.get('scorecard', {}).get('feature_selection', {}).get('max_tree_features', '')}，最终保留 {len(trainer.var_list or [])} 个入模变量。",
            f"最终入模变量并不是只看特征重要性，而是按“候选字段过滤 -> 正重要性预筛 -> 稳定性筛选 -> 取同时满足稳定性通过且重要性大于0的变量 -> 再按特征重要性排序取前{len(trainer.var_list or [])}个”的顺序选出；报告里保留的是原始排序口径，其中“正重要性排序”是稳定性筛选前的特征重要性排序，“稳定性排序”是按稳定性规则单独形成的排序。",
            f"训练集/测试集随机划分固定 random_state={trainer.config.get('seed', '')}，因此同一份数据在相同配置下可复现。",
            '训练集与测试集都来自同一开发期样本窗口，验证集为时间后置样本；模型选型时同时关注稳定性与过拟合风险。',
        ],
    })

    if not feature_stability_df.empty:
        feature_stability_df = feature_stability_df.rename(columns={'var': '变量名', 'dev_missing_rate': '开发样本缺失率', 'validation_missing_rate': '验证样本缺失率', 'missing_rate_diff': '缺失率差值', 'feature_psi': '变量PSI', 'stability_pass': '是否通过稳定性筛选', 'stability_rank': '稳定性排序'})
    time_window_summary_df = trainer.train_log.get('time_window_validation_summary', pd.DataFrame()).copy()
    if not time_window_summary_df.empty:
        time_window_summary_df = time_window_summary_df.rename(columns={
            'metric_name': '指标名称',
            'metric_value': '指标值',
            'metric': '指标名称',
            'value': '指标值',
            'repeat': '重复轮次',
            'repeat_id': '重复轮次',
            'window_index': '时间窗编号',
            'window_id': '时间窗编号',
            'train_count': '训练样本量',
            'train_sample': '训练样本量',
            'train_sample_validation_sample': '训练样本量/时间窗验证样本量',
            'valid_count': '时间窗验证样本量',
            'validation_sample': '时间窗验证样本量',
            'train_auc': '训练集AUC',
            'auc_train': '训练集AUC',
            'train_ks': '训练集KS',
            'ks_train': '训练集KS',
            'valid_auc': '时间窗验证集AUC',
            'validation_auc': '时间窗验证集AUC',
            'valid_ks': '时间窗验证集KS',
            'validation_ks': '时间窗验证集KS',
            'valid_score': '时间窗验证目标分',
            'validation_score': '时间窗验证目标分',
            'overfit_penalty': '过拟合惩罚',
            'stability_score': '稳定性得分',
        })
        metric_name_map = {
            'validation_auc_mean': '时间窗验证集AUC均值',
            'validation_auc_std': '时间窗验证集AUC标准差',
            'validation_ks_mean': '时间窗验证集KS均值',
            'validation_ks_std': '时间窗验证集KS标准差',
            'stability_score_mean': '稳定性得分均值',
            'stability_score_std': '稳定性得分标准差',
            'train_auc_mean': '训练集AUC均值',
            'train_auc_std': '训练集AUC标准差',
            'train_ks_mean': '训练集KS均值',
            'train_ks_std': '训练集KS标准差',
        }
        if '指标名称' in time_window_summary_df.columns:
            time_window_summary_df['指标名称'] = time_window_summary_df['指标名称'].replace(metric_name_map)
    time_window_detail_headers = [
        '重复轮次',
        '时间窗编号',
        '训练样本量',
        '时间窗验证样本量',
        '训练集AUC',
        '训练集KS',
        '时间窗验证集AUC',
        '时间窗验证集KS',
        '时间窗验证目标分',
        '过拟合惩罚',
        '稳定性得分',
    ]
    tuning_candidates_df = trainer.train_log.get('tuning_candidates', pd.DataFrame()).copy()
    if not tuning_candidates_df.empty:
        tuning_candidates_df = tuning_candidates_df.rename(columns={
            'candidate_rank': '候选排序',
            'trial_number': 'Trial编号',
            'optuna_score': 'Optuna原始得分',
            'selection_score': '稳定性复核得分',
            'mean_valid_score': '平均时间窗验证目标分',
            'std_valid_score': '时间窗验证目标分标准差',
            'mean_valid_auc': '平均时间窗验证集AUC',
            'mean_valid_ks': '平均时间窗验证集KS',
            'mean_train_auc': '平均训练集AUC',
            'mean_train_ks': '平均训练集KS',
            'params': '参数组合',
        })

    workbook_path = output_dir / f'{prefix}_oot_report.xlsx'
    if workbook_path.exists():
        wb = load_workbook(workbook_path)
    else:
        wb = _build_oot_workbook()

    def clear_block(ws, start_row, start_col, end_row, end_col):
        for r in range(start_row, end_row + 1):
            for c in range(start_col, end_col + 1):
                ws.cell(r, c).value = None

    def write_table(ws, start_row, start_col, df):
        for r_idx, row in enumerate(df.itertuples(index=False), start=start_row):
            for c_idx, value in enumerate(row, start=start_col):
                ws.cell(r_idx, c_idx).value = value

    psi_export_df = psi_df.rename(columns={'score_band': 'score', 'baseline_pct': '开发样本占比', 'validation_pct': '验证样本占比', 'psi': 'PSI'})
    if 'score' in psi_export_df.columns:
        psi_export_df['score'] = psi_export_df['score'].astype(str)
    band_export_df = band_report.rename(columns={'score_band': '分数段', 'sample_count': '样本量', 'bad_count': '坏样本量', 'bad_rate': '坏账率', 'avg_score': '平均分', 'sample_pct': '样本占比', 'cum_sample_pct': '累计样本占比', 'cum_bad_count': '累计坏样本量', 'cum_bad_pct': '累计坏样本占比'})
    if '分数段' in band_export_df.columns:
        band_export_df['分数段'] = band_export_df['分数段'].astype(str)
    strategy_export_df = strategy_threshold_df.rename(columns={'score_threshold': '分数阈值', 'approve_rate': '通过率', 'reject_rate': '拒绝率', 'approved_bad_rate': '通过样本坏账率', 'rejected_bad_rate': '拒绝样本坏账率', 'bad_capture_rate': '坏客户拦截率', 'good_pass_rate': '好客户通过率', 'approved_count': '通过样本量', 'rejected_count': '拒绝样本量'})

    for sheet_name in [
        '上线建议说明',
        '上线判定阈值',
        '模型概览',
        '分数分布PSI',
        '验证集分数分段表现',
        '策略阈值分析',
        '入模变量及重要性',
        '变量稳定性筛选',
        '多时间窗验证',
        '候选参数稳定性复核',
        candidate_sheet_name,
        '评估图表',
    ]:
        _ensure_sheet(wb, sheet_name)

    ws = wb['上线建议说明']
    ws.cell(1, 1).value = '项目'
    ws.cell(1, 2).value = '取值'
    clear_block(ws, 2, 1, max(ws.max_row, 40), 2)
    write_table(ws, 2, 1, decision_df)

    ws = wb['上线判定阈值']
    for c_idx, col in enumerate(criteria_df.columns, start=1):
        ws.cell(1, c_idx).value = col
    clear_block(ws, 2, 1, max(ws.max_row, 20), 4)
    write_table(ws, 2, 1, criteria_df)

    ws = wb['模型概览']
    ws.cell(1, 1).value = '项目'
    ws.cell(1, 2).value = '取值'
    clear_block(ws, 2, 1, max(ws.max_row, 30), 2)
    write_table(ws, 2, 1, model_summary_df)

    ws = wb['分数分布PSI']
    for c_idx, col in enumerate(psi_export_df.columns, start=1):
        ws.cell(1, c_idx).value = col
    clear_block(ws, 2, 1, max(ws.max_row, 200), 4)
    write_table(ws, 2, 1, psi_export_df)

    ws = wb['验证集分数分段表现']
    for c_idx, col in enumerate(band_export_df.columns, start=1):
        ws.cell(1, c_idx).value = col
    clear_block(ws, 2, 1, max(ws.max_row, 200), 9)
    write_table(ws, 2, 1, band_export_df)

    ws = wb['策略阈值分析']
    for c_idx, col in enumerate(strategy_export_df.columns, start=1):
        ws.cell(1, c_idx).value = col
    clear_block(ws, 2, 1, max(ws.max_row, 200), 9)
    write_table(ws, 2, 1, strategy_export_df)

    ws = wb['入模变量及重要性']
    for c_idx, col in enumerate(feature_importance_df.columns, start=1):
        ws.cell(1, c_idx).value = col
    clear_block(ws, 2, 1, max(ws.max_row, 1000), 5)
    write_table(ws, 2, 1, feature_importance_df)

    ws = wb['变量稳定性筛选']
    ws.cell(1, 1).value = '说明'
    ws.cell(1, 2).value = '内容'
    clear_block(ws, 2, 1, max(ws.max_row, 3000), 7)
    summary_rows = [
        ['筛选阶段', f"先基于树模型预筛选结果，仅对重要性大于 0 的变量做稳定性评估，共 {len(candidate_process_before_df)} 个变量。"],
        ['筛选阶段', '变量稳定性通过规则：变量PSI <= 0.25，且缺失率差值 <= 0.05。'],
        ['筛选阶段', f"最终通过稳定性筛选的变量数：{int(feature_stability_df['是否通过稳定性筛选'].fillna(False).astype(bool).sum()) if not feature_stability_df.empty else 0} 个。"],
        ['关键观察', '如果很多变量的开发样本缺失率、验证样本缺失率或缺失率差值显示为 0.00%，通常表示这些变量几乎没有缺失，属于正常现象，不是格式错误。'],
        ['关键结论', '建议先看顶部说明和关键结论，再看下方变量明细；“是否通过稳定性筛选”为否的变量，通常不建议优先入模。'],
    ]
    for i, row in enumerate(summary_rows, start=2):
        ws.cell(i, 1).value = row[0]
        ws.cell(i, 2).value = row[1]
    if not feature_stability_df.empty:
        for c_idx, col in enumerate(feature_stability_df.columns, start=1):
            ws.cell(9, c_idx).value = col
        write_table(ws, 10, 1, feature_stability_df)

    ws = wb['多时间窗验证']
    ws.cell(1, 1).value = '说明'
    ws.cell(1, 2).value = '内容'
    clear_block(ws, 2, 1, max(ws.max_row, 3000), 11)
    ws.cell(2, 1).value = '口径说明'
    ws.cell(2, 2).value = '本页展示的是开发期训练样本内部做滚动切窗后的时间窗验证结果，不是最终 OOT 验证集结果。'
    ws.cell(3, 1).value = '口径说明'
    ws.cell(3, 2).value = '因此这里的时间窗验证集AUC/KS通常会高于最终 OOT 报告中的验证集AUC/KS，两者不可直接视为同一口径。'
    if not time_window_summary_df.empty:
        for c_idx, col in enumerate(time_window_summary_df.columns, start=1):
            ws.cell(6, c_idx).value = col
        write_table(ws, 7, 1, time_window_summary_df)
    time_window_detail_df = trainer.train_log.get('time_window_validation_detail', pd.DataFrame()).copy()
    if not time_window_detail_df.empty:
        detail_df = time_window_detail_df.rename(columns={
            'repeat': '重复轮次',
            'repeat_id': '重复轮次',
            'window_index': '时间窗编号',
            'window_id': '时间窗编号',
            'train_count': '训练样本量',
            'train_sample': '训练样本量',
            'train_sample_validation_sample': '训练样本量/时间窗验证样本量',
            'valid_count': '时间窗验证样本量',
            'validation_sample': '时间窗验证样本量',
            'train_auc': '训练集AUC',
            'auc_train': '训练集AUC',
            'train_ks': '训练集KS',
            'ks_train': '训练集KS',
            'valid_auc': '时间窗验证集AUC',
            'validation_auc': '时间窗验证集AUC',
            'valid_ks': '时间窗验证集KS',
            'validation_ks': '时间窗验证集KS',
            'valid_score': '时间窗验证目标分',
            'validation_score': '时间窗验证目标分',
            'overfit_penalty': '过拟合惩罚',
            'stability_score': '稳定性得分',
        })
        detail_df = detail_df.reindex(columns=[col for col in time_window_detail_headers if col in detail_df.columns])
        for c_idx, col in enumerate(time_window_detail_headers, start=1):
            ws.cell(15, c_idx).value = col
        write_table(ws, 16, 1, detail_df)

    ws = wb['候选参数稳定性复核']
    ws.cell(1, 1).value = '说明'
    ws.cell(1, 2).value = '内容'
    clear_block(ws, 2, 1, max(ws.max_row, 3000), 11)
    ws.cell(2, 1).value = '口径说明'
    ws.cell(2, 2).value = '本页的候选参数复核结果来自开发期训练样本内部的滚动时间窗重复验证，用于挑选更稳定的参数组合。'
    ws.cell(3, 1).value = '口径说明'
    ws.cell(3, 2).value = '这里的“平均时间窗验证集AUC/KS”不是最终 OOT 指标；最终上线评估仍以模型概览和验证集分数分段表现中的 OOT 结果为准。'
    if not tuning_candidates_df.empty:
        for c_idx, col in enumerate(tuning_candidates_df.columns, start=1):
            ws.cell(6, c_idx).value = col
        write_table(ws, 7, 1, tuning_candidates_df)

    ws = wb[candidate_sheet_name]
    ws.cell(1, 1).value = '步骤'
    ws.cell(1, 2).value = '说明'
    clear_block(ws, 2, 1, max(ws.max_row, 4000), 6)
    for i, row in enumerate(candidate_summary_df.itertuples(index=False), start=2):
        ws.cell(i, 1).value = row[0]
        ws.cell(i, 2).value = row[1]
    for c_idx, col in enumerate(candidate_process_df.columns, start=1):
        ws.cell(13, c_idx).value = col
    write_table(ws, 14, 1, candidate_process_df)

    ws = wb['评估图表']
    _prepare_eval_chart_sheet(ws)
    next_row = 2
    fig = plt.figure(figsize=(8.8, 6.8))
    _configure_matplotlib()
    for name, color in [('train', 'blue'), ('test', 'green'), ('oot', 'red')]:
        fpr, tpr, _ = roc_curve(y_true_dict[name], y_pred_dict[name])
        auc = roc_auc_score(y_true_dict[name], y_pred_dict[name])
        ks = _calc_ks_value(y_true_dict[name], y_pred_dict[name])
        name_map = {'train': '训练集', 'test': '测试集', 'oot': '验证集'}
        plt.plot(fpr, tpr, color=color, label=f"{name_map[name]}: AUC={auc:.4f}, KS={ks:.4f}")
    plt.plot([0, 1], [0, 1], 'k--', lw=1)
    plt.legend(loc='lower right', fontsize=8)
    plt.title('ROC 曲线')
    roc_file = _safe_image(fig, target_size=CHART_SIZES['roc'])
    next_row = _append_chart_block(ws, next_row, 'ROC 曲线', roc_file)

    if hasattr(trainer.model, 'feature_importances_'):
        top_n = min(len(trainer.var_list or []), 25)
        fig = plt.figure(figsize=(9.2, max(10.5, top_n * 0.42)))
        _configure_matplotlib()
        imp_df = pd.DataFrame({'var': trainer.var_list, 'importance': trainer.model.feature_importances_}).sort_values('importance', ascending=False)
        top_n = min(len(imp_df), 25)
        imp_plot_df = imp_df.head(top_n).sort_values('importance', ascending=True)
        colors = ['#8fb8de'] * max(top_n - 8, 0) + ['#2b6cb0'] * min(8, top_n)
        plt.barh(imp_plot_df['var'], imp_plot_df['importance'], color=colors)
        plt.xlabel('特征分值')
        plt.title(f'特征重要性排序 Top {top_n}')
        plt.grid(True, axis='x', alpha=0.2)
        plt.tight_layout()
        imp_file = _safe_image(fig, target_size=CHART_SIZES['feature_importance'])

    _configure_matplotlib()
    trend_fig, trend_ax = plt.subplots(figsize=(8.8, 6.8))
    trend_labels = ['训练集', '测试集', '验证集']
    trend_auc = [train_auc, test_auc, oot_auc]
    trend_ks = [train_ks, test_ks, oot_ks]
    x = np.arange(len(trend_labels))
    trend_ax.plot(x, trend_auc, marker='o', linewidth=2.8, markersize=8, label='AUC', color='#1f77b4')
    trend_ax.plot(x, trend_ks, marker='s', linewidth=2.8, markersize=7, label='KS', color='#ff7f0e')
    trend_ax.set_title('训练/测试/验证集 AUC-KS 对比')
    trend_ax.set_ylabel('指标值')
    trend_ax.set_xticks(x)
    trend_ax.set_xticklabels(trend_labels)
    trend_ax.set_ylim(0, 1)
    trend_ax.grid(True, axis='y', alpha=0.25)
    trend_ax.axhline(0.65, color='#1f77b4', linestyle='--', linewidth=1, alpha=0.35)
    trend_ax.axhline(0.25, color='#ff7f0e', linestyle='--', linewidth=1, alpha=0.35)
    for idx, value in enumerate(trend_auc):
        trend_ax.text(idx, value + 0.03, f'{value:.3f}', color='#1f77b4', ha='center', va='bottom')
    for idx, value in enumerate(trend_ks):
        trend_ax.text(idx, value - 0.05, f'{value:.3f}', color='#ff7f0e', ha='center', va='top')
    trend_ax.legend(loc='upper right')
    trend_fig.tight_layout()
    trend_file = _safe_image(trend_fig, target_size=CHART_SIZES['trend'])
    next_row = _append_chart_block(ws, next_row, '训练/测试/验证 AUC-KS 对比', trend_file)

    _configure_matplotlib()
    psi_fig = plt.figure(figsize=(8.8, 6.8))
    psi_plot = psi_export_df.copy()
    x = np.arange(len(psi_plot))
    width = 0.36
    plt.bar(x - width / 2, psi_plot['开发样本占比'], width=width, color='#1f77b4', label='基准样本占比')
    plt.bar(x + width / 2, psi_plot['验证样本占比'], width=width, color='#ff7f0e', label='验证集样本占比')
    plt.title('分数分布 PSI')
    plt.ylabel('样本占比')
    plt.xticks(x, psi_plot['score'].astype(str), rotation=35, ha='right')
    plt.grid(True, axis='y', alpha=0.25)
    plt.legend(loc='upper right')
    plt.tight_layout()
    psi_file = _safe_image(psi_fig, target_size=CHART_SIZES['psi'])
    next_row = _append_chart_block(ws, next_row, '分数分布 PSI', psi_file)

    _configure_matplotlib()
    bad_fig, ax1 = plt.subplots(figsize=(8.8, 6.8))
    bad_plot = band_export_df.copy()
    x = np.arange(len(bad_plot))
    bars = ax1.bar(x, bad_plot['样本占比'], color='#9ec3e6', alpha=0.95, label='样本占比')
    ax1.set_ylabel('样本占比')
    ax1.set_xticks(x)
    ax1.set_xticklabels(bad_plot['分数段'].astype(str), rotation=35, ha='right')
    ax1.grid(True, axis='y', alpha=0.25)
    ax2 = ax1.twinx()
    ax2.plot(x, bad_plot['坏账率'], marker='o', linewidth=2.8, markersize=7, color='#d95f02', label='坏账率')
    ax2.set_ylabel('坏账率')
    ax1.set_title('验证集分数段样本占比与坏账率')
    top_idx = int(np.nanargmax(bad_plot['坏账率'].to_numpy(dtype=float))) if len(bad_plot) else 0
    if len(bad_plot):
        ax2.scatter([top_idx], [bad_plot['坏账率'].iloc[top_idx]], color='#b22222', s=60, zorder=5)
        ax2.text(top_idx, bad_plot['坏账率'].iloc[top_idx] + 0.03, f"峰值 {bad_plot['坏账率'].iloc[top_idx]:.1%}", color='#b22222', ha='center')
    lines1, labels1 = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax2.legend(lines1 + lines2, labels1 + labels2, loc='upper right')
    bad_fig.tight_layout()
    bad_file = _safe_image(bad_fig, target_size=CHART_SIZES['bad_rate'])
    next_row = _append_chart_block(ws, next_row, '验证集分数段坏账率', bad_file)

    if hasattr(trainer.model, 'feature_importances_'):
        next_row = _append_chart_block(ws, next_row, '特征重要性排序', imp_file)

    wb.save(workbook_path)
    _style_workbook(workbook_path)
    print(f'OOT report saved: {workbook_path}')


def save_model_artifacts(trainer, model_type, metrics, y_true_dict, y_pred_dict):
    output_dir = Path(trainer.config['output_dir'])
    output_dir.mkdir(parents=True, exist_ok=True)
    prefix = trainer.config['model_prefix']
    model_path = output_dir / f'{prefix}_{model_type}_model.pkl'
    joblib.dump(trainer.model, model_path)
    print(f'模型已保存: {model_path}')
    if model_type == 'logistic' and trainer.woe_binner:
        woe_path = output_dir / f'{prefix}_woe_binner.pkl'
        joblib.dump(trainer.woe_binner, woe_path)
        print(f'WOE分箱器已保存: {woe_path}')
    if model_type == 'logistic':
        export_scorecard_excel(trainer, output_dir, prefix)
    elif model_type in ['lightgbm', 'xgboost']:
        export_pmml(trainer, output_dir, prefix, model_type)
    export_oot_report(trainer, output_dir, prefix, y_true_dict=y_true_dict, y_pred_dict=y_pred_dict)
    print(f'\n所有文件已保存到: {output_dir}')
