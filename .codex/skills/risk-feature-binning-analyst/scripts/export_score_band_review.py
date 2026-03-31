#!/usr/bin/env python3
"""Export score-band review results into the project Excel template."""

from __future__ import annotations

import argparse
import math
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

SETTLESTATUS_NORMALIZATION = {
    4: 2,
    5: 3,
}

STATUS_ORDER = [0, 1, 2, 3]


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Export score-band review workbook from a template.")
    parser.add_argument("data_path", help="Input dataset path: .pkl / .csv / .xlsx / .xls")
    parser.add_argument("--score-column", required=True, help="Score column used to build bands")
    parser.add_argument("--cohort-column", required=True, help="Column distinguishing before/after launch cohorts")
    parser.add_argument("--before-value", required=True, help="Value in cohort column representing not launched / before launch")
    parser.add_argument("--after-value", required=True, help="Value in cohort column representing launched / after launch")
    parser.add_argument("--status-column", default="settlestatus", help="Repayment status column to normalize")
    parser.add_argument("--template-path", default="score-band-improvement-review-template.xlsx", help="Excel template path")
    parser.add_argument("--output-path", default="output/score-band-review.xlsx", help="Workbook output path")
    parser.add_argument("--sheet-name", default="复盘", help="Sheet name to fill")
    parser.add_argument("--band-step", type=int, default=20, help="Score band width")
    parser.add_argument("--band-start", type=int, help="Optional fixed band start")
    parser.add_argument("--band-end", type=int, help="Optional fixed band end")
    parser.add_argument("--note", default="", help="Optional note written to column T row 2")
    return parser


def load_table(path: Path) -> pd.DataFrame:
    suffix = path.suffix.lower()
    if suffix == ".pkl":
        return pd.read_pickle(path)
    if suffix == ".csv":
        return pd.read_csv(path)
    if suffix in {".xlsx", ".xls"}:
        return pd.read_excel(path)
    raise ValueError(f"Unsupported file type: {suffix}")


def normalize_settlestatus(frame: pd.DataFrame, status_column: str) -> pd.DataFrame:
    normalized = frame.copy()
    numeric_status = pd.to_numeric(normalized[status_column], errors="coerce")
    normalized[status_column] = numeric_status.replace(SETTLESTATUS_NORMALIZATION)
    normalized[f"{status_column}_normalized"] = normalized[status_column]
    return normalized


def resolve_band_limits(series: pd.Series, step: int, band_start: int | None, band_end: int | None) -> tuple[int, int]:
    valid = pd.to_numeric(series, errors="coerce").dropna()
    if valid.empty:
        raise ValueError("No valid score values found.")
    start = band_start if band_start is not None else int(math.floor(valid.min() / step) * step)
    end = band_end if band_end is not None else int(math.ceil((valid.max() + 1) / step) * step - 1)
    if end < start:
        raise ValueError("band_end must be greater than or equal to band_start.")
    return start, end


def build_band_labels(start: int, end: int, step: int) -> list[tuple[str, int, int]]:
    labels = []
    current = start
    while current <= end:
        upper = min(current + step - 1, end)
        labels.append((f"{current}-{upper}", current, upper))
        current += step
    return labels


def assign_score_band(series: pd.Series, labels: list[tuple[str, int, int]]) -> pd.Series:
    numeric = pd.to_numeric(series, errors="coerce")

    def locate(value: float) -> str | None:
        if pd.isna(value):
            return None
        for label, lower, upper in labels:
            if lower <= value <= upper:
                return label
        return None

    return numeric.apply(locate)


def summarize_block(frame: pd.DataFrame, cohort_column: str, cohort_value: str, score_column: str, status_column: str, labels: list[tuple[str, int, int]]) -> pd.DataFrame:
    subset = frame.loc[frame[cohort_column].astype(str) == str(cohort_value)].copy()
    subset["score_band"] = assign_score_band(subset[score_column], labels)
    subset = subset.dropna(subset=["score_band", status_column])
    subset[status_column] = pd.to_numeric(subset[status_column], errors="coerce").astype("Int64")
    subset = subset.loc[subset[status_column].isin(STATUS_ORDER)]

    rows: list[dict] = []
    total_count = len(subset)
    for label, _, _ in labels:
        band_df = subset.loc[subset["score_band"] == label]
        status_counts = band_df[status_column].value_counts().to_dict()
        total = int(len(band_df))
        bad_first = int(status_counts.get(0, 0) + status_counts.get(3, 0))
        bad_current = int(status_counts.get(0, 0))
        rows.append(
            {
                "label": label,
                0: int(status_counts.get(0, 0)),
                1: int(status_counts.get(1, 0)),
                2: int(status_counts.get(2, 0)),
                3: int(status_counts.get(3, 0)),
                "total": total,
                "first_overdue_rate": (bad_first / total) if total else 0,
                "current_overdue_rate": (bad_current / total) if total else 0,
                "share": (total / total_count) if total_count else 0,
            }
        )

    summary = pd.DataFrame(rows)
    total_row = {
        "label": "总计",
        0: int(summary[0].sum()),
        1: int(summary[1].sum()),
        2: int(summary[2].sum()),
        3: int(summary[3].sum()),
        "total": int(summary["total"].sum()),
    }
    total_row["first_overdue_rate"] = ((total_row[0] + total_row[3]) / total_row["total"]) if total_row["total"] else 0
    total_row["current_overdue_rate"] = (total_row[0] / total_row["total"]) if total_row["total"] else 0
    total_row["share"] = 1 if total_row["total"] else 0
    return pd.concat([summary, pd.DataFrame([total_row])], ignore_index=True)


def clear_block(ws, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            ws.cell(row, col).value = None


def write_block(ws, df: pd.DataFrame, label_col: int, status_start_col: int, total_col: int, first_rate_col: int, current_rate_col: int, share_col: int, start_row: int) -> None:
    for offset, row in enumerate(df.itertuples(index=False), start=0):
        excel_row = start_row + offset
        ws.cell(excel_row, label_col).value = row.label
        ws.cell(excel_row, status_start_col).value = int(getattr(row, "_1"))
        ws.cell(excel_row, status_start_col + 1).value = int(getattr(row, "_2"))
        ws.cell(excel_row, status_start_col + 2).value = int(getattr(row, "_3"))
        ws.cell(excel_row, status_start_col + 3).value = int(getattr(row, "_4"))
        ws.cell(excel_row, total_col).value = int(row.total)
        ws.cell(excel_row, first_rate_col).value = float(row.first_overdue_rate)
        ws.cell(excel_row, current_rate_col).value = float(row.current_overdue_rate)
        ws.cell(excel_row, share_col).value = float(row.share)


def ensure_template_rows(ws, required_rows: int, start_row: int = 3, total_row: int = 18) -> None:
    template_capacity = total_row - start_row
    data_rows = required_rows - 1
    delta = data_rows - template_capacity
    if delta > 0:
        ws.insert_rows(total_row, amount=delta)
    elif delta < 0:
        ws.delete_rows(total_row + delta, amount=-delta)


def main() -> int:
    args = build_parser().parse_args()
    data_path = Path(args.data_path)
    template_path = Path(args.template_path)
    output_path = Path(args.output_path)

    df = load_table(data_path)
    required_cols = [args.score_column, args.cohort_column, args.status_column]
    missing = [col for col in required_cols if col not in df.columns]
    if missing:
        raise SystemExit(f"Missing required columns: {', '.join(missing)}")

    df = normalize_settlestatus(df, args.status_column)
    start, end = resolve_band_limits(df[args.score_column], args.band_step, args.band_start, args.band_end)
    labels = build_band_labels(start, end, args.band_step)

    before_df = summarize_block(df, args.cohort_column, args.before_value, args.score_column, args.status_column, labels)
    after_df = summarize_block(df, args.cohort_column, args.after_value, args.score_column, args.status_column, labels)

    wb = load_workbook(template_path)
    ws = wb[args.sheet_name]

    ensure_template_rows(ws, max(len(before_df), len(after_df)))
    clear_block(ws, 3, ws.max_row, 1, 9)
    clear_block(ws, 3, ws.max_row, 11, 19)

    ws["A1"] = str(args.before_value)
    ws["K1"] = str(args.after_value)
    ws["A2"] = args.score_column
    ws["K2"] = args.score_column
    if args.note:
        ws["T2"] = args.note

    write_block(ws, before_df, 1, 2, 6, 7, 8, 9, 3)
    write_block(ws, after_df, 11, 12, 16, 17, 18, 19, 3)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    print(f"Template exported: {output_path}")
    print(f"Score bands: {start}-{end}, step={args.band_step}")
    print(f"Before rows: {len(before_df) - 1}, After rows: {len(after_df) - 1}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
