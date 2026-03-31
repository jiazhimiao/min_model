#!/usr/bin/env python3
"""Export multiple score and feature tables using precomputed analysis results."""

from __future__ import annotations

import argparse
import math
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

SETTLESTATUS_NORMALIZATION = {
    4: 2,
    5: 3,
}

STATUS_ORDER = [0, 1, 2, 3]


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Export template tables from analysis result files.")
    parser.add_argument("data_path", help="Input dataset path")
    parser.add_argument("--score-candidates", required=True, help="Path to score_candidates.csv")
    parser.add_argument("--feature-candidates", required=True, help="Path to feature_candidates.csv")
    parser.add_argument("--template-path", default="score-band-improvement-review-template.xlsx", help="Excel template path")
    parser.add_argument("--output-path", default="output/multi_score_feature_from_analysis.xlsx", help="Workbook output path")
    parser.add_argument("--status-column", default="settleStatus", help="Repayment status column")
    parser.add_argument("--sample-rows", type=int, default=1500, help="Number of rows to use")
    parser.add_argument("--max-score-tables", type=int, default=6, help="Maximum number of score tables")
    parser.add_argument("--max-feature-tables", type=int, default=12, help="Maximum number of feature tables")
    parser.add_argument("--feature-max-distinct", type=int, default=12, help="Maximum distinct count for feature output")
    parser.add_argument("--score-sheet", default="评分", help="Sheet name for score output")
    parser.add_argument("--feature-sheet", default="特征", help="Sheet name for feature output")
    return parser


def load_data(path: Path, sample_rows: int) -> pd.DataFrame:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return pd.read_csv(path, encoding="utf-8-sig", low_memory=False, nrows=sample_rows)
    if suffix == ".pkl":
        return pd.read_pickle(path).head(sample_rows)
    if suffix in {".xlsx", ".xls"}:
        return pd.read_excel(path).head(sample_rows)
    raise ValueError(f"Unsupported file type: {suffix}")


def normalize_status(frame: pd.DataFrame, status_column: str) -> pd.DataFrame:
    normalized = frame.copy()
    normalized[status_column] = pd.to_numeric(normalized[status_column], errors="coerce").replace(SETTLESTATUS_NORMALIZATION)
    return normalized


def resolve_score_bands(series: pd.Series, step: int = 20) -> list[tuple[str, int, int]]:
    numeric = pd.to_numeric(series, errors="coerce").dropna()
    if numeric.empty:
        return []
    start = int(math.floor(numeric.min() / step) * step)
    end = int(math.ceil((numeric.max() + 1) / step) * step - 1)
    bands = []
    current = start
    while current <= end:
        upper = min(current + step - 1, end)
        bands.append((f"{current}-{upper}", current, upper))
        current += step
    return bands


def assign_score_band(series: pd.Series, bands: list[tuple[str, int, int]]) -> pd.Series:
    numeric = pd.to_numeric(series, errors="coerce")

    def locate(value: float) -> str | None:
        if pd.isna(value):
            return None
        for label, lower, upper in bands:
            if lower <= value <= upper:
                return label
        return None

    return numeric.apply(locate)


def format_feature_label(value: object) -> str | None:
    if pd.isna(value):
        return None
    numeric = pd.to_numeric(pd.Series([value]), errors="coerce").iloc[0]
    if pd.notna(numeric):
        if float(numeric).is_integer():
            integer = int(numeric)
            return f"{integer}-{integer}"
        return str(numeric)
    return str(value)


def summarize(frame: pd.DataFrame, label_series: pd.Series, labels: list[str], status_column: str) -> pd.DataFrame:
    subset = frame.copy()
    subset["label"] = label_series
    subset = subset.dropna(subset=["label", status_column])
    subset[status_column] = pd.to_numeric(subset[status_column], errors="coerce").astype("Int64")
    subset = subset.loc[subset[status_column].isin(STATUS_ORDER)]

    rows = []
    total_count = len(subset)
    for label in labels:
        bucket = subset.loc[subset["label"] == label]
        status_counts = bucket[status_column].value_counts().to_dict()
        total = int(len(bucket))
        rows.append(
            {
                "label": label,
                "status_0": int(status_counts.get(0, 0)),
                "status_1": int(status_counts.get(1, 0)),
                "status_2": int(status_counts.get(2, 0)),
                "status_3": int(status_counts.get(3, 0)),
                "total": total,
                "first_overdue_rate": ((status_counts.get(0, 0) + status_counts.get(3, 0)) / total) if total else 0,
                "current_overdue_rate": (status_counts.get(0, 0) / total) if total else 0,
                "share": (total / total_count) if total_count else 0,
            }
        )

    summary = pd.DataFrame(rows)
    total_row = {
        "label": "总计",
        "status_0": int(summary["status_0"].sum()) if not summary.empty else 0,
        "status_1": int(summary["status_1"].sum()) if not summary.empty else 0,
        "status_2": int(summary["status_2"].sum()) if not summary.empty else 0,
        "status_3": int(summary["status_3"].sum()) if not summary.empty else 0,
        "total": int(summary["total"].sum()) if not summary.empty else 0,
    }
    total = total_row["total"]
    total_row["first_overdue_rate"] = ((total_row["status_0"] + total_row["status_3"]) / total) if total else 0
    total_row["current_overdue_rate"] = (total_row["status_0"] / total) if total else 0
    total_row["share"] = 1 if total else 0
    return pd.concat([summary, pd.DataFrame([total_row])], ignore_index=True)


def choose_scores(score_df: pd.DataFrame, max_count: int) -> list[tuple[str, str]]:
    filtered = score_df.loc[
        score_df["recommendation"].isin(["priority_review", "secondary_review", "score_monitoring"])
        & (score_df["constant_flag"] == False)
        & (score_df["valid_rows"] >= 100)
    ].copy()
    filtered = filtered.sort_values(["iv", "monotonic", "distinct_count"], ascending=[False, False, False])
    return list(filtered[["feature", "meaning"]].head(max_count).itertuples(index=False, name=None))


def choose_features(feature_df: pd.DataFrame, max_count: int, max_distinct: int) -> list[tuple[str, str]]:
    filtered = feature_df.loc[
        (feature_df["constant_flag"] == False)
        & (feature_df["valid_rows"] >= 100)
        & (feature_df["distinct_count"] >= 2)
        & (feature_df["distinct_count"] <= max_distinct)
        & (feature_df["missing_rate"] < 0.95)
    ].copy()
    filtered["meaning_known"] = filtered["meaning"].ne("未在特征码汇总中找到")
    filtered = filtered.sort_values(
        ["meaning_known", "distinct_count", "missing_rate", "feature"],
        ascending=[False, True, True, True],
    )
    return list(filtered[["feature", "meaning"]].head(max_count).itertuples(index=False, name=None))


def write_block(ws, start_row: int, title: str, table: pd.DataFrame) -> int:
    ws.cell(start_row, 1).value = title
    ws.cell(start_row, 2).value = 0
    ws.cell(start_row, 3).value = 1
    ws.cell(start_row, 4).value = 2
    ws.cell(start_row, 5).value = 3
    ws.cell(start_row, 6).value = "总计"
    ws.cell(start_row, 7).value = "首逾率"
    ws.cell(start_row, 8).value = "在逾率"
    ws.cell(start_row, 9).value = "占比"
    row_no = start_row + 1
    for row in table.itertuples(index=False):
        ws.cell(row_no, 1).value = row.label
        ws.cell(row_no, 2).value = int(row.status_0)
        ws.cell(row_no, 3).value = int(row.status_1)
        ws.cell(row_no, 4).value = int(row.status_2)
        ws.cell(row_no, 5).value = int(row.status_3)
        ws.cell(row_no, 6).value = int(row.total)
        ws.cell(row_no, 7).value = float(row.first_overdue_rate)
        ws.cell(row_no, 8).value = float(row.current_overdue_rate)
        ws.cell(row_no, 9).value = float(row.share)
        row_no += 1
    return row_no + 1


def clear_sheet(ws) -> None:
    for row in range(1, ws.max_row + 1):
        for col in range(1, 10):
            ws.cell(row, col).value = None


def main() -> int:
    args = build_parser().parse_args()

    data = normalize_status(load_data(Path(args.data_path), args.sample_rows), args.status_column)
    score_df = pd.read_csv(args.score_candidates, encoding="utf-8-sig")
    feature_df = pd.read_csv(args.feature_candidates, encoding="utf-8-sig")

    score_choices = choose_scores(score_df, args.max_score_tables)
    feature_choices = choose_features(feature_df, args.max_feature_tables, args.feature_max_distinct)

    wb = load_workbook(args.template_path)
    score_ws = wb[args.score_sheet]
    feature_ws = wb[args.feature_sheet]
    clear_sheet(score_ws)
    clear_sheet(feature_ws)

    score_row = 1
    for score_col, meaning in score_choices:
        bands = resolve_score_bands(data[score_col], step=20)
        labels = [label for label, _, _ in bands]
        label_series = assign_score_band(data[score_col], bands)
        table = summarize(data, label_series, labels, args.status_column)
        title = f"{score_col} | {meaning}"
        score_row = write_block(score_ws, score_row, title, table)

    feature_row = 1
    for feature_col, meaning in feature_choices:
        numeric = pd.to_numeric(data[feature_col], errors="coerce")
        if numeric.notna().all():
            labels = [format_feature_label(v) for v in sorted(numeric.dropna().unique())]
        else:
            labels = sorted({str(v) for v in data[feature_col].dropna().unique()})
        label_series = data[feature_col].apply(format_feature_label if numeric.notna().all() else lambda v: None if pd.isna(v) else str(v))
        table = summarize(data, label_series, labels, args.status_column)
        title = f"{feature_col} | {meaning}"
        feature_row = write_block(feature_ws, feature_row, title, table)

    output_path = Path(args.output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    print(f"Sample rows used: {len(data)}")
    print(f"Score tables written: {len(score_choices)}")
    print(f"Feature tables written: {len(feature_choices)}")
    print(f"Template exported: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
