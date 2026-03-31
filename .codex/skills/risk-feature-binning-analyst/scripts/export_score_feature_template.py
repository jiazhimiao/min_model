#!/usr/bin/env python3
"""Export score and feature review results into the two-sheet Excel template."""

from __future__ import annotations

import argparse
import math
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

SETTLESTATUS_NORMALIZATION = {
    4: 2,
    5: 3,
}

STATUS_ORDER = [0, 1, 2, 3]


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Export review results into the score/feature Excel template.")
    parser.add_argument("data_path", help="Input dataset path: .pkl / .csv / .xlsx / .xls")
    parser.add_argument("--status-column", default="settlestatus", help="Repayment status column to normalize")
    parser.add_argument("--template-path", default="score-band-improvement-review-template.xlsx", help="Excel template path")
    parser.add_argument("--output-path", default="output/score_feature_review.xlsx", help="Workbook output path")
    parser.add_argument("--score-column", help="Score column used to fill the score sheet")
    parser.add_argument("--score-sheet", default="评分", help="Sheet name for score output")
    parser.add_argument("--score-band-step", type=int, default=20, help="Score band width")
    parser.add_argument("--score-band-start", type=int, help="Optional fixed score band start")
    parser.add_argument("--score-band-end", type=int, help="Optional fixed score band end")
    parser.add_argument("--feature-column", help="Feature column used to fill the feature sheet")
    parser.add_argument("--feature-sheet", default="特征", help="Sheet name for feature output")
    parser.add_argument("--feature-sort", choices=["auto", "numeric", "string"], default="auto", help="Feature label sorting mode")
    return parser


def load_table(path: Path) -> pd.DataFrame:
    suffix = path.suffix.lower()
    if suffix == ".pkl":
        return pd.read_pickle(path)
    if suffix == ".csv":
        return pd.read_csv(path)
    if suffix in {".xlsx", ".xls"}:
        return pd.read_excel(path)
    raise ValueError(f"Unsupported file type: {suffix}")


def normalize_settlestatus(frame: pd.DataFrame, status_column: str) -> pd.DataFrame:
    normalized = frame.copy()
    numeric_status = pd.to_numeric(normalized[status_column], errors="coerce")
    normalized[status_column] = numeric_status.replace(SETTLESTATUS_NORMALIZATION)
    normalized[f"{status_column}_normalized"] = normalized[status_column]
    return normalized


def resolve_band_limits(series: pd.Series, step: int, band_start: int | None, band_end: int | None) -> tuple[int, int]:
    valid = pd.to_numeric(series, errors="coerce").dropna()
    if valid.empty:
        raise ValueError("No valid score values found.")
    start = band_start if band_start is not None else int(math.floor(valid.min() / step) * step)
    end = band_end if band_end is not None else int(math.ceil((valid.max() + 1) / step) * step - 1)
    if end < start:
        raise ValueError("score-band-end must be greater than or equal to score-band-start.")
    return start, end


def build_score_labels(start: int, end: int, step: int) -> list[tuple[str, int, int]]:
    labels = []
    current = start
    while current <= end:
        upper = min(current + step - 1, end)
        labels.append((f"{current}-{upper}", current, upper))
        current += step
    return labels


def assign_score_band(series: pd.Series, labels: list[tuple[str, int, int]]) -> pd.Series:
    numeric = pd.to_numeric(series, errors="coerce")

    def locate(value: float) -> str | None:
        if pd.isna(value):
            return None
        for label, lower, upper in labels:
            if lower <= value <= upper:
                return label
        return None

    return numeric.apply(locate)


def format_feature_value(value: float | int | str) -> str:
    if isinstance(value, str):
        return value
    if float(value).is_integer():
        int_value = int(value)
        return f"{int_value}-{int_value}"
    return str(value)


def infer_feature_labels(series: pd.Series, sort_mode: str) -> list[str]:
    valid = series.dropna()
    if valid.empty:
        return []
    if sort_mode == "numeric":
        numeric = pd.to_numeric(valid, errors="raise")
        return [format_feature_value(value) for value in sorted(pd.unique(numeric))]
    if sort_mode == "string":
        return sorted({str(value) for value in valid})
    numeric = pd.to_numeric(valid, errors="coerce")
    if numeric.notna().all():
        return [format_feature_value(value) for value in sorted(pd.unique(numeric))]
    return sorted({str(value) for value in valid})


def map_feature_label(value: object) -> str | None:
    if pd.isna(value):
        return None
    numeric = pd.to_numeric(pd.Series([value]), errors="coerce").iloc[0]
    if pd.notna(numeric):
        return format_feature_value(numeric)
    return str(value)


def summarize_rows(frame: pd.DataFrame, label_series: pd.Series, labels: list[str], status_column: str) -> pd.DataFrame:
    subset = frame.copy()
    subset["review_label"] = label_series
    subset = subset.dropna(subset=["review_label", status_column])
    subset[status_column] = pd.to_numeric(subset[status_column], errors="coerce").astype("Int64")
    subset = subset.loc[subset[status_column].isin(STATUS_ORDER)]

    rows: list[dict] = []
    total_count = len(subset)
    for label in labels:
        bucket = subset.loc[subset["review_label"] == label]
        status_counts = bucket[status_column].value_counts().to_dict()
        total = int(len(bucket))
        first_overdue = int(status_counts.get(0, 0) + status_counts.get(3, 0))
        current_overdue = int(status_counts.get(0, 0))
        rows.append(
            {
                "label": label,
                "status_0": int(status_counts.get(0, 0)),
                "status_1": int(status_counts.get(1, 0)),
                "status_2": int(status_counts.get(2, 0)),
                "status_3": int(status_counts.get(3, 0)),
                "total": total,
                "first_overdue_rate": (first_overdue / total) if total else 0,
                "current_overdue_rate": (current_overdue / total) if total else 0,
                "share": (total / total_count) if total_count else 0,
            }
        )

    summary = pd.DataFrame(rows)
    total_row = {
        "label": "总计",
        "status_0": int(summary["status_0"].sum()) if not summary.empty else 0,
        "status_1": int(summary["status_1"].sum()) if not summary.empty else 0,
        "status_2": int(summary["status_2"].sum()) if not summary.empty else 0,
        "status_3": int(summary["status_3"].sum()) if not summary.empty else 0,
        "total": int(summary["total"].sum()) if not summary.empty else 0,
    }
    total = total_row["total"]
    total_row["first_overdue_rate"] = ((total_row["status_0"] + total_row["status_3"]) / total) if total else 0
    total_row["current_overdue_rate"] = (total_row["status_0"] / total) if total else 0
    total_row["share"] = 1 if total else 0
    return pd.concat([summary, pd.DataFrame([total_row])], ignore_index=True)


def clear_block(ws, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            ws.cell(row, col).value = None


def ensure_template_rows(ws, required_rows: int, start_row: int = 2, total_row: int = 11) -> None:
    template_capacity = total_row - start_row
    data_rows = required_rows - 1
    delta = data_rows - template_capacity
    if delta > 0:
        ws.insert_rows(total_row, amount=delta)
    elif delta < 0:
        ws.delete_rows(total_row + delta, amount=-delta)


def write_single_block(ws, title: str, df: pd.DataFrame) -> None:
    ensure_template_rows(ws, len(df))
    clear_block(ws, 1, ws.max_row, 1, 9)
    ws["A1"] = title
    ws["B1"] = 0
    ws["C1"] = 1
    ws["D1"] = 2
    ws["E1"] = 3
    ws["F1"] = "总计"
    ws["G1"] = "首逾率"
    ws["H1"] = "在逾率"
    ws["I1"] = "占比"

    for offset, row in enumerate(df.itertuples(index=False), start=2):
        ws.cell(offset, 1).value = row.label
        ws.cell(offset, 2).value = int(row.status_0)
        ws.cell(offset, 3).value = int(row.status_1)
        ws.cell(offset, 4).value = int(row.status_2)
        ws.cell(offset, 5).value = int(row.status_3)
        ws.cell(offset, 6).value = int(row.total)
        ws.cell(offset, 7).value = float(row.first_overdue_rate)
        ws.cell(offset, 8).value = float(row.current_overdue_rate)
        ws.cell(offset, 9).value = float(row.share)


def build_score_summary(df: pd.DataFrame, score_column: str, status_column: str, step: int, start: int | None, end: int | None) -> pd.DataFrame:
    band_start, band_end = resolve_band_limits(df[score_column], step, start, end)
    score_labels = build_score_labels(band_start, band_end, step)
    label_series = assign_score_band(df[score_column], score_labels)
    return summarize_rows(df, label_series, [label for label, _, _ in score_labels], status_column)


def build_feature_summary(df: pd.DataFrame, feature_column: str, status_column: str, sort_mode: str) -> pd.DataFrame:
    feature_labels = infer_feature_labels(df[feature_column], sort_mode)
    label_series = df[feature_column].apply(map_feature_label)
    return summarize_rows(df, label_series, feature_labels, status_column)


def main() -> int:
    args = build_parser().parse_args()
    if not args.score_column and not args.feature_column:
        raise SystemExit("Provide at least one of --score-column or --feature-column.")

    data_path = Path(args.data_path)
    template_path = Path(args.template_path)
    output_path = Path(args.output_path)

    df = load_table(data_path)
    if args.status_column not in df.columns:
        raise SystemExit(f"Missing required column: {args.status_column}")
    df = normalize_settlestatus(df, args.status_column)

    wb = load_workbook(template_path)

    if args.score_column:
        if args.score_column not in df.columns:
            raise SystemExit(f"Missing score column: {args.score_column}")
        score_summary = build_score_summary(
            df=df,
            score_column=args.score_column,
            status_column=args.status_column,
            step=args.score_band_step,
            start=args.score_band_start,
            end=args.score_band_end,
        )
        write_single_block(wb[args.score_sheet], args.score_column, score_summary)
        print(f"Filled score sheet: {args.score_sheet}")

    if args.feature_column:
        if args.feature_column not in df.columns:
            raise SystemExit(f"Missing feature column: {args.feature_column}")
        feature_summary = build_feature_summary(
            df=df,
            feature_column=args.feature_column,
            status_column=args.status_column,
            sort_mode=args.feature_sort,
        )
        write_single_block(wb[args.feature_sheet], args.feature_column, feature_summary)
        print(f"Filled feature sheet: {args.feature_sheet}")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Template exported: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
